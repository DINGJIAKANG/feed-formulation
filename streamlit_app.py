"""
饲料配方优化系统 — Streamlit 图形化界面
========================================
基于线性规划的最低成本饲料配方求解工具。

启动方式：
    streamlit run streamlit_app.py
    或双击 run_gui.bat
"""

import sys
import os
from io import BytesIO

import pandas as pd
import numpy as np
import streamlit as st

# ── 路径设置 ──────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from feed_formulation import (  # noqa: E402
    load_ingredients,
    load_standards,
    load_formula_requirements,
    build_and_solve,
    filter_ingredients,
    convert_to_dm_basis,
    convert_result_to_asfed,  # 新增：DM转as-fed
    convert_requirements_to_dm,  # 新增：标准as-fed转DM
    save_result_to_excel,
    validate_constraints,
    load_feed_type_classification,
    _get_feed_type_indices,
    NUTRIENT_MAP,
    ENERGY_MAP,
)

from recipe_adjuster import (  # noqa: E402
    calc_concentrate_forage,
    calc_ca_p_ratio,
    _strip_tag,
)

# ── 文件路径 ──────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INGREDIENTS_FILE = os.path.join(BASE_DIR, "饲料常规营养含量.xls")
STANDARDS_FILE = os.path.join(BASE_DIR, "饲养标准.xlsx")

# ── 配方微调弹窗 ──────────────────────────────────────────
@st.dialog("🔧 配方微调", width="large")
def show_adjustment_dialog(result, all_ingredients):
    """配方微调弹窗：表格编辑式，实时查看各原料营养贡献和汇总对比。"""
    from recipe_adjuster import (
        calc_nutrients, calc_cost,
        build_adjustment_editable_df, calc_summary_rows,
    )

    st.caption("修改各原料的微调配比%，实时查看营养变化。")

    original_pct = result.get("ingredients", {})
    if not original_pct:
        st.info("无配方数据可供微调。")
        return

    # 获取求解时的原料数据和营养映射
    _adj_ing_df = result.get("_working_ingredients_df", all_ingredients)
    _adj_nutrient_map = result.get("nutrient_map", None)
    _requirements = result.get("requirements", {})

    # ── 初始化：仅在首次打开时 ─────
    _init_key = "_adj_initialized"
    if not st.session_state.get(_init_key):
        # 获取营养列名（用于显示表格）
        _, _nut_cols = build_adjustment_editable_df(
            original_pct=original_pct,
            current_adj_pct=dict(original_pct),
            ingredients_df=_adj_ing_df,
            requirements=_requirements,
            nutrient_map=_adj_nutrient_map,
        )
        st.session_state._adj_nut_cols = _nut_cols
        st.session_state._adj_ing_df = _adj_ing_df
        st.session_state._adj_requirements = _requirements
        st.session_state._adj_original_pct = dict(original_pct)
        st.session_state._adj_number_values = dict(original_pct)
        st.session_state[_init_key] = True

    # 从 session_state 读取（后续 rerun 直接使用）
    _nut_cols = st.session_state._adj_nut_cols
    _adj_ing_df_cached = st.session_state._adj_ing_df
    _requirements_cached = st.session_state._adj_requirements
    _adj_nutrient_map_cached = result.get("nutrient_map", None)
    _original_pct = st.session_state._adj_original_pct

    # ── 微调配比输入（逐行数字输入，避免 data_editor 状态问题）───
    st.markdown("**微调配比调整**")
    st.caption("修改各原料的微调配比%，实时查看营养变化。")

    # ── 添加新原料功能（放在微调区域顶部）──────────
    _existing_names_for_add = set(st.session_state.get("_adj_number_values", {}).keys())
    _adj_ing_df_for_add = st.session_state.get("_adj_ing_df", _adj_ing_df)
    _pool_candidates = []
    if hasattr(_adj_ing_df_for_add, 'index'):
        # 构建上传库和默认库名称集合，用于判断来源并加标签
        _upload_names_set = set()
        _uploaded_df_src = st.session_state.get("uploaded_ingredients_df")
        if _uploaded_df_src is not None:
            _upload_names_set = set(_uploaded_df_src.index)
        for _pname in _adj_ing_df_for_add.index:
            _real_pname = _strip_tag(str(_pname))
            _already_in = False
            for ename in _existing_names_for_add:
                if _strip_tag(ename) == _real_pname:
                    _already_in = True
                    break
            if not _already_in:
                # 根据来源添加 [上传]/[默认] 标签（与主界面手动选择一致）
                if _real_pname in _upload_names_set:
                    _pool_candidates.append(f"[上传] {_real_pname}")
                else:
                    _pool_candidates.append(f"[默认] {_real_pname}")

    if _pool_candidates:
        with st.expander("➕ 添加新原料到微调表"):
            _ing_search = st.text_input(
                "🔍 搜索原料（输入关键词过滤）",
                key="adj_new_ing_search",
                placeholder="如：玉米、豆粕、鱼粉...",
            )
            _sorted_pool = sorted(_pool_candidates)
            if _ing_search.strip():
                _kw = _ing_search.strip().lower()
                _filtered_pool = [x for x in _sorted_pool if _kw in x.lower()]
            else:
                _filtered_pool = _sorted_pool[:50]
            _finfo = f"（匹配 {len(_filtered_pool)} / 共 {len(_sorted_pool)} 种）"
            acol1, acol2, acol3 = st.columns([2, 1, 1])
            with acol1:
                _new_ing = st.selectbox(
                    f"选择原料 {_finfo}",
                    options=_filtered_pool,
                    key="adj_new_ingredient_sel",
                    placeholder="从列表中选择...",
                )
            with acol2:
                _new_val = st.number_input(
                    "初始配比%",
                    min_value=0.0, max_value=100.0,
                    value=0.0, step=0.5,
                    key="adj_new_ing_value",
                    format="%.2f",
                )
            with acol3:
                st.write("")
                if st.button("添加", key="adj_add_ing_btn", use_container_width=True):
                    if _new_ing and _new_val > 0:
                        st.session_state._adj_number_values[str(_new_ing)] = float(_new_val)
                        # 设置标志位让主页面自动重新打开dialog（避免st.rerun导致dialog关闭）
                        st.session_state._should_reopen_adj_dialog = True
                        st.rerun()

    # 初始化微调值在 session_state
    if "_adj_number_values" not in st.session_state:
        st.session_state._adj_number_values = dict(original_pct)

    adj_pct = {}
    # 显示顺序：按当前微调配比降序，包含所有已添加的原料
    _all_ings = list(st.session_state._adj_number_values.keys())

    # 用多列布局（每行2个原料）节省空间
    n_ui_cols = 2
    _rows = [_all_ings[i:i+n_ui_cols] for i in range(0, len(_all_ings), n_ui_cols)]
    for _row_ings in _rows:
        _cols = st.columns(n_ui_cols)
        for _j, _ing in enumerate(_row_ings):
            with _cols[_j]:
                _orig = _original_pct.get(_ing, 0.0)
                _val = st.session_state._adj_number_values.get(_ing, _orig)
                _new = st.number_input(
                    label=_ing,
                    min_value=0.0,
                    max_value=100.0,
                    value=_val,
                    step=0.1,
                    format="%.2f",
                    key=f"adj_num_{_ing}",
                )
                adj_pct[_ing] = _new

    # 同步到 session_state
    st.session_state._adj_number_values = adj_pct

    # ── 归一化配比（用于营养计算，确保合计=100%）─────
    _original_pct = st.session_state.get("_adj_original_pct", original_pct)
    _adj_ing_df_cached = st.session_state.get("_adj_ing_df", _adj_ing_df)
    _requirements_cached = st.session_state.get("_adj_requirements", _requirements)
    _adj_nutrient_map_cached = result.get("nutrient_map", None)

    _adj_total = sum(adj_pct.values())
    _orig_total = sum(_original_pct.values())
    if abs(_adj_total - 100.0) > 0.01 and _adj_total > 0:
        _adj_pct_normalized = {k: v / _adj_total * 100.0 for k, v in adj_pct.items()}
    else:
        _adj_pct_normalized = adj_pct

    # ── 显示只读营养贡献表格 ────────────────────
    st.markdown("---")
    st.markdown("**配方明细（营养贡献）**")
    _display_df, _nut_cols = build_adjustment_editable_df(
        original_pct=original_pct,
        current_adj_pct=_adj_pct_normalized,
        ingredients_df=_adj_ing_df_cached,
        requirements=_requirements_cached,
        nutrient_map=_adj_nutrient_map_cached,
    )
    # 只显示，不编辑
    st.dataframe(_display_df, use_container_width=True, hide_index=True)

    # ── 汇总指标（顶部卡片）───

    _orig_cost = result.get("cost", 0)
    _adj_cost = calc_cost(_adj_pct_normalized, _adj_ing_df_cached)

    mcol1, mcol2, mcol3, mcol4, mcol5 = st.columns([1, 1, 1, 1, 1])
    with mcol1:
        _tdelta = f"{_adj_total - _orig_total:+.1f}%"
        st.metric("配比合计", f"{_adj_total:.1f}%", delta=_tdelta)
    with mcol2:
        _cdelta = f"{_adj_cost - _orig_cost:+.4f}" if abs(_adj_cost - _orig_cost) > 0.0001 else "0"
        st.metric("配方成本", f"¥{_adj_cost:.4f}/kg", delta=_cdelta)
    with mcol3:
        if _orig_cost > 0:
            cr = (_adj_cost - _orig_cost) / _orig_cost * 100
            st.metric("成本变化率", f"{cr:+.2f}%")
        else:
            st.metric("成本变化率", "—")

    # 精粗比 & 钙磷比（使用归一化后的配比计算）
    _cf_data = calc_concentrate_forage(_adj_pct_normalized, _adj_ing_df_cached)
    _ca_p_ratio = calc_ca_p_ratio(_adj_pct_normalized, _adj_ing_df_cached)

    with mcol4:
        if _cf_data:
            st.metric(
                "精粗比(参考)",
                f"精{_cf_data['concentrate_pct']}%/粗{_cf_data['forage_pct']}%",
            )
        else:
            st.metric("精粗比(参考)", "—")
    with mcol5:
        if _ca_p_ratio is not None:
            st.metric("钙磷比(参考)", f"{_ca_p_ratio}:1")
        else:
            st.metric("钙磷比(参考)", "—")

    # 合计警告
    if abs(_adj_total - 100.0) > 0.5:
        st.warning(f"⚠️ 配比合计 {_adj_total:.1f}%，与100%偏差较大。保存时将自动归一化到100%。")
    elif abs(_adj_total - 100.0) > 0.01:
        st.info(f"ℹ️ 当前配比合计 {_adj_total:.1f}%，保存时将自动归一化到100%。")

    # ── 汇总行（合计 / 标准 / 差）───
    st.markdown("---")
    st.markdown("**营养汇总**")

    _summary_rows = calc_summary_rows(
        adj_pct=_adj_pct_normalized,
        ingredients_df=_adj_ing_df_cached,
        nutrient_names=_nut_cols,
        requirements=_requirements_cached,
        nutrient_map=_adj_nutrient_map_cached,
    )
    if len(_summary_rows) == 3:
        summary_data = []
        labels = ["**合计**", "**标准**", "**与标准的差**"]
        for label, srow in zip(labels, _summary_rows):
            rd = {"": label}
            for nc in _nut_cols:
                v = srow.get(nc, "")
                if isinstance(v, (int, float)):
                    rd[nc] = round(v, 4)
                else:
                    rd[nc] = v
            summary_data.append(rd)
        summary_df = pd.DataFrame(summary_data)

        try:
            styled_summary = summary_df.style.applymap(
                lambda x: "font-weight: bold;" if isinstance(x, str) and (x.startswith("✅") or x.startswith("❌")) or False else ""
            ).format({nc: "{:.4f}" for nc in _nut_cols})
            st.dataframe(styled_summary, use_container_width=True, hide_index=True)
        except:
            st.dataframe(summary_df, use_container_width=True, hide_index=True)

    # ── 达标情况快速判断（带容差，避免浮点精度误判）───
    # 容差：绝对0.01 或相对1%，取较大值（LP求解结果可能有微小偏差）
    _TOL_ABS = 0.01
    if _requirements_cached:
        _adj_nuts = calc_nutrients(_adj_pct_normalized, _adj_ing_df_cached, _adj_nutrient_map_cached)
        _pass_list, _fail_list = [], []
        for std_col, req_val in _requirements_cached.items():
            try:
                rf = float(req_val)
                av = _adj_nuts.get(std_col, 0)
                _tolerance = max(_TOL_ABS, abs(rf) * 0.01)
                if av >= rf - _tolerance:
                    _pass_list.append(f"**{std_col}**: {av:.4f}")
                else:
                    _fail_list.append(f"**{std_col}**: 需要≥{rf}, 实际{av:.4f}, 差{round(av-rf,4)}")
            except (ValueError, TypeError):
                continue

        if _fail_list:
            for item in _fail_list[:5]:
                st.error(item)
        elif _pass_list:
            st.success(f"✅ 微调后全部 {len(_pass_list)} 项营养指标达标！")

    # ── 重置按钮 ──
    st.divider()
    rcol1, rcol2, rcol3 = st.columns([1, 3, 3])
    with rcol1:
        if st.button("↩️ 重置为原始配比", key="adj_reset_btn"):
            st.session_state._adj_number_values = dict(original_pct)
            # 设置标志位让主页面自动重新打开dialog
            st.session_state._should_reopen_adj_dialog = True
            st.rerun()

    # ── 保存按钮 ──
    with rcol2:
        if st.button("💾 保存并关闭", key="adj_save_btn", type="primary"):
            # 将当前微调结果保存为"已保存微调"
            _saved_adj = dict(st.session_state.get("_adj_number_values", {}))
            
            # 自动归一化：如果配比合计不是100%，则归一化到100%
            _total = sum(_saved_adj.values())
            if abs(_total - 100.0) > 0.01 and _total > 0:
                _normalized = {k: v / _total * 100.0 for k, v in _saved_adj.items()}
                _saved_adj = _normalized
                st.info(f"ℹ️ 配比已自动归一化：{_total:.1f}% → 100%")
            
            st.session_state._saved_adjusted_pct = _saved_adj
            # 计算微调后的成本
            _adj_ing_for_save = st.session_state.get("_adj_ing_df", _adj_ing_df)
            _saved_cost = calc_cost(_saved_adj, _adj_ing_for_save)
            st.session_state._saved_adjusted_cost = _saved_cost
            # 不设置重开标志 → dialog关闭后不再自动打开
            st.rerun()

    with rcol3:
        # 显示保存状态
        _has_saved = "_saved_adjusted_pct" in st.session_state
        if _has_saved:
            st.caption("✅ 已保存微调版本，可再次微调或下载")
        else:
            st.caption("💡 保存后可导出微调后的配方")


# ── Session State 初始化 ─────────────────────────────────────
if "uploaded_ingredients_df" not in st.session_state:
    st.session_state.uploaded_ingredients_df = None
if "uploaded_standards_path" not in st.session_state:
    st.session_state.uploaded_standards_path = None  # 临时文件路径（不是文件对象）
if "uploaded_standards_data" not in st.session_state:
    st.session_state.uploaded_standards_data = {}


# ╔══════════════════════════════════════════════════════════════╗
# ║                   页面配置                                    ║
# ╚══════════════════════════════════════════════════════════════╝

st.set_page_config(
    page_title="饲料配方优化系统",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── 自定义CSS ─────────────────────────────────────────────
st.markdown("""
<style>
    /* 主色调 */
    :root {
        --primary: #1b5e20;
        --primary-light: #4caf50;
        --bg-light: #f1f8e9;
        --text-dark: #263238;
    }
    /* 主区域标题栏 */
    .app-header {
        background: linear-gradient(135deg, #1b5e20 0%, #2e7d32 50%, #388e3c 100%);
        padding: 24px 32px;
        border-radius: 16px;
        margin-bottom: 20px;
        color: white;
        box-shadow: 0 4px 20px rgba(27,94,32,0.25);
    }
    .app-header h1 {
        color: white !important;
        font-size: 2rem;
        font-weight: 700;
        margin: 0;
        padding: 0;
    }
    .app-header .subtitle {
        color: rgba(255,255,255,0.85);
        font-size: 0.95rem;
        margin-top: 6px;
    }
    .app-header .stats {
        display: flex;
        gap: 24px;
        margin-top: 12px;
    }
    .app-header .stat-item {
        background: rgba(255,255,255,0.15);
        padding: 6px 16px;
        border-radius: 20px;
        font-size: 0.85rem;
        backdrop-filter: blur(4px);
    }
    /* 指标卡片 */
    .metric-card {
        background: linear-gradient(135deg, #e8f5e9, #c8e6c9);
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        border: 1px solid #a5d6a7;
    }
    .metric-card.warn {
        background: linear-gradient(135deg, #fff3e0, #ffe0b2);
        border-color: #ffcc80;
    }
    .metric-card.error {
        background: linear-gradient(135deg, #ffebee, #ffcdd2);
        border-color: #ef9a9a;
    }
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #1b5e20;
    }
    div[data-testid="stMetricValue"] {
        font-size: 1.8rem;
    }
    /* 成功/警告标签 */
    .badge-pass {
        background: #e8f5e9;
        color: #2e7d32;
        padding: 2px 8px;
        border-radius: 10px;
        font-weight: 600;
    }
    .badge-fail {
        background: #ffebee;
        color: #c62828;
        padding: 2px 8px;
        border-radius: 10px;
        font-weight: 600;
    }
    /* 侧边栏标题 */
    .sidebar-title {
        font-size: 1.3rem;
        font-weight: 700;
        color: #2e7d32;
        margin-bottom: 1rem;
    }
    /* 分节标题 */
    .section-header {
        font-size: 1.1rem;
        font-weight: 600;
        color: #1b5e20;
        border-bottom: 2px solid #4caf50;
        padding-bottom: 4px;
        margin-top: 1rem;
    }
    /* 侧边栏按钮样式 */
    div[data-testid="stSidebar"] button[kind="primary"] {
        background: linear-gradient(135deg, #2e7d32, #43a047) !important;
        border: none !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        padding: 12px !important;
    }
    /* 侧边栏分段间距 */
    div[data-testid="stSidebar"] .stMarkdown hr {
        margin: 0.8rem 0;
    }
    /* 隐藏默认footer */
    footer {visibility: hidden;}
    #MainMenu {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ╔══════════════════════════════════════════════════════════════╗
# ║                   数据缓存                                    ║
# ╚══════════════════════════════════════════════════════════════╝

@st.cache_data(ttl=3600)
def load_all_data(uploaded_ingredients_df=None, uploaded_standards_path=None, _standards_mtime=0):
    """加载原料数据和饲养标准（缓存1小时）。
    返回 (ingredients, standards_data)
    standards_data: 始终为 {sheet_name: [standard_names]} 格式。
      单 sheet → {"饲养标准": [...]} 或 {sheet0: [...]}
      多 sheet → {"猪": [...], "牛": [...]}
    _standards_mtime: 内部参数，标准文件修改时间，用于缓存失效。调用方请勿传值。
    """
    # 原料数据
    if uploaded_ingredients_df is not None:
        ingredients = uploaded_ingredients_df
    else:
        ingredients = load_ingredients(INGREDIENTS_FILE,
                                       phos_enabled=st.session_state.get("phos_enabled", True),
                                       phos_custom_rates=st.session_state.get("phos_custom_rates") or None,
                                       phos_per_ingredient=st.session_state.get("phos_per_ingredient") or None)

    # 饲养标准：确定使用哪个文件
    standards_file_to_use = uploaded_standards_path if uploaded_standards_path else STANDARDS_FILE

    try:
        import openpyxl
        wb = openpyxl.load_workbook(standards_file_to_use, data_only=True)

        # 过滤出数据 sheet（排除说明类 sheet）
        NON_DATA_SHEETS = {"填写说明", "说明", "Notes", "README", "Sheet"}
        data_sheets = [s for s in wb.sheetnames if s not in NON_DATA_SHEETS]

        if not data_sheets:
            data_sheets = [wb.sheetnames[0]]

        if "饲养标准" in data_sheets:
            # 旧格式：单个「饲养标准」sheet
            ws = wb["饲养标准"]
            names = _extract_standard_names(ws)
            standards_data = {"饲养标准": names}
        elif len(data_sheets) > 1:
            # 新格式：按物种分多个 sheet
            standards_data = {}
            for sn in data_sheets:
                ws = wb[sn]
                names = _extract_standard_names(ws)
                if names:
                    standards_data[sn] = names
            if not standards_data:
                # 回退：取第一个 sheet
                ws = wb[wb.sheetnames[0]]
                names = _extract_standard_names(ws)
                standards_data = {wb.sheetnames[0]: names}
        else:
            # 单个数据 sheet
            ws = wb[data_sheets[0]]
            names = _extract_standard_names(ws)
            standards_data = {data_sheets[0]: names}

        wb.close()
    except Exception as e:
        standards_data = {"默认": ["标准0"]}

    return ingredients, standards_data


def _extract_standard_names(ws):
    """从 openpyxl worksheet 提取标准名称列表（跳过表头行）。"""
    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).strip():
            names.append(str(row[0]).strip())
    if not names:
        names = ["标准(第1行)"]
    return names


@st.cache_data(ttl=600, show_spinner=False)
def load_single_standard(standard_index: int, uploaded_standards_path=None, sheet_name=None, _file_mtime=0):
    """加载单个饲养标准（缓存10分钟）。
    若上传了文件，则从上传文件加载。
    uploaded_standards_path: 上传的标准文件路径（字符串），或 None
    sheet_name: 多 sheet 格式时指定物种 sheet（如 "猪"、"牛"），None 则自动检测
    _file_mtime: 内部参数，用于缓存失效（文件修改时间）。调用方请勿手动传值。

    支持两种标准格式：
      1. 浓度格式（常规）：每列是营养浓度（Mcal/kg、%等）
      2. 日摄入量格式（/d）：每列是每日需要量，程序自动根据采食量换算为浓度
    """
    # 确定使用哪个文件
    standards_file_to_use = uploaded_standards_path if uploaded_standards_path else STANDARDS_FILE

    MJ_TO_MCAL = 1.0 / 4.184  # MJ → Mcal 换算因子

    def _safe_float(val):
        """安全转换为浮点数，支持中文全角句号"""
        if isinstance(val, str):
            val = val.replace("。", ".")  # 中文全角句号 → 英文句号
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _parse_row_into_requirements(row, df, name_col):
        """从 Excel 行解析营养需要量字典"""
        nutrient_requirements = {}
        for col in df.columns:
            if col == name_col:
                continue
            val = row[col]
            if pd.notna(val):
                fval = _safe_float(val)
                if fval is None:
                    continue
                # 能量单位处理：
                # 仅当该指标在 NUTRIENT_MAP 中有映射时才换算（说明目标原料列使用 Mcal/kg）
                # 无映射的指标（如用户自定义的"其他消化能MJ/Kg"）保持原单位
                col_store = str(col)
                if ("kJ/kg" in col_store or "kJ/Kg" in col_store) and str(col) in NUTRIENT_MAP:
                    fval = fval / 4184.0
                    col_store = col_store.replace("kJ/kg", "Mcal/kg").replace("kJ/Kg", "Mcal/kg")
                elif ("MJ/kg" in col_store or "MJ/Kg" in col_store) and str(col) in NUTRIENT_MAP:
                    fval = fval * MJ_TO_MCAL
                    col_store = col_store.replace("MJ/kg", "Mcal/kg").replace("MJ/Kg", "Mcal/kg")
                nutrient_requirements[col_store] = fval
        return nutrient_requirements

    def _convert_per_day_to_concentration(nutrient_requirements, standard_name):
        """检测并转换「日摄入量格式」→「浓度格式」。

        日摄入量格式特征：列名含 /d（如 消化能MJ/d、粗蛋白质g./d），且存在 采食量 列。
        转换：每日量 ÷ 采食量 = 每kg浓度；
              g/kg → %（÷10）；mg/kg → 保留mg/kg；IU/kg → 保留IU/kg；
              MJ/kg → Mcal/kg（×MJ_TO_MCAL）；Mcal/kg → 直接保留。
        """
        # 检测是否为日摄入量格式
        cols_str = [str(k) for k in nutrient_requirements.keys()]
        has_per_day = any("/d" in c for c in cols_str)
        if not has_per_day:
            return nutrient_requirements, standard_name

        # 找采食量（支持多种列名变体）
        feed_intake = None
        for k in list(nutrient_requirements.keys()):
            k_str = str(k)
            if "采食" in k_str or "DMI" in k_str.upper() or "干物质采食" in k_str:
                feed_intake = nutrient_requirements.pop(k)
                break

        if not feed_intake or feed_intake <= 0:
            # 有/d列但无有效采食量 → 保留原样，附标注
            return nutrient_requirements, standard_name + " [⚠️ 缺少采食量]"

        converted = {}
        per_day_cols_found = []
        unconverted_cols = []

        for col, val in list(nutrient_requirements.items()):
            col_str = str(col)
            # 跳过非 /d 列
            if "/d" not in col_str:
                converted[col_str] = val
                continue

            per_day_cols_found.append(col_str)
            per_kg = val / feed_intake
            col_upper = col_str.upper()

            # ── 检测单位 ──
            is_energy = any(kw in col_str for kw in
                           ["消化能", "代谢能", "净能", "总能", "DE", "ME", "NE", "GE"])
            is_mj = "MJ" in col_upper
            is_mcal = "MCAL" in col_upper or "MC" in col_upper or "Mc" in col_str
            is_mg = "mg" in col_str.lower() or "MG" in col_upper
            is_iu = "IU" in col_upper or "iu" in col_str.lower()

            # ── 能量列：X/d → Mcal/kg ──
            if is_energy:
                if is_mj:
                    per_kg_mcal = per_kg * MJ_TO_MCAL
                elif is_mcal:
                    per_kg_mcal = per_kg  # 已是 Mcal
                else:
                    per_kg_mcal = per_kg * MJ_TO_MCAL  # 默认按 MJ

                if "消化能" in col_str or "DE" in col_upper:
                    converted["DE, Mcal/kg"] = round(per_kg_mcal, 6)
                elif "代谢能" in col_str or "ME" in col_upper:
                    converted["ME, Mcal/kg"] = round(per_kg_mcal, 6)
                elif "净能" in col_str or "NE" in col_upper:
                    converted["综合净能, Mcal/kg"] = round(per_kg_mcal, 6)
                elif "总能" in col_str or "GE" in col_upper:
                    converted["总能, Mcal/kg"] = round(per_kg_mcal, 6)
                else:
                    converted[f"{col_str.replace('/d','')}, Mcal/kg"] = round(per_kg_mcal, 6)

            # ── mg/d → mg/kg（微量矿物，不转百分比）──
            elif is_mg:
                per_kg_val = round(per_kg, 6)
                if "铜" in col_str and "氨基酸铜" not in col_str:
                    converted["铜, mg/kg"] = per_kg_val
                elif "碘" in col_str:
                    converted["碘, mg/kg"] = per_kg_val
                elif "铁" in col_str and "氨基酸铁" not in col_str:
                    converted["铁, mg/kg"] = per_kg_val
                elif "锰" in col_str and "氨基酸锰" not in col_str:
                    converted["锰, mg/kg"] = per_kg_val
                elif "硒" in col_str:
                    converted["硒, mg/kg"] = per_kg_val
                elif "锌" in col_str and "氨基酸锌" not in col_str:
                    converted["锌, mg/kg"] = per_kg_val
                elif "钴" in col_str:
                    converted["钴, mg/kg"] = per_kg_val
                elif "胡萝卜素" in col_str:
                    converted["胡萝卜素，mg/kg"] = per_kg_val
                else:
                    converted[col_str.replace("/d", "") + ", mg/kg"] = per_kg_val

            # ── IU/d → IU/kg（维生素）──
            elif is_iu:
                per_kg_val = round(per_kg, 6)
                col_lower = col_str.lower()
                if "维生素a" in col_lower or "va" in col_upper:
                    converted["维生素A，IU/kg"] = per_kg_val
                elif "维生素d" in col_lower or "vd" in col_upper:
                    converted["维生素D3，IU/kg"] = per_kg_val
                elif "维生素e" in col_lower or "ve" in col_upper:
                    converted["维生素E，IU/kg"] = per_kg_val
                elif "维生素k" in col_lower or "vk" in col_upper:
                    converted["维生素K，IU/kg"] = per_kg_val
                else:
                    converted[col_str.replace("/d", "") + ", IU/kg"] = per_kg_val

            # ── 营养列：g/d → g/kg → % ──
            elif "粗蛋白" in col_str or "CP" in col_upper or "蛋白质" in col_str:
                converted["粗蛋白质CP,%"] = round(per_kg / 10.0, 6)
            elif "钙" in col_str and "磷" not in col_str and "氨基酸钙" not in col_str:
                converted["钙，%"] = round(per_kg / 10.0, 6)
            elif "总磷" in col_str:
                converted["总磷，%"] = round(per_kg / 10.0, 6)
            elif "有效磷" in col_str:
                converted["有效磷，%"] = round(per_kg / 10.0, 6)
            elif "食盐" in col_str or ("钠" in col_str and "钾" not in col_str):
                converted["食盐%"] = round(per_kg / 10.0, 6)
            elif "氯" in col_str or "CL" in col_upper or "氯化" in col_str:
                converted["氯，%"] = round(per_kg / 10.0, 6)
            elif "镁" in col_str or "MG" in col_upper:
                converted["镁，%"] = round(per_kg / 10.0, 6)
            elif "钾" in col_str:
                converted["钾，%"] = round(per_kg / 10.0, 6)
            elif "硫" in col_str:
                converted["硫，%"] = round(per_kg / 10.0, 6)
            elif "赖氨酸" in col_str:
                converted["赖氨酸%"] = round(per_kg / 10.0, 6)
            elif "蛋氨酸" in col_str and "胱" not in col_str:
                converted["蛋氨酸%"] = round(per_kg / 10.0, 6)
            elif "蛋胱氨酸" in col_str or "胱氨酸" in col_str:
                converted["蛋胱氨酸%"] = round(per_kg / 10.0, 6)
            elif "苏氨酸" in col_str:
                converted["苏氨酸%"] = round(per_kg / 10.0, 6)
            elif "色氨酸" in col_str:
                converted["色氨酸%"] = round(per_kg / 10.0, 6)
            elif "精氨酸" in col_str:
                converted["精氨酸%"] = round(per_kg / 10.0, 6)
            elif "组氨酸" in col_str:
                converted["组氨酸%"] = round(per_kg / 10.0, 6)
            elif "异亮氨酸" in col_str:
                converted["异亮氨酸%"] = round(per_kg / 10.0, 6)
            elif "亮氨酸" in col_str:
                converted["亮氨酸%"] = round(per_kg / 10.0, 6)
            elif "缬氨酸" in col_str:
                converted["缬氨酸%"] = round(per_kg / 10.0, 6)
            elif "苯丙氨酸" in col_str:
                converted["苯丙氨酸%"] = round(per_kg / 10.0, 6)
            elif "酪氨酸" in col_str:
                converted["酪氨酸%"] = round(per_kg / 10.0, 6)
            elif "亚油酸" in col_str:
                converted["亚油酸，%"] = round(per_kg / 10.0, 6)
            elif "粗脂肪" in col_str or "EE" in col_upper:
                converted["粗脂肪%"] = round(per_kg / 10.0, 6)
            elif "粗纤维" in col_str or "CF" in col_upper:
                converted["粗纤维%"] = round(per_kg / 10.0, 6)
            else:
                # 未识别的 /d 列 → 按通用 g/d→% 换算
                converted[col_str.replace("/d", "") + ", %"] = round(per_kg / 10.0, 6)

        # 附注：记录未识别列（调试用）
        if unconverted_cols:
            _ = unconverted_cols  # 保留调试入口

        if per_day_cols_found:
            standard_name = f"{standard_name} [日粮→浓度换算]"

        return converted, standard_name

    try:
        try:
            df = pd.read_excel(standards_file_to_use, sheet_name=sheet_name or "饲养标准")
        except Exception:
            try:
                df = pd.read_excel(standards_file_to_use, sheet_name=sheet_name or 0)
            except Exception:
                df = pd.read_excel(standards_file_to_use, sheet_name=0)

        # 新格式：每行一个标准，第一列是标准名称
        first_col = df.columns[0]
        if first_col != "Unnamed: 0":
            # 新格式
            name_col = first_col
            if standard_index < len(df):
                row = df.iloc[standard_index]
                standard_name = str(row[name_col]) if pd.notna(row[name_col]) else f"标准{standard_index}"
            else:
                row = df.iloc[0]
                standard_name = str(row[name_col])

            nutrient_requirements = _parse_row_into_requirements(row, df, name_col)
            nutrient_requirements, standard_name = _convert_per_day_to_concentration(
                nutrient_requirements, standard_name
            )
            return nutrient_requirements, standard_name
        else:
            # 旧格式：第一列是指标名，第一行是标准名称
            df = df.set_index(first_col)
            df = df.T
            if standard_index < len(df):
                row = df.iloc[standard_index]
                standard_name = df.index[standard_index]
            else:
                row = df.iloc[0]
                standard_name = df.index[0]
            nutrient_requirements = {
                k: (_safe_float(v) if pd.notna(v) else None)
                for k, v in row.to_dict().items()
            }
            return nutrient_requirements, standard_name
    except Exception as e:
        # 异常时返回空标准
        return {"粗蛋白质%": 16.0}, f"标准{standard_index}"


# ╔══════════════════════════════════════════════════════════════╗
# ║                   辅助函数                                    ║
# ╚══════════════════════════════════════════════════════════════╝

def _build_nutrient_df(result, ingredients_df):
    """构建营养对比 DataFrame（结果 vs 标准）。"""
    if result["status"] != "optimal":
        return pd.DataFrame()
    rows = []
    requirements = result.get("requirements", {})
    for std_col, req_val in requirements.items():
        if req_val is None:
            continue
        actual = result.get("nutrients", {}).get(std_col, 0.0)
        diff = actual - float(req_val)
        rows.append({
            "营养指标": std_col,
            "需要量": float(req_val),
            "实际含量": actual,
            "差值": round(diff, 4),
            "达成": "✅" if round(diff, 4) >= 0 else "❌",
        })
    return pd.DataFrame(rows)


def _build_formula_df(result, ingredients_df, auto_ingredients=None):
    """构建配方组成 DataFrame。无论求解状态都尝试构建。
    auto_ingredients: 自动补充的原料列表，这些原料会在名称后标注 '*自动补充*'
    """
    ingredients = result.get("ingredients", {})
    if not ingredients:
        return pd.DataFrame()

    if auto_ingredients is None:
        auto_ingredients = []

    rows = []
    prices = ingredients_df["价格"].to_dict()
    total = sum(float(v) for v in ingredients.values())
    for name, pct in ingredients.items():
        try:
            pct_val = float(pct)
            price_val = float(prices.get(name, 0))
        except (ValueError, TypeError):
            continue

        # 处理带标签的名称：去掉 "[上传] " 或 "[默认] " 前缀，但保留来源信息
        source_emoji = ""
        real_name = name
        if name.startswith("[上传] "):
            real_name = name[5:]  # 去掉 "[上传] "
            source_emoji = " 📤"  # 用 emoji 表示来源
        elif name.startswith("[默认] "):
            real_name = name[5:]  # 去掉 "[默认] "
            source_emoji = " 📚"  # 用 emoji 表示来源
        
        # 如果是自动补充的原料，在名称后标注
        display_name = real_name + source_emoji
        if name in auto_ingredients:
            display_name = f"{real_name} *自动补充*" + source_emoji

        rows.append({
            "原料名称": display_name,
            "配比(%)": round(pct_val, 2),
            "单价(元/kg)": price_val,
            "成本贡献(元)": round(pct_val / 100 * price_val, 4),
        })
    if not rows:
        return pd.DataFrame()

    # 按配比降序
    df = pd.DataFrame(rows).sort_values("配比(%)", ascending=False)
    # 添加合计行
    cost_sum = sum(r["成本贡献(元)"] for r in rows)
    total_row = {
        "原料名称": "【合计】",
        "配比(%)": round(total, 2),
        "单价(元/kg)": "-",
        "成本贡献(元)": round(cost_sum, 4),
    }
    df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)
    return df


def _build_nutrient_df(result, ingredients_df):
    """构建营养对比 DataFrame（结果 vs 标准）。无论状态都尝试构建。"""
    requirements = result.get("requirements", {})
    if not requirements:
        return pd.DataFrame()

    rows = []
    for std_col, req_val in requirements.items():
        if req_val is None:
            continue
        try:
            req_float = float(req_val)
        except (ValueError, TypeError):
            continue
        actual = result.get("nutrients", {}).get(std_col, 0.0)
        try:
            actual = float(actual)
        except (ValueError, TypeError):
            actual = 0.0
        diff = actual - req_float
        rows.append({
            "营养指标": std_col,
            "需要量": req_float,
            "实际含量": actual,
            "差值": round(diff, 4),
            "达成": "✅" if round(diff, 4) >= 0 else "❌",
        })
    return pd.DataFrame(rows)


def _highlight_nutrition_row(row):
    """根据达标状态高亮行。"""
    if row.get("达成") == "✅":
        return ["background-color: #f1f8e9;"] * len(row)
    elif row.get("达成") == "❌":
        return ["background-color: #fbe9e7;"] * len(row)
    return [""] * len(row)


def _color_diff(val):
    """差值颜色：正数绿色，负数红色。"""
    if isinstance(val, (int, float)):
        if val >= 0:
            return "color: #2e7d32; font-weight: 600;"
        else:
            return "color: #c62828; font-weight: 600;"
    return ""


# ── 中国饲料原料分类（基于国标 GB/T 10647-2008）──────────────

# 分类规则：按原料名称关键字 + 营养特征 归类
_FEED_CATEGORY_RULES = [
    # (类别名, 关键字列表, 特征说明, emoji)
    ("能量饲料", ["玉米", "小麦", "大麦", "高粱", "稻谷", "碎米", "糙米", "甘薯",
                  "木薯", "栗", "黑麦", "燕麦", "次粉", "麦麸", "米糠", "米糠粕",
                  "小麦麸", "玉米DDGS", "玉米胚芽", "葡萄糖", "蔗糖", "乳糖",
                  "牛奶乳糖", "乳清粉", "淀粉"],
     "提供能量、淀粉", "🌽"),
    ("蛋白质饲料", ["豆粕", "豆饼", "大豆", "菜粕", "菜饼", "菜籽", "棉粕", "棉饼",
                   "棉籽", "花生", "芝麻", "亚麻", "向日葵", "玉米蛋白", "鱼粉",
                   "肉骨粉", "血粉", "羽毛粉", "肉粉", "啤酒酵母", "酵母",
                   "粉浆蛋白", "膨化豆粕", "全脂大豆", "去皮豆粕", "复合氨基酸",
                   "DDGS", "酪蛋白", "明胶", "蚕豆", "豌豆", "蚕蛹"],
     "提供粗蛋白、氨基酸", "🥩"),
    ("粗饲料", ["干草", "青贮", "青割", "稻草", "秸秆", "秸", "秕壳", "草粉",
                "糟", "酒糟", "啤酒糟", "木薯渣", "苜蓿", "槐叶", "柳叶",
                "胡枝子", "骆驼刺", "沙蒿", "芨芨草", "芦苇", "混合牧草",
                "干稗", "干冰草", "狗尾草", "谷草", "碱草", "羊草",
                "野干草", "野青草", "黑麦草", "冰草", "白茅"],
     "提供纤维、NDF/ADF", "🌿"),
    ("矿物质饲料", ["石粉", "贝壳粉", "磷酸氢钙", "磷酸二氢钙", "磷酸一二钙",
                   "磷酸三钙", "骨粉", "食盐", "碳酸钙", "硫酸钙",
                   "磷酸二氢钠", "磷酸二氢钾", "磷酸氢二钠", "磷酸氢二铵",
                   "磷酸氢铵", "磷酸二氢铵"],
     "提供钙、磷、钠等矿物元素", "🪨"),
    ("氨基酸添加剂", ["赖氨酸", "蛋氨酸", "苏氨酸", "色氨酸", "精氨酸",
                    "缬氨酸", "异亮氨酸", "胱氨酸", "甘氨酸", "MHA",
                    "L-Lys", "L-Thr", "L-Try", "DL-"],
     "补充必需氨基酸", "🧬"),
    ("其他添加剂", ["复合多维", "预混料", "维生素", "氯化胆碱", "胆碱[",
                   "酶", "植酸酶", "酸化剂", "柠檬酸", "富马酸",
                   "甜味", "香味", "色素", "防霉", "抗氧化",
                   "微量元素", "微量", "硫酸铜", "硫酸锌", "硫酸锰",
                   "硫酸亚铁", "硫酸铁", "氧化锌", "亚硒酸钠", "碘",
                   "钴", "硒", "多维", "VB", "VC", "VA", "VD", "VE", "VK",
                   "生物素", "烟酸", "叶酸", "肌醇", "泛酸", "吡哆",
                   "甜菜碱", "L-肉碱"],
     "提供维生素、微量元素、酶制剂等", "⚗️"),
]

# 命名中包含这些关键字的原料按关键字优先级判定（先匹配先得）
# 在某些原料可以同时匹配多个时，靠前的规则优先
_FEED_CATEGORY_OVERRIDE = {
    # 明确归类到特定类别的特殊原料
    "玉米DDGS": "蛋白质饲料",  # DDGS CP通常>20%，应归蛋白质饲料
    "乳清粉": "能量饲料",
    "尿素": "蛋白质饲料",  # 非蛋白氮
    "缩二脲": "蛋白质饲料",
    "双缩脲": "蛋白质饲料",
    "磷酸脲": "蛋白质饲料",
    "复合氨基酸": "蛋白质饲料",
    "羽毛粉": "蛋白质饲料",
    "全脂大豆": "蛋白质饲料",
    "大豆": "蛋白质饲料",  # 大豆CP约35%
    "大豆荚": "蛋白质饲料",
    "花生蔓": "粗饲料",
    "棉籽": "蛋白质饲料",
    "棉籽壳": "粗饲料",
    "玉米秸": "粗饲料",
    "小麦秸": "粗饲料",
    "大豆秸": "粗饲料",
    "小麦秕壳": "粗饲料",
    "玉米蛋白饲料": "蛋白质饲料",
    "玉米蛋白粉": "蛋白质饲料",
    "水花生": "粗饲料",
    "甘薯藤": "粗饲料",
    "甘薯蔓": "粗饲料",
    "甘薯叶粉": "粗饲料",
    "槐树叶粉": "粗饲料",
    # 青贮类 → 粗饲料（即使名字含"玉米"等能量关键字）
    "青贮": "粗饲料",
    "青割": "粗饲料",
    "大麦青割": "粗饲料",
    "高粱青割": "粗饲料",
    "燕麦青割": "粗饲料",
    "玉米青割": "粗饲料",
    "向日葵青割": "粗饲料",
    "稻草青割": "粗饲料",
}


def _classify_feed_cn(name: str) -> tuple:
    """将原料归类到中国饲料分类体系。
    返回 (类别名, emoji, 特征说明)
    """
    name_str = str(name).strip()

    # 检查手动覆盖
    for key, category in _FEED_CATEGORY_OVERRIDE.items():
        if key in name_str:
            for cat_name, _, desc, emoji in _FEED_CATEGORY_RULES:
                if cat_name == category:
                    return (cat_name, emoji, desc)
            break

    # 按优先级匹配关键字
    for cat_name, keywords, desc, emoji in _FEED_CATEGORY_RULES:
        for kw in keywords:
            if kw in name_str:
                return (cat_name, emoji, desc)

    return ("其他", "📦", "其他功能添加剂")


def _build_auto_ingredient_detail(results, ingredients_df, nutrient_map):
    """构建自动补充原料的营养详情 DataFrame。
    results: result dict 包含 ingredients, requirements, nutrients
    ingredients_df: 求解时用的原料 DataFrame（含营养数据）
    nutrient_map: 标准指标名 → 原料库列名 的映射
    """
    auto_ingredients = results.get("auto_ingredients", [])
    formula = results.get("ingredients", {})
    if not auto_ingredients or not formula:
        return pd.DataFrame(), {}

    # 选择要展示的营养指标：能量 + 粗蛋白 + 钙磷 + 核心氨基酸
    priority_cols = []
    for key in ["猪消化能MC/Kg", "禽代谢能MC/Kg", "综合净能(牛)MC/Kg",
                "羊消化能MC/Kg", "消化能(其他)MC/Kg",
                "粗蛋白%", "粗脂肪%", "粗纤维%",
                "钙%", "总磷%", "非植酸磷%", "食盐%",
                "赖氨酸%", "蛋氨酸%", "蛋胱氨酸%", "苏氨酸%", "色氨酸%",
                "干物质%", "中性洗涤纤维%", "酸性洗涤纤维%"]:
        if key in ingredients_df.columns:
            priority_cols.append(key)

    # 补充 nutrient_map 中的列
    mapped_cols = list(nutrient_map.values()) if nutrient_map else []
    for col in mapped_cols:
        if col in ingredients_df.columns and col not in priority_cols:
            priority_cols.append(col)

    # 限制展示数量
    display_cols = priority_cols[:15]

    # 按配比排序
    rows = []
    category_stats = {}  # {category: [ingredient_names]}
    for name in sorted(auto_ingredients,
                       key=lambda n: float(formula.get(n, 0)),
                       reverse=True):
        pct = float(formula.get(name, 0))
        if pct <= 0.01:
            continue
        if name not in ingredients_df.index:
            continue

        # 处理带标签的名称：获取真实名称（去掉标签）
        real_name = name
        if name.startswith("[上传] "):
            real_name = name[5:]
        elif name.startswith("[默认] "):
            real_name = name[5:]
        
        cat_name, emoji, desc = _classify_feed_cn(real_name)  # 使用真实名称分类
        if cat_name not in category_stats:
            category_stats[cat_name] = []
        category_stats[cat_name].append(real_name)  # 存储真实名称

        row_data = ingredients_df.loc[name]  # 使用带标签的名称获取数据
        row = {
            "原料名称": real_name,  # 显示真实名称
            "配比(%)": round(pct, 2),
            "中国分类": f"{emoji} {cat_name}",
        }
        for col in display_cols:
            val = row_data.get(col, 0)
            try:
                val_f = float(val)
                if pd.isna(val_f):
                    val_f = 0.0
                row[col] = round(val_f, 3)
            except (ValueError, TypeError):
                row[col] = 0.0
        rows.append(row)

    if not rows:
        return pd.DataFrame(), {}

    df = pd.DataFrame(rows)

    # 对每列用数值格式化（除了前3列）
    fmt_dict = {}
    for col in df.columns:
        if col not in ("原料名称", "配比(%)", "中国分类"):
            fmt_dict[col] = "{:.3f}"
    fmt_dict["配比(%)"] = "{:.2f}"

    return df, category_stats


def _build_feed_recommendation(auto_ingredients, category_stats,
                                ingredients_df, nutrient_map):
    """根据自动补充的原料分类生成补充建议。"""
    if not auto_ingredients or not category_stats:
        return None

    # 各类别典型原料（静态推荐列表）
    _TYPICAL = {
        "能量饲料": ["玉米[1级8.7%]", "小麦[2级13.9%]", "麦麸", "米糠[2级]",
                     "次粉[1级15.4%]", "玉米DDGS[28%]", "高粱[1级]"],
        "蛋白质饲料": ["豆粕(43%)", "菜籽粕[38.6%]", "棉籽粕[47%]",
                      "鱼粉(进口)", "花生仁粕[47.8%]", "玉米蛋白粉[63.5%]",
                      "棉籽饼[36.3%]"],
        "粗饲料": ["苜蓿干草", "玉米青贮", "稻草", "小麦秸",
                    "苜蓿草粉[14%]", "花生蔓", "甘薯蔓"],
        "矿物质饲料": ["石粉", "磷酸氢钙", "贝壳粉", "食盐", "骨粉22"],
        "氨基酸添加剂": ["赖氨酸(98%)", "DL-蛋氨酸", "L-苏氨酸", "L-色氨酸"],
        "其他添加剂": ["复合多维", "氯化胆碱", "植酸酶5000", "复合氨基酸"],
    }

    lines = []

    # === 1. 已补充原料分类汇总 ===
    lines.append("#### 📋 本次自动补充了以下类别原料：\n")
    for cat_name, names in category_stats.items():
        emoji = ""
        desc = ""
        for cn, _, d, e in _FEED_CATEGORY_RULES:
            if cn == cat_name:
                emoji, desc = e, d
                break
        lines.append(f"- {emoji} **{cat_name}**（{desc}）")
        names_str = "、".join(names[:4])
        if len(names) > 4:
            names_str += f"等{len(names)}种"
        lines.append(f"  → 已补充：{names_str}")

    # === 2. 建议用户补充的原料类别 ===
    lines.append("\n#### 💡 建议您在自有原料库中补充以下类别原料：\n")

    recs = []
    for cat_name in category_stats:
        emoji = ""
        desc = ""
        for cn, _, d, e in _FEED_CATEGORY_RULES:
            if cn == cat_name:
                emoji, desc = e, d
                break
        typical = _TYPICAL.get(cat_name, [])[:5]
        if typical:
            typical_str = "、".join(typical)
        else:
            typical_str = "相关原料"
        recs.append(
            f"- {emoji} **{cat_name}**（{desc}）\n"
            f"  → 建议补充：{typical_str} 等"
        )

    # 额外建议：如果缺能量+蛋白，也提示粗饲料
    has_coarse = any(c == "粗饲料" for c in category_stats)
    if not has_coarse and len(category_stats) >= 2:
        recs.append(
            f"- 🌿 **粗饲料**\n"
            f"  → 如需要配制牛羊配方，建议同时补充："
            f"苜蓿干草、玉米青贮、稻草 等"
        )

    lines.extend(recs)

    # === 3. 操作提示 ===
    lines.append("\n#### 📝 操作提示：\n")
    lines.append(
        "1. 将上述推荐原料的**营养数据**填入「原料数据模板.xlsx」"
        "（列名须与本程序数据库完全一致）\n"
        "2. 填写**市场价格**后上传「用户原料库」\n"
        "3. 下次优化时选择「仅用上传原料」，即可完全使用自有原料求解\n"
        "4. 建议各类别至少包含 **2-3 种原料**，以提高配方灵活性"
    )

    return "\n".join(lines)


def _write_feeding_sheet(wb, dm_ingredients: dict, asfed_ingredients: dict, result: dict):
    """在Excel工作簿中新增「投料单」工作表（原样基础 as-fed）。

    参数:
        wb: openpyxl.Workbook 对象
        dm_ingredients: {原料名: DM基础配比%}（来自result["ingredients"]）
        asfed_ingredients: {原料名: 原样基础配比%}（来自convert_result_to_asfed()）
        result: 求解结果dict（含cost, dm_pct_map等）
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title="投料单(原样基础)")

    # 样式
    title_font = Font(name="微软雅黑", bold=True, size=11)
    header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    bold_font = Font(name="微软雅黑", bold=True, size=10)
    normal_font = Font(name="微软雅黑", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(r, c, value, font=None, fill=None, alignment=None):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font = font or normal_font
        cell.alignment = alignment or center
        cell.border = border
        if fill:
            cell.fill = fill
        return cell

    # DM%信息
    dm_pct_map = result.get("dm_pct_map", {})
    headers = ["原料名称", "DM配比(%)", "干物质(%)", "原样投料(%)"]
    cost_yuan_kg = result.get("cost", 0)

    # 标题行
    cost_t = cost_yuan_kg  # DM基础成本
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = _cell(1, 1, f"投料单 — 原样基础(as-fed) | 配方成本: {cost_t:.4f}元/kg(DM)",
              font=title_font, alignment=center)
    ws.row_dimensions[1].height = 25

    # 表头
    for ci, hdr in enumerate(headers, 1):
        _cell(2, ci, hdr, font=header_font, fill=header_fill, alignment=center)
    ws.row_dimensions[2].height = 30

    # 数据行
    row_idx = 3
    total_asfed = 0.0
    total_dm = 0.0
    # 按as-fed配比降序排列（投料单关注实际投料量）
    for ing_name, pct_asfed in sorted(asfed_ingredients.items(), key=lambda x: -x[1]):
        pct_dm = dm_ingredients.get(ing_name, 0)
        dm_pct = dm_pct_map.get(ing_name, 88.0)
        total_asfed += pct_asfed
        total_dm += pct_dm

        _cell(row_idx, 1, ing_name,
              font=normal_font, alignment=Alignment(horizontal="left"))
        _cell(row_idx, 2, round(pct_dm, 2), alignment=right)
        _cell(row_idx, 3, round(dm_pct, 1), alignment=right)
        _cell(row_idx, 4, round(pct_asfed, 2), alignment=right)
        row_idx += 1

    # 合计行
    ws.row_dimensions[row_idx].height = 20
    _cell(row_idx, 1, "合计", font=bold_font, alignment=center)
    _cell(row_idx, 2, round(total_dm, 2), font=bold_font, alignment=right)
    _cell(row_idx, 3, "", alignment=center)
    _cell(row_idx, 4, round(total_asfed, 2), font=bold_font, alignment=right)
    row_idx += 1

    # 说明行
    row_idx += 1
    ws.cell(row=row_idx, column=1, value=(
        "说明：本表为原样基础(as-fed)实际投料用量，已按各原料干物质含量换算。"
        "换算公式：原样投料% = (DM配比% × DM%) / Σ(DM配比% × DM%)"
    ))
    ws.cell(row=row_idx, column=1).font = Font(name="微软雅黑", size=9, italic=True, color="666666")
    ws.merge_cells(f"A{row_idx}:{get_column_letter(len(headers))}{row_idx}")

    # 列宽
    col_widths = [22, 12, 12, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _diagnose_and_suggest(ing_df, nutrient_requirements, nutrient_map,
                          deduped_matched=None, animal_key="猪"):
    """求解失败时诊断原因并建议补充什么类型的原料。

    分析当前原料池的营养覆盖情况：
    1. 哪些营养指标的上限/下限最难满足
    2. 当前原料池缺乏哪些类型的原料（能量/蛋白/氨基酸/矿物质）
    """
    import numpy as np

    st.markdown("#### 🔍 失败诊断与补充建议")
    st.caption("以下分析基于当前原料池的营养覆盖情况：")

    # 1. 统计当前原料池各列的最大可能贡献
    ing_cols = list(ing_df.columns)
    # 获取原料名称列表
    ing_names = list(ing_df.index)

    # 分析每种营养指标在原料中的最大值
    nutrient_analysis = []
    for std_key, req_val in (deduped_mapped or nutrient_requirements).items():
        if req_val is None:
            continue
        try:
            req_f = float(req_val)
        except (ValueError, TypeError):
            continue
        col = nutrient_map.get(std_key) if nutrient_map else std_key
        if col not in ing_cols:
            continue
        series = pd.to_numeric(ing_df[col], errors="coerce").dropna()
        if len(series) == 0:
            max_possible = 0.0
            avg_available = 0.0
        else:
            max_possible = float(series.max())
            avg_available = float(series.mean())
        nutrient_analysis.append({
            "指标": std_key,
            "原料列": col,
            "需求值": req_f,
            "原料最大值": max_possible,
            "原料平均值": avg_available,
            "差距": max(0, req_f - max_possible),
        })

    if not nutrient_analysis:
        st.info("无法分析：未找到匹配的营养指标数据。请检查原料和标准的列名是否一致。")
        return

    # 按差距排序，找出最缺的指标
    nutrient_analysis.sort(key=lambda x: x["差距"], reverse=True)

    # 显示分析表格
    st.markdown("**营养缺口分析（按严重程度排序）：**")
    diag_data = []
    for na in nutrient_analysis[:10]:
        status = "❌ 缺口大" if na["差距"] > 0.01 else ("⚠️ 紧张" if na["需求值"] > 0 and na["原料最大值"] / max(na["需求值"], 0.001) < 1.5 else "✅ 可满足")
        diag_data.append({
            "营养指标": na["指标"],
            "需求值": f"{na['需求值']:.2f}",
            "原料最大可提供": f"{na['原料最大值']:.3f}",
            "缺口": f"{na['差距']:.3f}" if na["差距"] > 0 else "—",
            "状态": status,
        })
    if diag_data:
        st.dataframe(diag_data, use_container_width=True, hide_index=True)

    # 2. 根据缺口类型推荐原料类别
    gaps = [na for na in nutrient_analysis if na["差距"] > 0.005]
    if not gaps:
        st.info("💡 所有营养指标的**理论上限均可满足**，但约束组合（如多约束同时作用、精粗比等）导致无解。"
                "建议：适当放宽某些约束或检查上下限设置是否矛盾。")
        return

    # 分类判断需要补什么类型
    _ENERGY_COLS = {"猪消化能MC/Kg", "禽代谢能MC/Kg", "综合净能(牛)MC/Kg", "羊消化能MC/Kg", "消化能(其他)MC/Kg"}
    _PROTEIN_COLS = {"粗蛋白%"}
    _AA_COLS = {"赖氨酸%", "蛋氨酸%", "蛋胱氨酸%", "苏氨酸%", "色氨酸%",
                "异亮氨酸%", "亮氨酸%", "缬氨酸%", "苯丙氨酸%"}
    _MINERAL_COLS = {"钙%", "总磷%", "非植酸磷%", "钠%", "氯%"}
    _FIBER_COLS = {"粗纤维%", "中性洗涤纤维%", "酸性洗涤纤维%"}

    needed_types = set()
    for g in gaps:
        col = g["原料列"]
        if col in _ENERGY_COLS:
            needed_types.add("energy")
        elif col in _PROTEIN_COLS:
            needed_types.add("protein")
        elif col in _AA_COLS:
            needed_types.add("amino_acid")
        elif col in _MINERAL_COLS:
            needed_types.add("mineral")
        elif col in _FIBER_COLS:
            needed_types.add("fiber")

    # 生成建议
    rec_lines = []
    rec_lines.append("**根据以上缺口，建议补充以下类型的原料：**\n")

    type_recommendations = {
        "energy": ("🌽 能量饲料", "当前能量原料不足，建议添加玉米、小麦、次粉、麦麸、高粱等高能谷物"),
        "protein": ("🥩 蛋白质饲料", "蛋白质含量偏低，建议添加豆粕、棉籽粕、菜籽粕、鱼粉、花生仁粕等"),
        "amino_acid": ("🧬 氨基酸添加剂", "必需氨基酸不足，建议添加赖氨酸(98%)、DL-蛋氨酸、L-苏氨酸、L-色氨酸等"),
        "mineral": ("🪨 矿物质饲料", "钙磷等矿物元素不足，建议添加石粉(碳酸钙)、磷酸氢钙、骨粉、食盐等"),
        "fiber": ("🌿 粗饲料", "粗纤维不足（反刍动物必需），建议添加苜蓿干草、玉米青贮、稻草、羊草等"),
    }

    for ttype, (title, desc) in type_recommendations.items():
        if ttype in needed_types:
            rec_lines.append(f"- {title}：{desc}")

    # 额外提示
    if len(gaps) >= 3:
        rec_lines.append("\n> ⚠️ **注意**：缺口指标较多，说明当前原料池种类过于单一或数量太少。"
                         "建议至少包含 **能量饲料+蛋白质饲料+矿物质** 三类基础原料。")

    if animal_key in ("牛", "羊"):
        if "fiber" not in needed_types:
            rec_lines.append("- 🌿 **粗饲料**：反刍动物配方通常需要粗饲料，建议添加苜蓿干草/草粉、玉米青贮等")

    st.markdown("\n".join(rec_lines))

    # 快捷操作提示
    st.markdown("---")
    col_tip1, col_tip2 = st.columns(2)
    with col_tip1:
        st.info(
            "**快捷方案 A**：勾选上方「求解失败时自动从默认原料库补充」\n\n"
            "程序会自动从内置的 **548 种原料** 中挑选合适的来补充，无需手动录入。"
        )
    with col_tip2:
        st.info(
            "**快捷方案 B**：切换为「使用系统默认原料库(548种)」\n\n"
            "不使用自定义模板，直接用完整的 548 种原料库求解。\n"
            "成功后再对照结果调整自有原料库。"
        )


# ╔══════════════════════════════════════════════════════════════╗
# ║                   UI — 侧边栏                                 ║
# ╚══════════════════════════════════════════════════════════════╝

with st.sidebar:
    st.markdown('<div class="sidebar-title">🌾 饲料配方优化系统</div>',
                unsafe_allow_html=True)

    # ── 1. 数据上传 ────────────────────────────────────
    st.markdown('<p class="section-header">📥 数据文件</p>',
                unsafe_allow_html=True)
    with st.expander("📥 上传数据文件", expanded=False):
        st.caption("上传自定义原料或饲养标准文件，覆盖默认数据")

        # ── 下载模板按钮 ──
        st.caption("📋 不确定文件格式？先下载模板参考：")
        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            with open(os.path.join(BASE_DIR, "测试文件", "原料数据模板.xlsx"), "rb") as f_ing:
                st.download_button(
                    label="📥 原料模板",
                    data=f_ing,
                    file_name="原料数据模板.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_ing_template",
                    use_container_width=True,
                )
        with col_dl2:
            with open(os.path.join(BASE_DIR, "测试文件", "饲养标准模板.xlsx"), "rb") as f_std:
                st.download_button(
                    label="📥 标准模板",
                    data=f_std,
                    file_name="饲养标准模板.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_std_template",
                    use_container_width=True,
                )

        st.markdown("---")

        # 上传原料文件
        uploaded_ing_file = st.file_uploader(
            "上传原料数据文件（.xls/.xlsx）",
            type=["xls", "xlsx"],
            key="upload_ingredients",
            help="格式：第1列为原料名称，第2列为价格，后续列为营养成分"
        )
        if uploaded_ing_file is not None:
            try:
                uploaded_ing_df = pd.read_excel(uploaded_ing_file)

                # ── 列名标准化（兼容模板/数据库不同列名）──
                price_aliases = ["价格", "价格(元/kg)", "价格（元/kg）", "价格(元/公斤)", "price"]
                for alias in price_aliases:
                    if alias in uploaded_ing_df.columns and alias != "价格":
                        uploaded_ing_df = uploaded_ing_df.rename(columns={alias: "价格"})
                        break
                name_aliases = ["原料名称", "饲料原料名称", "原料名", "名称", "饲料原料", "ingredient"]
                for alias in name_aliases:
                    if alias in uploaded_ing_df.columns and alias != "原料名称":
                        uploaded_ing_df = uploaded_ing_df.rename(columns={alias: "原料名称"})
                        break
                ft_aliases = ["精粗类型", "饲料类型", "类型", "feed_type", "type"]
                for alias in ft_aliases:
                    if alias in uploaded_ing_df.columns and alias != "精粗类型":
                        uploaded_ing_df = uploaded_ing_df.rename(columns={alias: "精粗类型"})
                        break

                # 检查格式：第一列作为索引
                first_col = uploaded_ing_df.columns[0]
                if first_col != "原料名称":
                    uploaded_ing_df = uploaded_ing_df.rename(
                        columns={first_col: "原料名称"}
                    )
                uploaded_ing_df = uploaded_ing_df.set_index("原料名称")

                # 价格数值化（与 load_ingredients 行为一致）
                uploaded_ing_df["价格"] = pd.to_numeric(
                    uploaded_ing_df["价格"], errors="coerce"
                ).fillna(0)

                # 计算有效磷（与 load_ingredients 行为一致）
                from feed_formulation import _add_effective_phosphorus
                uploaded_ing_df = _add_effective_phosphorus(uploaded_ing_df)

                # ── 干物质%(DM)合理性校验 ──────────────────────
                if "干物质%" in uploaded_ing_df.columns:
                    dm_vals = pd.to_numeric(uploaded_ing_df["干物质%"], errors="coerce")
                    # 检测异常值：DM% < 10% 或 > 100%
                    bad_dm = []
                    for idx_name in uploaded_ing_df.index:
                        v = dm_vals.get(idx_name, None)
                        if pd.notna(v):
                            vf = float(v)
                            if vf < 5 or vf > 100:
                                bad_dm.append((idx_name, vf))
                        else:
                            bad_dm.append((idx_name, "空"))
                    if bad_dm:
                        st.warning(
                            f"⚠️ **以下原料的干物质%(DM)值异常**，可能导致营养数据计算错误：\n"
                            + "\n".join(f"  - {n}: DM%={v}" for n, v in bad_dm[:8])
                            + (f"\n... 等 {len(bad_dm)} 种" if len(bad_dm) > 8 else "")
                            + "\n\n正常范围应为 **10%~99%**（配合饲料通常88%，矿物质/添加剂95~99%）。"
                            "\n请检查上传文件中「干物质%」列的数据是否正确。"
                        )

                st.session_state.uploaded_ingredients_df = uploaded_ing_df
                st.success(f"✅ 已上传原料数据：{len(uploaded_ing_df)} 种原料")
            except Exception as e:
                st.error(f"❌ 原料文件读取失败：{e}")
                st.session_state.uploaded_ingredients_df = None
        else:
            # 用户清空了上传，恢复使用默认原料库
            if st.session_state.uploaded_ingredients_df is not None:
                st.session_state.uploaded_ingredients_df = None
                # 注意：不要在这里 st.rerun()，否则会与其他 rerun 冲突
                # 清除后下一次 load_all_data 会自动读到默认库

        st.markdown("")  # 空行

        # 上传饲养标准文件
        uploaded_std_file = st.file_uploader(
            "上传饲养标准文件（.xls/.xlsx）",
            type=["xls", "xlsx"],
            key="upload_standards",
            help="格式：每行一个标准，第一列为标准名称，后续列为营养指标需要量"
        )
        if uploaded_std_file is not None:
            # 立即保存为临时文件（避免文件对象缓存问题）
            try:
                import tempfile
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(uploaded_std_file.read())
                    tmp_std_path = tmp.name
                st.session_state.uploaded_standards_path = tmp_std_path

                # 检测多 sheet 格式
                import openpyxl
                wb = openpyxl.load_workbook(tmp_std_path, data_only=True)
                NON_DATA = {"填写说明", "说明", "Notes", "README"}
                data_sheets = [s for s in wb.sheetnames if s not in NON_DATA]
                if "饲养标准" in data_sheets or len(data_sheets) == 1:
                    # 单 sheet：flat list（向后兼容）
                    df = pd.read_excel(tmp_std_path, sheet_name=data_sheets[0] if data_sheets else 0)
                    first_col = df.columns[0]
                    if first_col != "Unnamed: 0":
                        names = df[first_col].dropna().astype(str).tolist()
                    else:
                        names = df.columns[1:].dropna().astype(str).tolist()
                    st.session_state.uploaded_standards_data = {data_sheets[0]: names}
                else:
                    # 多 sheet：按物种分组
                    uploaded_data = {}
                    for sn in data_sheets:
                        df = pd.read_excel(tmp_std_path, sheet_name=sn)
                        first_col = df.columns[0]
                        if first_col != "Unnamed: 0":
                            names = df[first_col].dropna().astype(str).tolist()
                        else:
                            names = []
                        if names:
                            uploaded_data[sn] = names
                    st.session_state.uploaded_standards_data = uploaded_data
                wb.close()

                total = sum(len(v) for v in st.session_state.uploaded_standards_data.values())
                st.success(f"✅ 已上传饲养标准：{total} 个标准（{len(st.session_state.uploaded_standards_data)} 个物种）")
            except Exception as e:
                st.error(f"❌ 标准文件读取失败：{e}")
                st.session_state.uploaded_standards_path = None
                st.session_state.uploaded_standards_data = {}
        else:
            # 用户清空了上传，恢复默认
            st.session_state.uploaded_standards_path = None
            st.session_state.uploaded_standards_data = {}

    # ── 数据加载 ────────────────────────────────────
    with st.spinner("正在加载数据..."):
        try:
            # 计算标准文件修改时间（用于缓存失效）
            _std_file_for_load = st.session_state.get("uploaded_standards_path") or STANDARDS_FILE
            _std_mtime_for_load = os.path.getmtime(_std_file_for_load) if os.path.exists(_std_file_for_load) else 0

            all_ingredients, standards_data = load_all_data(
                uploaded_ingredients_df=st.session_state.uploaded_ingredients_df,
                uploaded_standards_path=st.session_state.uploaded_standards_path,
                _standards_mtime=_std_mtime_for_load,
            )
            # 始终加载一份默认原料库（用于自动扩充时从全部548种原料中查找）
            if "default_ingredients" not in st.session_state:
                st.session_state.default_ingredients = load_ingredients(
                    INGREDIENTS_FILE,
                    phos_enabled=st.session_state.get("phos_enabled", True),
                    phos_custom_rates=st.session_state.get("phos_custom_rates") or None,
                    phos_per_ingredient=st.session_state.get("phos_per_ingredient") or None)
            # 若上传了标准文件，使用上传的标准数据
            if st.session_state.uploaded_standards_data:
                standards_data = st.session_state.uploaded_standards_data
            total_standards = sum(len(v) for v in standards_data.values())
            st.success(f"✅ 已加载 {len(all_ingredients)} 种原料, {total_standards} 个饲养标准")
        except Exception as e:
            st.error(f"❌ 数据加载失败: {e}")
            st.stop()

    # ── 2. 配方设置 ────────────────────────────────────
    st.markdown('<p class="section-header">🎯 配方设置</p>',
                unsafe_allow_html=True)

    # 检测是否为多 sheet 格式 → 显示物种选择器
    sheet_names = list(standards_data.keys())
    if len(sheet_names) > 1:
        selected_sheet = st.selectbox(
            "动物种类",
            options=sheet_names,
            index=0,
            help="选择要配方的动物种类，不同物种的营养指标不同"
        )
    else:
        selected_sheet = sheet_names[0]

    # 按选中物种过滤标准列表
    species_standards = standards_data.get(selected_sheet, [])
    if species_standards:
        selected_standard_name = st.selectbox(
            "饲养标准",
            options=species_standards,
            index=0,
            label_visibility="collapsed",
            help="从当前物种的饲养标准中选择"
        )
        st.session_state.selected_standard_name = selected_standard_name  # 供后续使用
        standard_index = species_standards.index(selected_standard_name)
    else:
        st.error("未找到饲养标准")
        st.stop()

    # ── 2.5. 预加载标准（供高级约束设置使用）──────
    # 仅在标准选择变化时重新加载
    _std_file = st.session_state.get("uploaded_standards_path") or STANDARDS_FILE
    _std_mtime = os.path.getmtime(_std_file) if os.path.exists(_std_file) else 0
    _load_key = (standard_index, selected_sheet, _std_file, _std_mtime)
    if st.session_state.get("_last_std_load_key") != _load_key:
        try:
            std_reqs, _ = load_single_standard(
                standard_index,
                uploaded_standards_path=st.session_state.uploaded_standards_path,
                sheet_name=selected_sheet if len(sheet_names) > 1 and selected_sheet != "饲养标准" else None,
                _file_mtime=_std_mtime,
            )
            st.session_state.current_standard_reqs = std_reqs
            st.session_state._last_std_load_key = _load_key
        except Exception:
            st.session_state.current_standard_reqs = {}
    preloaded_reqs = st.session_state.get("current_standard_reqs", {})

    # ── 饲养标准基础选择 ──
    # 默认：干物质基础(DM)，标准值直接使用
    # 风干饲粮基础：标准值需÷DM%转为DM基础
    standard_basis = st.radio(
        "饲养标准基础",
        options=["干物质基础(DM)", "风干饲粮基础"],
        index=0,
        horizontal=True,
        help="选择饲养标准给出的营养浓度基础。"
             "干物质基础：标准值已经是DM基础，直接使用；"
             "风干饲粮基础：标准值是风干基础，需输入DM%转换"
    )

    air_dried_dm_pct = 88.0  # 默认值
    if standard_basis == "风干饲粮基础":
        air_dried_dm_pct = st.number_input(
            "标准饲粮干物质含量(%)",
            min_value=10.0, max_value=100.0,
            value=88.0, step=1.0,
            help="风干饲粮的干物质含量，猪/禽配合饲料通常约88%，牛约55-65%，羊约60-70%"
        )

    # 原料始终转为DM基础进行优化
    use_dm = True

    st.markdown("---")

    # ── 3. 原料选择 ────────────────────────────────────
    st.markdown('<p class="section-header">🧂 原料选择</p>',
                unsafe_allow_html=True)

    selection_mode = st.radio(
        "选择模式",
        options=["使用全部可用原料", "手动选择原料"],
        index=0,
        label_visibility="collapsed",
        help="手动选择可以排除不需要的原料"
    )

    selected_ingredients = None
    auto_expand_from_upload = False  # 先从上传原料中未选择的补充
    auto_expand = False              # 再从默认548原料库补充
    min_bounds = {}
    max_bounds = {}

    # ── 自动补充选项（两种选择模式均显示）───
    _has_upload = st.session_state.uploaded_ingredients_df is not None
    with st.container():
        st.markdown("---")
        st.info(
            "💡 **智能补充**：若选中的原料无法满足全部营养约束，"
            "可勾选下方选项让程序自动从更多原料中补充。"
        )
        acol_a, acol_b = st.columns([1, 1])
        with acol_a:
            # "使用全部可用原料"模式下不显示①（已全部包含上传库，无意义）
            if _has_upload and selection_mode != "使用全部可用原料":
                auto_expand_from_upload = st.checkbox(
                    "① 从上传库未选原料补充",
                    value=True,
                    key="auto_expand_upload_cb",
                    help="求解失败时，从您上传的原料库中寻找未勾选的原料来补充",
                )
            elif _has_upload and selection_mode == "使用全部可用原料":
                # 全部可用模式下，①始终为False（不补充上传库，已全部包含）
                auto_expand_from_upload = False
                st.caption("ℹ️ 「使用全部可用原料」模式下，选项①自动忽略（已全部包含上传原料）")
            else:
                auto_expand_from_upload = False
        with acol_b:
            _step_label = "② " if _has_upload else ""
            auto_expand = st.checkbox(
                f"{_step_label}从默认原料库(548种)自动补充",
                value=False,
                key="auto_expand_default_cb",
                help="求解失败时，自动从系统内置的548种默认原料库中补充合适原料"
                if not _has_upload else
                "从548种默认原料库补充。若同时勾选①和②则使用合并池求解"
            )
        st.markdown("---")

    # 预加载默认原料库（供「使用全部可用原料」模式和自动补充使用）
    _default_df_preload = st.session_state.get("default_ingredients")
    if _default_df_preload is None:
        try:
            _default_df_preload = load_ingredients(
                INGREDIENTS_FILE,
                phos_enabled=st.session_state.get("phos_enabled", True),
                phos_custom_rates=st.session_state.get("phos_custom_rates") or None,
                phos_per_ingredient=st.session_state.get("phos_per_ingredient") or None)
            st.session_state.default_ingredients = _default_df_preload
        except Exception:
            _default_df_preload = None

    if selection_mode == "手动选择原料":
        # 候选原料池：上传原料 + 默认库合并
        # 显示时加标签区分来源：[上传] 玉米, [默认] 豆粕
        _has_upload = st.session_state.uploaded_ingredients_df is not None
        
        if _has_upload:
            # 合并：上传原料 + 默认库
            _default_df = st.session_state.get("default_ingredients")
            if _default_df is None:
                _default_df = load_ingredients(
                    INGREDIENTS_FILE,
                    phos_enabled=st.session_state.get("phos_enabled", True),
                    phos_custom_rates=st.session_state.get("phos_custom_rates") or None,
                    phos_per_ingredient=st.session_state.get("phos_per_ingredient") or None)
                st.session_state.default_ingredients = _default_df
            
            # 检测重名原料
            _up_names = set(all_ingredients.index)
            _df_names = set(_default_df.index)
            _duplicates = _up_names & _df_names
            if _duplicates:
                st.warning(
                    f"⚠️ 检测到 {len(_duplicates)} 种原料在上传库和默认库中重名："
                    f"{', '.join(sorted(_duplicates)[:5])}"
                    f"{'...' if len(_duplicates) > 5 else ''}。"
                    f"合并时将优先使用上传版本（上传优先）。"
                )
            
            # 对齐列：取两者共有的列
            _up_cols = set(all_ingredients.columns)
            _df_cols = set(_default_df.columns)
            _common = list(_up_cols & _df_cols)
            if not _common:
                _common = list(_up_cols | _df_cols)
            _up_sub = all_ingredients[[c for c in _common if c in all_ingredients.columns]]
            _df_sub = _default_df[[c for c in _common if c in _default_df.columns]]
            
            # 构建带标签的索引（用于显示）
            _up_sub_display = _up_sub.copy()
            _up_sub_display.index = ["[上传] " + n for n in _up_sub_display.index]
            _df_sub_display = _df_sub.copy()
            _df_sub_display.index = ["[默认] " + n for n in _df_sub_display.index]
            
            # 合并（上传优先：如果重名，保留上传版本）
            _merged_display = pd.concat([_up_sub_display, _df_sub_display])
            _merged_display = _merged_display[~_merged_display.index.duplicated(keep="first")]
            
            # 真实名称映射：显示名称 → 真实名称
            display_to_real = {}
            for display_name in _merged_display.index:
                if display_name.startswith("[上传] "):
                    display_to_real[display_name] = display_name[5:]
                elif display_name.startswith("[默认] "):
                    display_to_real[display_name] = display_name[5:]
                else:
                    display_to_real[display_name] = display_name
            
            selection_pool = _merged_display
            st.session_state.display_to_real_map = display_to_real
        else:
            selection_pool = all_ingredients
            st.session_state.display_to_real_map = {}  # 全部原料无需映射
        
        # 保存到 session_state，供求解时使用
        st.session_state.selection_pool = selection_pool
        all_names = list(selection_pool.index)
        # 初始化/保持之前的选中状态
        if "manual_selected" not in st.session_state:
            st.session_state.manual_selected = []

        # 搜索框
        search_query = st.text_input(
            "搜索原料",
            placeholder="输入关键词搜索...",
            label_visibility="collapsed"
        )
        if search_query:
            matched = [n for n in all_names if search_query in n]
        else:
            matched = all_names

        # filtered_names = 搜索匹配的 + 之前已选的（确保已选的不因搜索而消失）
        prev_selected = st.session_state.manual_selected
        filtered_set = set(matched) | set(prev_selected)
        filtered_names = sorted(filtered_set)

        selected_ingredients = st.multiselect(
            f"可选原料（共{len(filtered_names)}种，筛选{len(matched)}种）",
            options=filtered_names,
            default=prev_selected,
            key="manual_ingredient_select",
            help="勾选需要使用的原料。未勾选的原料不会进入配方。"
        )
        # 同步到 session_state（供其他逻辑使用）
        st.session_state.manual_selected = selected_ingredients

        # ── 用量限制表格（与选择合并）───
        if selected_ingredients:
            st.caption(f"已选择 {len(selected_ingredients)} 种原料，默认不限量（0~100%），可修改下方数值设置限制")

            # 初始化/同步 limit_pairs：以 selected_ingredients 为准
            if "limit_pairs" not in st.session_state:
                st.session_state.limit_pairs = []
            existing_map = {p["name"]: p for p in st.session_state.limit_pairs}

            # 新选中的自动加入，被移除的自动剔除
            current_set = set(selected_ingredients)
            synced_pairs = []
            for name in selected_ingredients:
                if name in existing_map:
                    synced_pairs.append(existing_map[name])
                else:
                    synced_pairs.append({"name": name, "min": 0.0, "max": 100.0})
            # 移除不再选择的
            removed = [n for n in existing_map if n not in current_set]
            if removed:
                st.session_state.limit_pairs = synced_pairs
                # 不立即 rerun，让用户看到变化即可

            st.markdown("**用量限制**（留空默认 0~100%）：")
            to_remove = []
            for i, pair in enumerate(synced_pairs):
                col_name, col_min, col_max, col_del = st.columns([3.5, 1.5, 1.5, 0.6])
                with col_name:
                    # 显示名称：去掉前缀标签，但保留来源信息
                    display_name = pair['name']
                    if display_name.startswith("[上传] "):
                        short_name = display_name[5:]  # 去掉 "[上传] "
                        st.markdown(f"**{short_name}** 📤")  # 用 emoji 表示来源
                    elif display_name.startswith("[默认] "):
                        short_name = display_name[5:]  # 去掉 "[默认] "
                        st.markdown(f"**{short_name}** 📚")  # 用 emoji 表示来源
                    else:
                        st.markdown(f"**{display_name}**")
                with col_min:
                    new_min = st.number_input(
                        "最小(%)", min_value=0.0, max_value=100.0,
                        value=float(pair["min"]), step=1.0,
                        key=f"limit_min_{pair['name']}",
                        label_visibility="collapsed",
                    )
                    pair["min"] = new_min
                with col_max:
                    new_max = st.number_input(
                        "最大(%)", min_value=0.0, max_value=100.0,
                        value=float(pair["max"]), step=1.0,
                        key=f"limit_max_{pair['name']}",
                        label_visibility="collapsed",
                    )
                    pair["max"] = new_max
                with col_del:
                    # 删除按钮 = 从选中列表移除该原料
                    if st.button("✕", key=f"limit_del_{pair['name']}", help="从配方中移除此原料"):
                        to_remove.append(pair['name'])

                if pair["min"] > 0:
                    min_bounds[pair["name"]] = pair["min"] / 100.0
                if pair["max"] < 100:
                    max_bounds[pair["name"]] = pair["max"] / 100.0

            # 将 min_bounds/max_bounds 的 key 从显示名（带 [上传]/[默认] 标签）转回真实名称
            # 因为后续求解/验证时 ing_df 的索引可能是真实名（upload_only模式）或显示名（merged模式）
            # 统一转为真实名可保证两种模式都正确
            _dtr = st.session_state.get("display_to_real_map")
            if _dtr:
                min_bounds = {_dtr.get(k, k): v for k, v in min_bounds.items()}
                max_bounds = {_dtr.get(k, k): v for k, v in max_bounds.items()}
            else:
                # 无标签映射时，手动去掉前缀
                def _strip_tag(name):
                    if name.startswith("[上传] ") or name.startswith("[默认] "):
                        return name[5:]
                    return name
                min_bounds = {_strip_tag(k): v for k, v in min_bounds.items()}
                max_bounds = {_strip_tag(k): v for k, v in max_bounds.items()}

            # 同步回 session_state
            st.session_state.limit_pairs = synced_pairs

            # 处理删除（从选中列表中移除）
            if to_remove:
                st.session_state.manual_selected = [
                    n for n in selected_ingredients if n not in to_remove
                ]
                st.rerun()

            # ── 高级设置：额外原料用量限制（expander 形式）───
            with st.expander("⚙️ 高级设置：原料用量限制"):
                st.caption("对特定原料设置最小/最大用量限制（可包含未选中的原料，自动补充时生效）")

                # 初始化 session state 中的额外用量限制数据
                if "extra_limit_pairs" not in st.session_state:
                    st.session_state.extra_limit_pairs = []  # [{"name": str, "min": float, "max": float}, ...]

                extra_candidates = list(selection_pool.index)

                col_e1, col_e2 = st.columns([3, 1])
                with col_e1:
                    existing_extra_names = {p["name"] for p in st.session_state.extra_limit_pairs}
                    available_extra = [n for n in sorted(extra_candidates) if n not in existing_extra_names]
                    # 搜索式选择
                    _extra_search = st.text_input(
                        "🔍 搜索原料",
                        key="extra_limit_search",
                        placeholder="输入关键词...",
                        label_visibility="collapsed",
                    )
                    if _extra_search.strip():
                        _ekw = _extra_search.strip().lower()
                        _available_filtered = [x for x in available_extra if _ekw in x.lower()]
                    else:
                        _available_filtered = available_extra[:50]
                    new_extra_ing = st.selectbox(
                        "添加用量限制",
                        options=[""] + _available_filtered,
                        format_func=lambda x: "— 选择原料 —" if x == "" else x,
                        key="new_extra_limit_select",
                        label_visibility="collapsed",
                    )
                with col_e2:
                    if st.button("➕ 添加", key="add_extra_limit_btn", use_container_width=True):
                        if new_extra_ing and new_extra_ing not in existing_extra_names:
                            st.session_state.extra_limit_pairs.append({
                                "name": new_extra_ing, "min": 0.0, "max": 100.0
                            })
                            st.rerun()

                # 显示当前额外用量限制列表
                if st.session_state.extra_limit_pairs:
                    st.markdown("---")
                    extra_to_remove = []
                    for i, pair in enumerate(st.session_state.extra_limit_pairs):
                        col_name, col_min, col_max, col_del = st.columns([3, 2, 2, 1])
                        with col_name:
                            st.markdown(f"**{pair['name']}**")
                        with col_min:
                            new_min = st.number_input(
                                "最小(%)", min_value=0.0, max_value=100.0,
                                value=float(pair["min"]), step=1.0,
                                key=f"extra_limit_min_{i}",
                                label_visibility="collapsed",
                            )
                            pair["min"] = new_min
                        with col_max:
                            new_max = st.number_input(
                                "最大(%)", min_value=0.0, max_value=100.0,
                                value=float(pair["max"]), step=1.0,
                                key=f"extra_limit_max_{i}",
                                label_visibility="collapsed",
                            )
                            pair["max"] = new_max
                        with col_del:
                            if st.button("✕", key=f"extra_limit_del_{i}"):
                                extra_to_remove.append(i)

                        # 将额外限制合并到 min_bounds / max_bounds
                        real_name = pair["name"]
                        # 如果是显示名（带标签），转回真实名
                        _dtr = st.session_state.get("display_to_real_map")
                        if _dtr and real_name in _dtr:
                            real_name = _dtr[real_name]
                        elif real_name.startswith("[上传] ") or real_name.startswith("[默认] "):
                            real_name = real_name[5:]

                        if pair["min"] > 0:
                            min_bounds[real_name] = pair["min"] / 100.0
                        if pair["max"] < 100:
                            max_bounds[real_name] = pair["max"] / 100.0

                    # 移除标记的项
                    if extra_to_remove:
                        for idx in sorted(extra_to_remove, reverse=True):
                            st.session_state.extra_limit_pairs.pop(idx)
                        st.rerun()

                    if st.button("🗑 清空全部限制", key="clear_all_extra_limits"):
                        st.session_state.extra_limit_pairs = []
                        st.rerun()
                else:
                    st.info("尚未添加任何用量限制，可从上方下拉框选择原料添加")

        else:
            st.warning("未选择任何原料，将使用全部可用原料")
            selected_ingredients = None
            # 注意：自动补充复选框已统一移至 selection_mode 判断之前（L1731），无需在此重复定义

    else:
        # 全部原料模式：保留原有的高级设置 expander 作为用量限制入口
        with st.expander("⚙️ 高级设置：原料用量限制"):
            st.caption("设置特定原料的最小/最大用量（百分比，0~100%）")

            # 初始化 session state 中的用量限制数据
            if "limit_pairs" not in st.session_state:
                st.session_state.limit_pairs = []  # [{"name": str, "min": float, "max": float}, ...]

            # 构建候选列表：带标签的合并池（上传+默认548库），与主界面手动选择一致
            _has_upload_for_limit = st.session_state.uploaded_ingredients_df is not None
            _default_for_limit = st.session_state.get("default_ingredients")
            if _has_upload_for_limit and _default_for_limit is not None:
                _up_cols_set_l = set(all_ingredients.columns)
                _df_cols_set_l = set(_default_for_limit.columns)
                _common_l = list(_up_cols_set_l & _df_cols_set_l)
                if not _common_l:
                    _common_l = list(_up_cols_set_l | _df_cols_set_l)
                _up_sub_l = all_ingredients[[c for c in _common_l if c in all_ingredients.columns]]
                _df_sub_l = _default_for_limit[[c for c in _common_l if c in _default_for_limit.columns]]
                _up_display_l = _up_sub_l.copy()
                _up_display_l.index = ["[上传] " + n for n in _up_display_l.index]
                _df_display_l = _df_sub_l.copy()
                _df_display_l.index = ["[默认] " + n for n in _df_display_l.index]
                _merged_limit_pool = pd.concat([_up_display_l, _df_display_l])
                _merged_limit_pool = _merged_limit_pool[~_merged_limit_pool.index.duplicated(keep="first")]
                limit_candidates = list(_merged_limit_pool.index)
            else:
                limit_candidates = list(all_ingredients.index)

            # 添加用量限制
            col_add1, col_add2 = st.columns([3, 1])
            with col_add1:
                existing_names = {p["name"] for p in st.session_state.limit_pairs}
                available_for_limit = [n for n in limit_candidates if n not in existing_names]
                # 搜索式选择
                _all_limit_search = st.text_input(
                    "🔍 搜索原料",
                    key="all_limit_search",
                    placeholder="输入关键词...",
                    label_visibility="collapsed",
                )
                if _all_limit_search.strip():
                    _akw = _all_limit_search.strip().lower()
                    _avail_filtered = [x for x in available_for_limit if _akw in x.lower()]
                else:
                    _avail_filtered = sorted(available_for_limit)[:50]
                new_limit_ing = st.selectbox(
                    "添加用量限制",
                    options=[""] + _avail_filtered,
                    format_func=lambda x: "— 选择原料 —" if x == "" else x,
                    key="new_limit_select_all",
                    label_visibility="collapsed",
                )
            with col_add2:
                if st.button("➕ 添加", key="add_limit_btn", use_container_width=True):
                    if new_limit_ing and new_limit_ing not in existing_names:
                        st.session_state.limit_pairs.append({
                            "name": new_limit_ing, "min": 0.0, "max": 100.0
                        })
                        st.rerun()

            # 显示当前用量限制列表
            if st.session_state.limit_pairs:
                st.markdown("---")
                st.caption("当前用量限制：")
                to_remove_idx = []
                for i, pair in enumerate(st.session_state.limit_pairs):
                    col_name, col_min, col_max, col_del = st.columns([3, 2, 2, 1])
                    with col_name:
                        st.markdown(f"**{pair['name']}**")
                    with col_min:
                        new_min = st.number_input(
                            "最小(%)", min_value=0.0, max_value=100.0,
                            value=float(pair["min"]), step=1.0,
                            key=f"limit_min_{i}",
                            label_visibility="collapsed",
                        )
                        pair["min"] = new_min
                    with col_max:
                        new_max = st.number_input(
                            "最大(%)", min_value=0.0, max_value=100.0,
                            value=float(pair["max"]), step=1.0,
                            key=f"limit_max_{i}",
                            label_visibility="collapsed",
                        )
                        pair["max"] = new_max
                    with col_del:
                        if st.button("✕", key=f"limit_del_{i}"):
                            to_remove_idx.append(i)

                    if pair["min"] > 0:
                        min_bounds[pair["name"]] = pair["min"] / 100.0
                    if pair["max"] < 100:
                        max_bounds[pair["name"]] = pair["max"] / 100.0

                # 移除标记的项
                if to_remove_idx:
                    for idx in sorted(to_remove_idx, reverse=True):
                        st.session_state.limit_pairs.pop(idx)
                    st.rerun()

                # 一键清空
                if st.button("🗑 清空全部限制", key="clear_all_limits"):
                    st.session_state.limit_pairs = []
                    st.rerun()

            # 注意：自动补充复选框已统一移至 selection_mode 判断之前（L1731），无需在此重复定义

    # ── 4.5. 高级约束设置 ──────────────────────────
    with st.expander("🛡 高级约束设置"):
        st.caption("设置营养上限、精粗比、钙磷比等额外约束（可选）")

        # 初始化 session state
        if "nutrient_max_settings" not in st.session_state:
            st.session_state.nutrient_max_settings = {}  # {营养指标: max_value}
        if "cf_ratio_settings" not in st.session_state:
            st.session_state.cf_ratio_settings = {
                "concentrate_min": None, "concentrate_max": None,
                "forage_min": None, "forage_max": None,
            }
        if "ca_p_settings" not in st.session_state:
            st.session_state.ca_p_settings = {"ca_p_min": None, "ca_p_max": None}

        # ── 有效磷处理逻辑 ──────────────────────────
        # 检测原料数据中的有效磷/总磷情况，决定默认行为
        if "phos_enabled" not in st.session_state:
            st.session_state.phos_enabled = False  # 默认不启用
        if "phos_custom_rates" not in st.session_state:
            st.session_state.phos_custom_rates = {}
        if "phos_auto_detected" not in st.session_state:
            st.session_state.phos_auto_detected = None

        _has_eff_p_col = ("有效磷%" in all_ingredients.columns) if all_ingredients is not None else False
        _has_total_p_col = ("总磷%" in all_ingredients.columns) if all_ingredients is not None else False
        _eff_p_has_values = False
        if _has_eff_p_col and all_ingredients is not None:
            _ep_vals = pd.to_numeric(all_ingredients["有效磷%"], errors="coerce")
            _eff_p_has_values = _ep_vals.notna().any()

        # 自动检测并设置状态（只在首次检测时）
        if st.session_state.phos_auto_detected is None:
            if _eff_p_has_values and _has_total_p_col:
                # 原料已有有效磷数据 → 不启用换算
                st.session_state.phos_enabled = False
                st.session_state.phos_auto_detected = "has_data"
            elif _has_total_p_col and not _eff_p_has_values:
                # 有总磷但无有效磷 → 默认不启用，用户可手动开启
                st.session_state.phos_enabled = False
                st.session_state.phos_auto_detected = "need_calc"
            elif _eff_p_has_values and not _has_total_p_col:
                # 只有有效磷没有总磷 → 直接使用有效磷
                st.session_state.phos_enabled = False
                st.session_state.phos_auto_detected = "eff_only"
            else:
                st.session_state.phos_enabled = False
                st.session_state.phos_auto_detected = "none"

        # 显示有效磷状态提示 + 控制开关
        _phos_status_msg = ""
        if st.session_state.phos_auto_detected == "has_data":
            _phos_status_msg = "ℹ️ 原料数据中已包含「有效磷%」列且有数值，将直接使用该列参与约束。无需启用换算。"
        elif st.session_state.phos_auto_detected == "need_calc":
            _phos_status_msg = "ℹ️ 原料数据中有「总磷%」但无有效磷数据。可勾选下方开关启用总磷→有效磷换算。"
        elif st.session_state.phos_auto_detected == "eff_only":
            _phos_status_msg = "ℹ️ 原料数据中只有「有效磷%」无「总磷%」，将直接以有效磷值参与约束。"
        else:
            _phos_status_msg = "⚠️ 原料数据中未检测到磷相关数据。"

        st.markdown("##### ⚙️ 有效磷处理")
        st.caption(_phos_status_msg)

        phos_on = st.checkbox(
            "启用总磷→有效磷生物利用率换算",
            value=st.session_state.phos_enabled,
            key="phos_enable_checkbox",
            help="勾选后，程序按原料类别自动将「总磷×利用率」折算为有效磷参与约束；不勾选则直接使用原料中的有效磷值（若有）或总磷值",
        )
        st.session_state.phos_enabled = phos_on

        if phos_on:
            with st.expander("📐 自定义有效磷利用率（可选，留空用默认值）"):
                st.caption("按大类设置默认值；下方可对每种原料单独设置（单独设置优先于大类值）")

                # ── 第一层：大类默认值 ──
                _ph_cols = st.columns(2)
                with _ph_cols[0]:
                    _cereal_rate = st.number_input(
                        "谷类利用率(%)", min_value=0.0, max_value=100.0,
                        value=st.session_state.phos_custom_rates.get("cereal", 30.0),
                        step=1.0, format="%.0f", key="phos_cereal",
                    )
                    _oilmeal_rate = st.number_input(
                        "粕类利用率(%)", min_value=0.0, max_value=100.0,
                        value=st.session_state.phos_custom_rates.get("oil_meal", 50.0),
                        step=1.0, format="%.0f", key="phos_oilmeal",
                    )
                with _ph_cols[1]:
                    _animal_rate = st.number_input(
                        "动物性利用率(%)", min_value=0.0, max_value=100.0,
                        value=st.session_state.phos_custom_rates.get("animal", 70.0),
                        step=1.0, format="%.0f", key="phos_animal",
                    )
                    _other_rate = st.number_input(
                        "其他利用率(%)", min_value=0.0, max_value=100.0,
                        value=st.session_state.phos_custom_rates.get("other", 100.0),
                        step=1.0, format="%.0f", key="phos_other",
                    )
                # 收集大类自定义值
                _defaults = {"cereal": 30, "oil_meal": 50, "animal": 70, "other": 100}
                _custom = {}
                for _k, _v in [("cereal", _cereal_rate), ("oil_meal", _oilmeal_rate),
                               ("animal", _animal_rate), ("other", _other_rate)]:
                    if abs(_v - _defaults[_k]) > 0.001:
                        _custom[_k] = _v / 100.0
                st.session_state.phos_custom_rates = _custom

                st.markdown("---")

                # ── 第二层：逐个原料设置 ──
                if "phos_per_ingredient" not in st.session_state:
                    st.session_state.phos_per_ingredient = {}  # {原料名: 利用率%}

                if _has_total_p_col and all_ingredients is not None:
                    st.markdown("**逐个原料设置有效磷利用率**（仅列出需要换算的原料：有总磷%但无有效磷%）")
                    st.caption(
                        "以下原料含有「总磷%」但缺少「有效磷%」数据，需要设置利用率来换算。"
                        "已有有效磷数据的原料无需设置，将直接使用原值。留空 = 使用上方大类默认值。"
                    )

                    # 只筛选"有总磷但无有效磷（或有效磷为空）"的原料
                    _tp_ing_list = []
                    for _name in all_ingredients.index:
                        _tp_val = pd.to_numeric(all_ingredients.at[_name, "总磷%"], errors="coerce")
                        if pd.notna(_tp_val) and float(_tp_val) > 0:
                            # 检查该原料是否已有有效磷数据
                            _needs_setting = True
                            if _has_eff_p_col:
                                _ep_val = pd.to_numeric(
                                    all_ingredients.at[_name, "有效磷%"], errors="coerce"
                                )
                                if pd.notna(_ep_val) and float(_ep_val) > 0:
                                    _needs_setting = False  # 已有有效磷，不需要设置
                            if _needs_setting:
                                _tp_ing_list.append(_name)
                    _tp_ing_list.sort()

                    if _tp_ing_list:
                        st.caption(f"共 {len(_tp_ing_list)} 种原料需要设置有效磷利用率")
                        # 分两列显示
                        _half = (len(_tp_ing_list) + 1) // 2
                        _col_left, _col_right = st.columns([1, 1])
                        with _col_left:
                            for i, _ing_name in enumerate(_tp_ing_list[:_half]):
                                _cur_val = st.session_state.phos_per_ingredient.get(_ing_name, None)
                                # 显示当前大类推断值作为占位提示
                                _hint = ""
                                if _cur_val is None:
                                    _n_lower = str(_ing_name).lower()
                                    if any(kw in _n_lower for kw in ["玉米", "小麦", "稻", "米", "高粱"]):
                                        _hint = f" (默认{_cereal_rate:.0f}%)"
                                    elif any(kw in _n_lower for kw in ["粕", "豆", "菜籽", "棉籽"]):
                                        _hint = f" (默认{_oilmeal_rate:.0f}%)"
                                    elif any(kw in _n_lower for kw in ["鱼", "肉骨", "血", "羽毛", "蚕蛹"]):
                                        _hint = f" (默认{_animal_rate:.0f}%)"
                                    else:
                                        _hint = f" (默认{_other_rate:.0f}%)"

                                _new_val = st.number_input(
                                    f"{_ing_name}{_hint}",
                                    min_value=0.0, max_value=100.0,
                                    value=float(_cur_val) if _cur_val is not None else 0.0,
                                    step=1.0, format="%.0f",
                                    key=f"phos_ing_{i}_{_ing_name}",
                                )
                                # 0 表示留空（使用默认），非0才记录
                                if _new_val > 0:
                                    st.session_state.phos_per_ingredient[_ing_name] = _new_val
                                elif _ing_name in st.session_state.phos_per_ingredient:
                                    del st.session_state.phos_per_ingredient[_ing_name]

                        with _col_right:
                            for i, _ing_name in enumerate(_tp_ing_list[_half:]):
                                _cur_val = st.session_state.phos_per_ingredient.get(_ing_name, None)
                                _hint = ""
                                if _cur_val is None:
                                    _n_lower = str(_ing_name).lower()
                                    if any(kw in _n_lower for kw in ["玉米", "小麦", "稻", "米", "高粱"]):
                                        _hint = f" (默认{_cereal_rate:.0f}%)"
                                    elif any(kw in _n_lower for kw in ["粕", "豆", "菜籽", "棉籽"]):
                                        _hint = f" (默认{_oilmeal_rate:.0f}%)"
                                    elif any(kw in _n_lower for kw in ["鱼", "肉骨", "血", "羽毛", "蚕蛹"]):
                                        _hint = f" (默认{_animal_rate:.0f}%)"
                                    else:
                                        _hint = f" (默认{_other_rate:.0f}%)"

                                _new_val = st.number_input(
                                    f"{_ing_name}{_hint}",
                                    min_value=0.0, max_value=100.0,
                                    value=float(_cur_val) if _cur_val is not None else 0.0,
                                    step=1.0, format="%.0f",
                                    key=f"phos_ing_{_half+i}_{_ing_name}",
                                )
                                if _new_val > 0:
                                    st.session_state.phos_per_ingredient[_ing_name] = _new_val
                                elif _ing_name in st.session_state.phos_per_ingredient:
                                    del st.session_state.phos_per_ingredient[_ing_name]
                    else:
                        st.caption("✅ 所有含「总磷%」的原料都已有有效磷数据，无需单独设置利用率")
                else:
                    st.caption("原料数据中无「总磷%」列")

        # NPN 原料控制
        if "exclude_npn" not in st.session_state:
            st.session_state.exclude_npn = True

        # 从已选饲养标准名称判断动物类型（animal_key_map 在后面才定义）
        _std_name_for_npn = str(st.session_state.get("selected_standard_name", "") or "")
        _is_ruminant_npn = any(kw in _std_name_for_npn for kw in ["牛", "羊"])
        _show_npn_option = (
            not _is_ruminant_npn
            or st.session_state.get("_always_show_npn", False)
        )
        if _show_npn_option:
            _npn_exclude = st.checkbox(
                "排除非蛋白氮(NPN)原料（尿素等）",
                value=st.session_state.exclude_npn,
                key="npn_exclude_checkbox",
                help="✅ 勾选 = 配方中排除尿素等NPN原料（单胃动物无法利用NPN）；"
                     "☐ 取消勾选 = 允许NPN原料参与配方计算。",
            )
            st.session_state.exclude_npn = _npn_exclude
        else:
            st.caption("ℹ️ 当前动物类型为反刍动物，NPN原料可用")
            st.session_state.exclude_npn = False

        st.markdown("---")

        # ── 营养上限 ─────────────────────────────────
        st.markdown("##### 📊 营养上限约束")
        st.caption("为选定的营养指标设置含量上限。留空 = 不限制。")

        available_nutrients = sorted(preloaded_reqs.keys()) if preloaded_reqs else []
        # 同步已选择的营养指标（可能因标准切换而变化）
        current_max_keys = set(st.session_state.nutrient_max_settings.keys())
        valid_max_keys = current_max_keys & set(available_nutrients)
        # 清理无效键
        invalid_keys = current_max_keys - set(available_nutrients)
        for k in invalid_keys:
            del st.session_state.nutrient_max_settings[k]

        if available_nutrients:
            selected_nutrients = st.multiselect(
                "选择需要设置上限的营养指标",
                options=available_nutrients,
                default=sorted(valid_max_keys),
                help="从选定饲养标准中勾选需要上限约束的营养指标",
                key="nutrient_max_select",
            )
            # 为每个选中的指标显示上限输入
            if selected_nutrients:
                for i, nut in enumerate(selected_nutrients):
                    col_n1, col_n2 = st.columns([3, 1])
                    with col_n1:
                        st.markdown(f"**{nut}** ≤ ")
                    with col_n2:
                        cur_val = st.session_state.nutrient_max_settings.get(nut, None)
                        max_val = st.number_input(
                            f"上限_{nut}",
                            min_value=0.0,
                            value=float(cur_val) if cur_val is not None else 0.0,
                            step=0.01,
                            format="%.4f",
                            key=f"nut_max_{nut}",
                            label_visibility="collapsed",
                        )
                        if max_val > 0:
                            st.session_state.nutrient_max_settings[nut] = max_val
                        elif nut in st.session_state.nutrient_max_settings:
                            del st.session_state.nutrient_max_settings[nut]
                # 清除未选中的指标
                for nut in list(st.session_state.nutrient_max_settings.keys()):
                    if nut not in selected_nutrients:
                        del st.session_state.nutrient_max_settings[nut]
        else:
            st.info("请先在左侧选择饲养标准")

        # ── 抗营养因子上限 ─────────────────────────
        st.markdown("---")
        st.markdown("##### 🧪 抗营养因子上限约束")
        st.caption(
            "选择饲料原料中的抗营养因子指标（如黄曲霉素μg/kg），"
            "设置配方中该因子的总含量上限。留空 = 不限制。"
        )

        # 初始化 session state
        if "anti_factor_max_settings" not in st.session_state:
            st.session_state.anti_factor_max_settings = {}  # {列名: max_value}

        # 从原料数据中收集所有可作为抗营养因子上限的数值列
        _anti_candidates = []
        _anti_priority = []   # 含关键词的（排前面）
        _anti_other = []      # 其他数值列（排后面）
        if all_ingredients is not None:
            _af_keywords = ["μg", "mg", "毒素", "霉素", "toxin", "afla", "deox", "zearalenone", "ochra"]
            _exclude = {"原料名称", "价格(元/kg)", "精粗类型", "干物质%"}
            for col in all_ingredients.columns:
                if col in _exclude:
                    continue
                _vals = pd.to_numeric(all_ingredients[col], errors="coerce")
                if not _vals.notna().any():
                    continue  # 全为非数值，跳过
                col_s = str(col)
                if any(kw in col_s.lower() for kw in _af_keywords):
                    _anti_priority.append(col)
                else:
                    _anti_other.append(col)
            # 合并：疑似抗营养因子排前面，其他数值列排后面
            _anti_candidates = sorted(_anti_priority) + sorted(_anti_other)

        if _anti_candidates:
            _selected_factors = st.multiselect(
                "选择抗营养因子指标",
                options=sorted(_anti_candidates),
                default=sorted(st.session_state.anti_factor_max_settings.keys()),
                help="选择饲料原料中含有的抗营养因子列（数据来自原料文件）",
                key="anti_factor_select",
            )

            # 为每个选中的因子显示上限输入
            if _selected_factors:
                for _af in _selected_factors:
                    _col1, _col2 = st.columns([3, 1])
                    with _col1:
                        st.markdown(f"**{_af}** ≤")
                    with _col2:
                        _cur = st.session_state.anti_factor_max_settings.get(_af, None)
                        _val = st.number_input(
                            f"上限_{_af}",
                            min_value=0.0,
                            value=float(_cur) if _cur is not None else 0.0,
                            step=0.1,
                            format="%.4f",
                            key=f"anti_factor_max_{_af}",
                            label_visibility="collapsed",
                        )
                        if _val > 0:
                            st.session_state.anti_factor_max_settings[_af] = _val
                        elif _af in st.session_state.anti_factor_max_settings:
                            del st.session_state.anti_factor_max_settings[_af]

            # 清除未选中的因子
            for _af in list(st.session_state.anti_factor_max_settings.keys()):
                if _af not in _selected_factors:
                    del st.session_state.anti_factor_max_settings[_af]
        else:
            st.info("饲料原料数据中未检测到抗营养因子相关列。请在原料数据中添加如「黄曲霉素μg/kg」等列。")

        st.markdown("---")

        # ── 精粗比约束 ────────────────────────
        st.markdown("##### 🌿 精粗比约束")
        st.caption("设置精料/粗料在配方中的占比范围（留空 = 不限制）")

        # 检查原料中是否有精粗类型标签
        has_feed_type_col = "精粗类型" in all_ingredients.columns
        cf_warn_missing = False
        if st.session_state.cf_ratio_settings.get("concentrate_min") is not None or \
           st.session_state.cf_ratio_settings.get("concentrate_max") is not None or \
           st.session_state.cf_ratio_settings.get("forage_min") is not None or \
           st.session_state.cf_ratio_settings.get("forage_max") is not None:
            if not has_feed_type_col:
                feed_type_dict = load_feed_type_classification()
                unclassified = 0
                for name in all_ingredients.index:
                    ft = None
                    if name in feed_type_dict:
                        ft = feed_type_dict[name]
                    else:
                        from feed_formulation import classify_ingredient
                        ft = classify_ingredient(name, feed_type_dict)
                    if ft not in ("精料", "粗料"):
                        unclassified += 1
                if unclassified > 0:
                    cf_warn_missing = True

        # 统一样式：每行「指标名 | 最小 | 最大」
        _cf_label_w, _cf_min_w, _cf_max_w = st.columns([2, 1, 1])
        with _cf_label_w:
            st.markdown("**指标**")
        with _cf_min_w:
            st.markdown("**最小(%)**")
        with _cf_max_w:
            st.markdown("**最大(%)**")

        # 精料占比行
        _r1_l, _r1_min, _r1_max = st.columns([2, 1, 1])
        with _r1_l:
            st.markdown("精料占比")
        with _r1_min:
            conc_min = st.number_input("精料min", min_value=0.0, max_value=100.0,
                value=float(st.session_state.cf_ratio_settings["concentrate_min"] * 100)
                    if st.session_state.cf_ratio_settings["concentrate_min"] is not None else 0.0,
                step=1.0, key="cf_conc_min", label_visibility="collapsed",
                help="精料占总配方的比例下限")
            st.session_state.cf_ratio_settings["concentrate_min"] = conc_min / 100.0 if conc_min > 0 else None
        with _r1_max:
            conc_max = st.number_input("精料max", min_value=0.0, max_value=100.0,
                value=float(st.session_state.cf_ratio_settings["concentrate_max"] * 100)
                    if st.session_state.cf_ratio_settings["concentrate_max"] is not None else 0.0,
                step=1.0, key="cf_conc_max", label_visibility="collapsed",
                help="精料占总配方的比例上限")
            st.session_state.cf_ratio_settings["concentrate_max"] = conc_max / 100.0 if conc_max > 0 else None

        # 粗料占比行
        _r2_l, _r2_min, _r2_max = st.columns([2, 1, 1])
        with _r2_l:
            st.markdown("粗料占比")
        with _r2_min:
            for_min = st.number_input("粗料min", min_value=0.0, max_value=100.0,
                value=float(st.session_state.cf_ratio_settings["forage_min"] * 100)
                    if st.session_state.cf_ratio_settings["forage_min"] is not None else 0.0,
                step=1.0, key="cf_for_min", label_visibility="collapsed",
                help="粗料占总配方的比例下限")
            st.session_state.cf_ratio_settings["forage_min"] = for_min / 100.0 if for_min > 0 else None
        with _r2_max:
            for_max = st.number_input("粗料max", min_value=0.0, max_value=100.0,
                value=float(st.session_state.cf_ratio_settings["forage_max"] * 100)
                    if st.session_state.cf_ratio_settings["forage_max"] is not None else 0.0,
                step=1.0, key="cf_for_max", label_visibility="collapsed",
                help="粗料占总配方的比例上限")
            st.session_state.cf_ratio_settings["forage_max"] = for_max / 100.0 if for_max > 0 else None

        if cf_warn_missing:
            st.warning(
                "⚠️ 已设置精粗比约束，但部分原料的「精粗类型」未标注。"
                "请在原料模板的「精粗类型」列中为每个原料标注「精料」或「粗料」，"
                "否则软件通过关键词自动分类可能不准确。"
            )

        st.markdown("---")

        # ── 钙磷比约束 ─────────────────────────────────
        st.markdown("##### 🦴 钙磷比约束")
        st.caption("设置钙与总磷的比例范围（如 1.2~2.0:1，留空 = 不限制）")

        # 统一表头样式：与精粗比一致
        _cap_label_w, _cap_min_w, _cap_max_w = st.columns([2, 1, 1])
        with _cap_label_w:
            st.markdown("**指标**")
        with _cap_min_w:
            st.markdown("**最小值**")
        with _cap_max_w:
            st.markdown("**最大值**")

        _cap_l, _cap_min_c, _cap_max_c = st.columns([2, 1, 1])
        with _cap_l:
            st.markdown("Ca : P 比值")
        with _cap_min_c:
            ca_p_min = st.number_input("CaPmin", min_value=0.0, max_value=10.0,
                value=float(st.session_state.ca_p_settings["ca_p_min"])
                    if st.session_state.ca_p_settings["ca_p_min"] is not None else 0.0,
                step=0.1, format="%.1f", key="cap_min", label_visibility="collapsed",
                help="如 1.2 表示 Ca:P ≥ 1.2:1")
            st.session_state.ca_p_settings["ca_p_min"] = ca_p_min if ca_p_min > 0 else None
        with _cap_max_c:
            ca_p_max = st.number_input("CaPmax", min_value=0.0, max_value=10.0,
                value=float(st.session_state.ca_p_settings["ca_p_max"])
                    if st.session_state.ca_p_settings["ca_p_max"] is not None else 0.0,
                step=0.1, format="%.1f", key="cap_max", label_visibility="collapsed",
                help="如 2.0 表示 Ca:P ≤ 2.0:1")
            st.session_state.ca_p_settings["ca_p_max"] = ca_p_max if ca_p_max > 0 else None

    # ── 5. 求解按钮 ────────────────────────────────────
    solve_btn = st.button(
        "🚀 开始求解最低成本配方",
        type="primary",
        use_container_width=True,
    )


# ╔══════════════════════════════════════════════════════════════╗
# ║                   UI — 主区域                                 ║
# ╚══════════════════════════════════════════════════════════════╝

# ── 标题栏 ─────────────────────────────────────────────
ing_count = len(all_ingredients)
total_std_count = sum(len(v) for v in standards_data.values())
species_count = len(standards_data)
st.markdown(f"""
<div class="app-header">
    <h1>🌾 线性规划饲料配方优化系统</h1>
    <div class="subtitle">基于线性规划的最低成本饲料配方求解工具 | 支持猪/禽/牛/羊</div>
    <div class="stats">
        <span class="stat-item">📦 {ing_count} 种原料</span>
        <span class="stat-item">📋 {total_std_count} 个饲养标准</span>
        <span class="stat-item">🐾 {species_count} 种动物</span>
    </div>
</div>
""", unsafe_allow_html=True)

# ── 使用教程（页面顶部，可折叠）──────────────────────
with st.expander("📖 使用教程", expanded=False):
    col_t1, col_t2, col_t3 = st.columns(3)
    with col_t1:
        st.markdown("""
**第一步：选择配方参数**
- 在左侧侧边栏选择**动物种类**和**饲养标准**
- 选择**饲养标准基础**（干物质/风干饲粮）
- 选择原料范围（全部/手动）
""")
    with col_t2:
        st.markdown("""
**第二步：设置用量限制（可选）**
- 展开「高级设置」
- 添加原料用量上下限
- 如玉米不低于30%、不高于70%
""")
    with col_t3:
        st.markdown("""
**第三步：求解与导出**
- 点击「开始求解」按钮
- 查看配方组成、营养对比
- 导出结果为 Excel 文件
""")
    st.markdown("---")
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        st.markdown("""
##### 📂 数据文件格式

| 文件 | 必填列 |
|------|--------|
| 原料数据 | 第1列：原料名称，第2列：价格 |
| 饲养标准 | 第1列：标准名称，按物种分Sheet |

> 💡 不知道格式？侧边栏 **「📥 上传数据文件」** 中有模板下载按钮。
""")
    with col_f2:
        st.markdown("""
##### ❓ 常见问题

**Q: 求解失败/无解？**
A: 放宽某些营养约束，或增加原料种类。

**Q: 上传文件后没反应？**
A: 确保第1列为原料名/标准名，列名与模板一致。

**Q: 如何添加禽的能量数据？**
A: 原料文件中添加「禽代谢能MC/Kg」列即可。
""")

# ── 状态显示 ─────────────────────────────────────────────
if "result" not in st.session_state:
    st.session_state.result = None
if "solved" not in st.session_state:
    st.session_state.solved = False

if solve_btn:
    with st.status("🔍 正在求解...", expanded=True) as status:
        st.write("加载饲养标准数据...")

        # 保存用户选择到 result 中
        st.session_state.user_selected_ingredients = selected_ingredients
        st.session_state.user_auto_expand_from_upload = auto_expand_from_upload
        st.session_state.user_auto_expand = auto_expand

        # 确定使用哪个标准文件
        standards_file_to_use = (
            st.session_state.uploaded_standards_path
            if st.session_state.uploaded_standards_path
            else STANDARDS_FILE
        )
        _std_mtime = os.path.getmtime(standards_file_to_use) if os.path.exists(standards_file_to_use) else 0

        # 加载标准
        nutrient_requirements, standard_name = load_single_standard(
            standard_index,
            uploaded_standards_path=st.session_state.uploaded_standards_path,
            sheet_name=selected_sheet if len(sheet_names) > 1 and selected_sheet != "饲养标准" else None,
            _file_mtime=_std_mtime,
        )

        if not nutrient_requirements:
            status.update(label="❌ 求解失败", state="error")
            st.error("饲养标准为空，请检查数据文件")
            st.stop()

        st.write(f"标准: **{standard_name}**")
        st.write(f"营养约束数: {len(nutrient_requirements)}")

        # ── 确定动物种类（需在DM转换前） ──
        animal_key_map = "禽"
        if any(kw in standard_name for kw in ["猪", "仔猪", "母猪"]):
            animal_key_map = "猪"
        elif any(kw in standard_name for kw in ["牛", "奶牛", "肉牛"]):
            animal_key_map = "牛"
        elif any(kw in standard_name for kw in ["羊"]):
            animal_key_map = "羊"
        elif any(kw in standard_name for kw in ["鸡", "鸭", "鹅", "禽", "肉鸡", "蛋鸡"]):
            animal_key_map = "禽"
        elif any(kw in standard_name for kw in ["猫", "狗", "马", "鱼", "其他"]):
            animal_key_map = "其他"

        # ── 饲养标准基础转换 ──
        # 干物质基础(DM)：标准值已是DM基础，直接使用，无需转换
        # 风干饲粮基础：标准值需÷DM%转为DM基础
        if standard_basis == "风干饲粮基础":
            nutrient_requirements = convert_requirements_to_dm(
                nutrient_requirements, typical_dm_pct=air_dried_dm_pct
            )
            st.info(f"📐 饲养标准已从风干基础转为干物质基础(DM)（DM%={air_dried_dm_pct}%）")
            st.write(f"DM基础营养约束数: {len(nutrient_requirements)}")
        else:
            st.info("📐 饲养标准为干物质基础(DM)，直接使用")
            st.write(f"DM基础营养约束数: {len(nutrient_requirements)}")

        # 干物质转换
        # 根据自动补充选项决定使用哪个原料池（三模式）：
        #   - 仅勾选①（有上传时）→ upload_only：只用上传原料库，补充也仅从上传库找
        #   - 仅勾选②           → default_only：用合并池做基础，但补充仅从默认548库找
        #   - ①+②都勾选（有上传）→ merged：用合并池，补充先从合并池再从默认548库
        #   - 都不勾选           → default：不做自动补充，仅用已选原料求解
        #   - 使用全部可用原料模式 → full_pool：使用 all_ingredients
        _use_auto_expand_1 = st.session_state.get("user_auto_expand_from_upload", False)
        _use_auto_expand_2 = st.session_state.get("user_auto_expand", False)
        _has_upload_file = st.session_state.uploaded_ingredients_df is not None

        if selection_mode == "手动选择原料":
            _selection_pool = st.session_state.get("selection_pool")
            # 仅勾选①且有上传 → 用纯上传库（不含默认库原料），补充也仅限上传库
            if _use_auto_expand_1 and not _use_auto_expand_2 and _has_upload_file:
                _base_df = all_ingredients.copy()  # all_ingredients 就是上传的原料库
                st.session_state._solve_mode = "upload_only"
            # 仅勾选② 或 ①+②都勾选 → 根据具体勾选情况决定基础池
            elif _use_auto_expand_2 or (_use_auto_expand_1 and _use_auto_expand_2 and _has_upload_file):
                if _selection_pool is not None:
                    _base_df = _selection_pool.copy()
                    st.session_state._solve_mode = "merged" if (_use_auto_expand_1 and _use_auto_expand_2) else "default_only"
                else:
                    # 无手动选择时，根据勾选组合决定基础池：
                    #   - ①+② 都勾选 → 合并池（上传+默认548库）
                    #   - 仅勾选 ②   → 仅默认548库（不含上传原料）
                    _preloaded_default_sel = st.session_state.get("default_ingredients")
                    if _use_auto_expand_1 and _use_auto_expand_2 and _has_upload_file and _preloaded_default_sel is not None:
                        # ①+②：合并上传 + 默认548库
                        _up_cols_set = set(all_ingredients.columns)
                        _df_cols_set = set(_preloaded_default_sel.columns)
                        _common_cols = list(_up_cols_set & _df_cols_set)
                        if not _common_cols:
                            _common_cols = list(_up_cols_set | _df_cols_set)
                        _up_part = all_ingredients[[c for c in _common_cols if c in all_ingredients.columns]]
                        _df_part = _preloaded_default_sel[[c for c in _common_cols if c in _preloaded_default_sel.columns]]
                        _base_df = pd.concat([_up_part, _df_part])
                        _base_df = _base_df[~_base_df.index.duplicated(keep="first")]
                        st.session_state._solve_mode = "merged"
                    elif _use_auto_expand_2 and not _use_auto_expand_1 and _preloaded_default_sel is not None:
                        # 仅②：只用默认548库，不用上传原料
                        _base_df = _preloaded_default_sel.copy()
                        st.session_state._solve_mode = "default_only"
                    else:
                        # 兜底：回退到上传库
                        _base_df = all_ingredients.copy()
                        st.session_state._solve_mode = "default_only"
            elif _selection_pool is not None:
                # 没勾选任何自动补充 → 用合并池但不自动补充
                _base_df = _selection_pool.copy()
                st.session_state._solve_mode = "default"
            else:
                _base_df = all_ingredients.copy()
                st.session_state._solve_mode = "default"
        else:
            # 使用全部可用原料模式：
            #   - 默认仅使用上传原料（快速）
            #   - 勾选②（从默认原料库补充）时，合并默认548库
            #   - ①（从上传库补充）在"全部可用"模式下无意义（已全部包含），忽略
            _preloaded_default = st.session_state.get("default_ingredients")
            # 只有勾选②时才合并默认548库（①在全部可用模式下忽略）
            _use_expand_in_full = _use_auto_expand_2

            if _use_expand_in_full and _preloaded_default is not None:
                # 勾选了自动补充 → 合并上传 + 默认548库
                if _has_upload_file:
                    _up_cols_set = set(all_ingredients.columns)
                    _df_cols_set = set(_preloaded_default.columns)
                    _common_cols = list(_up_cols_set & _df_cols_set)
                    if not _common_cols:
                        _common_cols = list(_up_cols_set | _df_cols_set)
                    _up_part = all_ingredients[[c for c in _common_cols if c in all_ingredients.columns]]
                    _df_part = _preloaded_default[[c for c in _common_cols if c in _preloaded_default.columns]]
                    _base_df = pd.concat([_up_part, _df_part])
                    _base_df = _base_df[~_base_df.index.duplicated(keep="first")]
                    st.session_state._solve_mode = "full_pool_merged"
                else:
                    _base_df = _preloaded_default.copy()
                    st.session_state._solve_mode = "full_pool_default"
            else:
                # 未勾选自动补充 → 仅使用上传原料
                _base_df = all_ingredients.copy()
                st.session_state._solve_mode = "full_pool"
            # 清除可能存在的 selection_pool（避免残留）
            if "selection_pool" in st.session_state:
                del st.session_state.selection_pool
            if "display_to_real_map" in st.session_state:
                del st.session_state.display_to_real_map

        if use_dm:
            st.write("转换为干物质基础(DM)...")
            working_ingredients = convert_to_dm_basis(_base_df)
        else:
            working_ingredients = _base_df

        # 过滤原料 — 根据物种自动选择能量列
        if "猪" in selected_sheet:
            energy_col = ENERGY_MAP.get("猪", "猪消化能MC/Kg")
        elif "牛" in selected_sheet or "奶牛" in selected_sheet:
            energy_col = ENERGY_MAP.get("牛", "综合净能(牛)MC/Kg")
        elif "禽" in selected_sheet or "鸡" in selected_sheet or "鸭" in selected_sheet:
            energy_col = ENERGY_MAP.get("禽", "禽代谢能MC/Kg")
        elif "羊" in selected_sheet:
            energy_col = ENERGY_MAP.get("羊", "羊消化能MC/Kg")
        elif "其他" in selected_sheet:
            energy_col = ENERGY_MAP.get("其他", "消化能(其他)MC/Kg")
        else:
            energy_col = ENERGY_MAP.get("禽", "禽代谢能MC/Kg")

        # 动物种类已提前确定

        ing_df = filter_ingredients(working_ingredients, selected_ingredients, energy_col, animal_key_map,
                                   exclude_npn=st.session_state.get("exclude_npn", True))
        st.write(f"可用原料: {len(ing_df)} 种")

        # 加载分类
        feed_type_dict = load_feed_type_classification()
        conc_idx, forage_idx = _get_feed_type_indices(ing_df, feed_type_dict)

        # 配方要求（含精粗比约束）—— 必须在精粗比检查前加载
        formula_min, formula_max, ratio_constraints, ca_p_constraints = load_formula_requirements(
            standards_file_to_use, ing_df, feed_type_dict,
        )
        merged_min = dict(formula_min)
        if min_bounds:
            merged_min.update(min_bounds)
        merged_max = dict(formula_max)
        if max_bounds:
            merged_max.update(max_bounds)

        # ── 关键：统一 min/max_bounds 的 key 与 ing_df.index 一致 ──
        # 背景：ing_df.index 在不同模式下可能是真实名或带标签名
        #   upload_only / full_pool 模式：真实名（无标签）
        #   merged / default_only 模式：带 [上传]/[默认] 标签
        # min_bounds/max_bounds 的 key 可能是真实名（从 limit_pairs 转换来的）
        # 需要把 key 映射到 ing_df.index 中的标准名称
        _ing_index_set = set(ing_df.index)
        # 构建映射：真实名 → 标准索引名（带标签或不带）
        _real_to_idx = {}
        for _idx_name in _ing_index_set:
            if _idx_name.startswith("[上传] ") or _idx_name.startswith("[默认] "):
                _real_to_idx[_idx_name[5:]] = _idx_name
                _real_to_idx[_idx_name] = _idx_name
            else:
                _real_to_idx[_idx_name] = _idx_name
        # 也保留 display_to_real_map 的反向映射
        _dtr = st.session_state.get("display_to_real_map") or {}
        _display_to_idx = {}
        for _display, _real in _dtr.items():
            if _real in _real_to_idx:
                _display_to_idx[_display] = _real_to_idx[_real]
            # 也直接映射
            if _display in _ing_index_set:
                _display_to_idx[_display] = _display

        def _normalize_key(k):
            """将 min/max_bounds 的 key 映射为 ing_df.index 中的标准名称"""
            if k in _ing_index_set:
                return k
            if k in _display_to_idx:
                return _display_to_idx[k]
            if k in _real_to_idx:
                return _real_to_idx[k]
            return k  # 找不到时保留原值，让后续验证报警告

        merged_min = {_normalize_key(k): v for k, v in merged_min.items()}
        merged_max = {_normalize_key(k): v for k, v in merged_max.items()}
        # 同时更新 min_bounds/max_bounds（供自动补充逻辑使用）
        if min_bounds:
            min_bounds = {_normalize_key(k): v for k, v in min_bounds.items()}
        if max_bounds:
            max_bounds = {_normalize_key(k): v for k, v in max_bounds.items()}

        # ── 精粗比约束未分类原料警告 ────────────────────
        has_cf_constraint = any(
            ratio_constraints.get(k) is not None
            for k in ("concentrate_min", "concentrate_max", "forage_min", "forage_max")
        ) or any(
            st.session_state.get("cf_ratio_settings", {}).get(k) is not None
            for k in ("concentrate_min", "concentrate_max", "forage_min", "forage_max")
        )
        if has_cf_constraint:
            classified_count = len(conc_idx) + len(forage_idx)
            total_ing = len(ing_df)
            if classified_count == 0:
                st.warning(
                    "⚠️ 已设置精粗比约束，但无法通过「精粗类型」列或关键词识别任何原料的精粗分类。"
                    "请在原料数据的「精粗类型」列中标注「精料」或「粗料」，"
                    "或在「饲料类型分类.xlsx」中配置分类规则。"
                )
            elif classified_count < total_ing:
                unclassified_names = [
                    ing_df.index[i] for i in range(total_ing)
                    if i not in conc_idx and i not in forage_idx
                ]
                st.info(
                    f"💡 {len(unclassified_names)} 种原料未被分类为精料/粗料，将不参与精粗比计算。"
                    f"如需精确控制，请在原料数据中补充「精粗类型」列。"
                )

        # ── 合并高级约束设置（UI优先）──────────────────
        # 营养上限
        ui_nutrient_max = st.session_state.get("nutrient_max_settings", {})
        nutrient_max_reqs = ui_nutrient_max if ui_nutrient_max else None

        # 精粗比：文件约束为基础，UI设置覆盖
        ui_cf = st.session_state.get("cf_ratio_settings", {})
        for key in ("concentrate_min", "concentrate_max", "forage_min", "forage_max"):
            ui_val = ui_cf.get(key)
            if ui_val is not None:
                ratio_constraints[key] = ui_val

        # 钙磷比：文件约束为基础，UI设置覆盖
        ui_cap = st.session_state.get("ca_p_settings", {})
        for key in ("ca_p_min", "ca_p_max"):
            ui_val = ui_cap.get(key)
            if ui_val is not None:
                ca_p_constraints[key] = ui_val

        # 营养映射（动态匹配：标准指标 → 原料库列名）
        from feed_formulation import _find_best_column_matches

        # 动态匹配所有标准指标到原料列
        std_keys = list(nutrient_requirements.keys())
        matched_map, unmatched_keys = _find_best_column_matches(
            std_keys, list(ing_df.columns), animal_key=animal_key_map
        )

        # 检测是否存在"用户上传原料库特有的营养指标"
        # 即：用户标准要求 + 用户原料库有对应列 + 但主原料库(548种)无此列
        # 如果存在，主库补充无法提供这些指标的值，自动补充没有意义
        user_has_exclusive_nutrients = False
        exclusive_cols = set()
        if st.session_state.uploaded_ingredients_df is not None:
            default_ing = st.session_state.get("default_ingredients")
            if default_ing is not None:
                default_cols = set(default_ing.columns)
                matched_cols = set(matched_map.values())
                exclusive_cols = matched_cols - default_cols
                user_has_exclusive_nutrients = bool(exclusive_cols)

        # 构建 nutrient_map（用于后续 build_and_solve）
        nutrient_map = matched_map.copy()

        # ── 合并抗营养因子上限约束 ────────────────────────
        _anti_factor = st.session_state.get("anti_factor_max_settings", {})
        if _anti_factor:
            if nutrient_max_reqs is None:
                nutrient_max_reqs = {}
            for _col, _val in _anti_factor.items():
                nutrient_max_reqs[_col] = _val
                # 更新 nutrient_map，将该列映射到自身
                if _col not in nutrient_map:
                    nutrient_map[_col] = _col
            st.info(f"🧪 已加载 {len(_anti_factor)} 个抗营养因子上限约束")

        # 构建 deduped_mapped: {std_key: req_value}  仅包含匹配成功的，且同列去重
        deduped_mapped = {}
        seen_cols = set()
        for k, v in nutrient_requirements.items():
            if k in matched_map and v is not None:
                effective_col = matched_map[k]
                if effective_col in seen_cols:
                    continue  # 跳过映射到同一原料列的重复指标
                seen_cols.add(effective_col)
                deduped_mapped[k] = v

        # 报告匹配结果
        if unmatched_keys:
            st.warning(f"⚠️ {len(unmatched_keys)} 个标准指标在原料库中无对应数据，已跳过：{', '.join(sorted(unmatched_keys))}")

        # 显示匹配成功的指标
        if matched_map:
            # 找到实际匹配的能量列
            actual_energy_col = None
            for std_key, ing_col in matched_map.items():
                if "能" in std_key or any(tag in std_key for tag in ["DE", "ME", "NE", "GE"]):
                    actual_energy_col = ing_col
                    break
            if actual_energy_col is None:
                actual_energy_col = ENERGY_MAP.get(animal_key_map, "禽代谢能MC/Kg")
            st.write(f"动态匹配成功 {len(matched_map)} 项（能量列：{actual_energy_col}，动物种类：{animal_key_map}）")

        # 约束验证
        warnings_list = validate_constraints(
            ing_df, deduped_mapped, nutrient_map,
            merged_min if merged_min else None,
            merged_max if merged_max else None,
        )
        if warnings_list:
            st.write("**约束验证警告:**")
            for w in warnings_list:
                st.warning(w)

        # 清除上次自动补充记录
        st.session_state.auto_expand_pool = None
        st.session_state.auto_expand_nutrient_map = None

        # 求解
        st.write("求解线性规划...")
        result = build_and_solve(
            ingredients_df=ing_df,
            nutrient_requirements=deduped_mapped,
            nutrient_map=nutrient_map,
            min_bounds=merged_min if merged_min else None,
            max_bounds=merged_max if merged_max else None,
            concentrate_indices=conc_idx if conc_idx else None,
            forage_indices=forage_idx if forage_idx else None,
            concentrate_min=ratio_constraints.get("concentrate_min"),
            concentrate_max=ratio_constraints.get("concentrate_max"),
            forage_min=ratio_constraints.get("forage_min"),
            forage_max=ratio_constraints.get("forage_max"),
            ca_p_min=ca_p_constraints.get("ca_p_min"),
            ca_p_max=ca_p_constraints.get("ca_p_max"),
            animal_key=animal_key_map,
            nutrient_max_requirements=nutrient_max_reqs,
        )

        # ── 自动扩充逻辑 ──
        # 三种模式（由复选框①/②的组合决定）：
        #   upload_only（仅勾选①）：只从上传原料中补充，失败则提示勾选②
        #   default_only（仅勾选②）：跳过上传库，直接从默认548库补充
        #   merged（①+②都勾选）：先从合并池补充，不足再补默认548库
        #   default（都不勾选）：不做自动补充
        auto_expand_from_upload = st.session_state.get("user_auto_expand_from_upload", False)
        auto_expand_default = st.session_state.get("user_auto_expand", False)
        _has_upload = st.session_state.uploaded_ingredients_df is not None
        _solve_mode = st.session_state.get("_solve_mode", "default")
        _any_expand = auto_expand_from_upload or auto_expand_default
        _has_selection = selected_ingredients is not None
        _has_upload_or_selection = _has_selection or _has_upload

        def _solve_with_pool(
            pool_df, label_hint="",
        ):
            """对给定的原料池求解，返回 result dict。辅助函数，避免重复代码。"""
            fm2, fx2, rc2, cp2 = load_formula_requirements(
                standards_file_to_use, pool_df, feed_type_dict,
            )
            ex_min = dict(fm2)
            if min_bounds:
                ex_min.update(min_bounds)
            ex_max = dict(fx2)
            if max_bounds:
                ex_max.update(max_bounds)
            ui_cf2 = st.session_state.get("cf_ratio_settings", {})
            for key in ("concentrate_min", "concentrate_max", "forage_min", "forage_max"):
                ui_val = ui_cf2.get(key)
                if ui_val is not None:
                    rc2[key] = ui_val
            ui_cap2 = st.session_state.get("ca_p_settings", {})
            for key in ("ca_p_min", "ca_p_max"):
                ui_val = ui_cap2.get(key)
                if ui_val is not None:
                    cp2[key] = ui_val
            ci2, fi2 = _get_feed_type_indices(pool_df, feed_type_dict)
            return build_and_solve(
                ingredients_df=pool_df,
                nutrient_requirements=deduped_mapped,
                nutrient_map=nutrient_map,
                min_bounds=ex_min if ex_min else None,
                max_bounds=ex_max if ex_max else None,
                concentrate_indices=ci2 if ci2 else None,
                forage_indices=fi2 if fi2 else None,
                concentrate_min=rc2.get("concentrate_min"),
                concentrate_max=rc2.get("concentrate_max"),
                forage_min=rc2.get("forage_min"),
                forage_max=rc2.get("forage_max"),
                ca_p_min=cp2.get("ca_p_min"),
                ca_p_max=cp2.get("ca_p_max"),
                animal_key=animal_key_map,
                nutrient_max_requirements=nutrient_max_reqs,
            )

        if result["status"] != "optimal" and _any_expand and not user_has_exclusive_nutrients and _has_upload_or_selection:

            # ── 构建基础原料池 ──────────────────────────
            if selected_ingredients:
                current_pool = working_ingredients.loc[
                    working_ingredients.index.intersection(selected_ingredients)
                ].copy().fillna(0.0)
            elif _has_upload or _solve_mode in ("merged", "full_pool_merged", "full_pool_default"):
                current_pool = working_ingredients.copy().fillna(0.0)
            else:
                current_pool = pd.DataFrame(columns=working_ingredients.columns)

            original_set = set(selected_ingredients) if selected_ingredients else set()
            _solved = False

            # ── 步骤①：根据 _solve_mode 决定补充策略 ────
            # upload_only：仅从上传库补充（不含默认库）
            # merged：从合并池（上传+默认）补充
            # default_only：跳过步骤①，直接走步骤②（默认548库）

            if _solve_mode == "upload_only" and auto_expand_from_upload and _has_upload:
                # ── 仅上传模式：只从上传原料库中找未选的 ──
                _upload_ing = all_ingredients.copy()
                if use_dm and "干物质%" in _upload_ing.columns:
                    _upload_ing = convert_to_dm_basis(_upload_ing)
                # 将选中的真实名称与上传库匹配
                real_selected = set()
                for s in (original_set if original_set else set()):
                    rn = s
                    if s.startswith("[上传] "): rn = s[5:]
                    elif s.startswith("[默认] "): rn = s[5:]
                    real_selected.add(rn)
                _unselected_mask = ~_upload_ing.index.isin(real_selected)
                unselected_df = _upload_ing.loc[_unselected_mask]
                if not unselected_df.empty:
                    st.warning(
                        f"① 正在从上传原料库中未选择的原料补充 "
                        f"({len(unselected_df)} 种，不含默认库)..."
                    )
                    user_expanded = pd.concat([current_pool, unselected_df])
                    # 去重：当用户未选择任何原料时，current_pool已包含全部原料，
                    # unselected_df 可能与 current_pool 有重叠
                    user_expanded = user_expanded[~user_expanded.index.duplicated(keep="first")]
                    common_cols = [c for c in current_pool.columns if c in user_expanded.columns]
                    user_expanded = user_expanded[common_cols]

                    r_user = _solve_with_pool(user_expanded, "上传库未选原料")
                    if r_user["status"] == "optimal":
                        result = r_user
                        result["auto_ingredients"] = [
                            n for n in result["ingredients"] if n not in original_set
                        ]
                        st.session_state.auto_expand_pool = user_expanded
                        st.session_state.auto_expand_nutrient_map = nutrient_map
                        auto_ing = result["auto_ingredients"]
                        st.success(
                            f"✅ 从上传原料库补充了 {len(auto_ing)} 种原料："
                            f"{', '.join(auto_ing[:5])}"
                            f"{'...' if len(auto_ing) > 5 else ''}"
                        )
                        _solved = True
                        current_pool = user_expanded
                    else:
                        # 仅上传模式失败 → 提示用户勾选②
                        st.error(
                            f"❌ 从全部上传原料（{len(user_expanded)} 种）中仍无法求解。"
                            f"\n💡 建议同时勾选选项「②」以获得更多候选原料。"
                        )
                else:
                    st.info("⚠️ 上传原料库中没有更多可补充的原料了")

            elif _solve_mode == "merged" and (auto_expand_from_upload or auto_expand_default):
                # ── 合并库模式（①+②都勾选）：从合并池中找未选的 ──
                unselected_mask = ~working_ingredients.index.isin(current_pool.index)
                unselected_df = working_ingredients.loc[unselected_mask]
                if not unselected_df.empty:
                    st.warning(
                        f"①② 正在从候选原料池（上传+默认）中未选择的原料补充 "
                        f"({len(unselected_df)} 种)..."
                    )
                    user_expanded = pd.concat([current_pool, unselected_df])
                    # 去重（防御性处理）
                    user_expanded = user_expanded[~user_expanded.index.duplicated(keep="first")]
                    common_cols = [c for c in current_pool.columns if c in user_expanded.columns]
                    user_expanded = user_expanded[common_cols]

                    r_user = _solve_with_pool(user_expanded, "候选池未选原料")
                    if r_user["status"] == "optimal":
                        result = r_user
                        result["auto_ingredients"] = [
                            n for n in result["ingredients"] if n not in original_set
                        ]
                        st.session_state.auto_expand_pool = user_expanded
                        st.session_state.auto_expand_nutrient_map = nutrient_map
                        auto_ing = result["auto_ingredients"]
                        st.success(
                            f"✅ 从候选原料库补充了 {len(auto_ing)} 种原料："
                            f"{', '.join(auto_ing[:5])}"
                            f"{'...' if len(auto_ing) > 5 else ''}"
                        )
                        _solved = True
                        current_pool = user_expanded
                    else:
                        st.info(
                            f"⚠️ 从候选原料库补充后仍无法求解"
                            f"（共 {len(user_expanded)} 种），尝试步骤②..."
                        )
                        current_pool = user_expanded

            elif _solve_mode in ("full_pool_merged", "full_pool_default"):
                # 全部可用原料模式：已使用上传+默认548合并池，无需额外补充
                st.info(
                    "📋 当前「使用全部可用原料」模式已包含上传原料和默认548库全部原料"
                    f"（共 {len(working_ingredients)} 种），无更多可补充来源。"
                    "\n如仍无法求解，建议：放宽营养约束或调整用量限制。"
                )
                current_pool = working_ingredients.copy().fillna(0.0)

            elif _solve_mode == "default_only" and auto_expand_default:
                # default_only 模式：跳过合并池补充，直接进入步骤②
                st.info("⏭️ 跳过上传/合并池补充（仅勾选了②），直接从默认548库查找...")

            # ── 步骤②：从默认548原料库补充（勾选②时执行）────
            #   default_only 模式：这是主要（且唯一）的自动补充来源
            #   merged 模式：步骤①失败后的后备方案
            if auto_expand_default and not _solved and result["status"] != "optimal":
                st.warning("② 正在从默认原料库(548种)自动补充...")

                default_ing = st.session_state.get("default_ingredients", working_ingredients)
                if isinstance(default_ing, pd.DataFrame):
                    default_ing = default_ing.copy()
                else:
                    default_ing = load_ingredients(
                        INGREDIENTS_FILE,
                        phos_enabled=st.session_state.get("phos_enabled", True),
                        phos_custom_rates=st.session_state.get("phos_custom_rates") or None,
                        phos_per_ingredient=st.session_state.get("phos_per_ingredient") or None)

                if use_dm and "干物质%" in default_ing.columns:
                    default_ing = convert_to_dm_basis(default_ing)

                lib_pool = filter_ingredients(default_ing, None, energy_col, animal_key_map,
                                             exclude_npn=st.session_state.get("exclude_npn", True))
                common_cols2 = [c for c in lib_pool.columns if c in current_pool.columns]
                lib_pool = lib_pool[common_cols2]
                current_pool = current_pool[common_cols2]

                # 将 current_pool 中的标签名转回真实名称，用于去重比较
                existing_real_names = set()
                for idx_name in current_pool.index:
                    if idx_name.startswith("[上传] ") or idx_name.startswith("[默认] "):
                        existing_real_names.add(idx_name[5:])
                    else:
                        existing_real_names.add(idx_name)

                supplement_rows = lib_pool.loc[~lib_pool.index.isin(existing_real_names)]
                if not supplement_rows.empty:
                    ing_df_full = pd.concat([current_pool, supplement_rows])
                    # 去重（防御性处理，防止标签名/真实名匹配遗漏）
                    ing_df_full = ing_df_full[~ing_df_full.index.duplicated(keep="first")]
                else:
                    ing_df_full = (
                        current_pool.copy()
                        if not current_pool.empty
                        else lib_pool.copy()
                    )

                result = _solve_with_pool(ing_df_full, "默认548库")
                if result["status"] == "optimal":
                    result["auto_ingredients"] = [
                        n for n in result["ingredients"] if n not in original_set
                    ]
                    st.session_state.auto_expand_pool = ing_df_full
                    st.session_state.auto_expand_nutrient_map = nutrient_map
                    auto_ing2 = result["auto_ingredients"]
                    if auto_ing2:
                        st.success(
                            f"✅ 自动补充了 {len(auto_ing2)} 种原料："
                            f"{', '.join(auto_ing2[:5])}"
                            f"{'...' if len(auto_ing2) > 5 else ''}"
                        )
                    else:
                        st.success("✅ 当前原料池已能满足全部约束")
                    _solved = True
                else:
                    st.error(
                        "❌ 即使从默认原料库(548种)补充后仍无法求解，"
                        "请检查营养约束是否合理"
                    )
        elif result["status"] != "optimal" and _any_expand and user_has_exclusive_nutrients:
            st.info(
                f"💡 用户标准包含主原料库不存在的营养指标"
                f"（{', '.join(sorted(exclusive_cols))}），"
                f"已禁用主库自动补充，仅使用用户上传原料求解。"
                f"如需优化配方，请在用户原料库中添加更多原料。"
            )

        # ── 失败诊断：提示用户应补充什么类型的原料 ────────
        if result["status"] != "optimal":
            # 如果用户没有勾选自动补充，先提示
            if not _any_expand and not user_has_exclusive_nutrients:
                st.warning(
                    "💡 **当前原料池无法满足所有约束。**\n\n"
                    "建议操作："
                    "\n1. 勾选上方智能补充选项（☐ 从默认原料库自动补充），程序会自动从548种原料中筛选合适的补充进来；"
                    "\n2. 或在「手动选择原料」模式中选择更多种类的原料（如蛋白质料、氨基酸等）；"
                    "\n3. 或适当放宽用量限制/营养约束。"
                )
            # 分析当前原料池的营养覆盖情况，给出建议
            _diagnose_and_suggest(
                working_ingredients, nutrient_requirements, nutrient_map,
                deduped_matched=deduped_mapped, animal_key=animal_key_map
            )

        result["dm_basis"] = True  # 始终为DM基础（原料已转为DM）
        result["standard_basis"] = standard_basis  # 记录标准的基础类型
        result["air_dried_dm_pct"] = air_dried_dm_pct if standard_basis == "风干饲粮基础" else None
        result["standard_name"] = standard_name
        result["animal_key"] = animal_key_map
        # 保存精粗比/钙磷比约束设置（用于结果展示）
        result["ratio_constraints"] = ratio_constraints
        result["ca_p_constraints"] = ca_p_constraints
        # 保存营养映射和求解用原料（供Excel导出使用）
        result["nutrient_map"] = nutrient_map
        result["_working_ingredients_df"] = working_ingredients

        # 计算实际配方的精粗比和钙磷比（用于结果展示区显示参考值）
        _solved_ing = result.get("ingredients", {})
        if _solved_ing:
            _cf_actual = calc_concentrate_forage(_solved_ing, working_ingredients)
            result["actual_concentrate_pct"] = _cf_actual.get("concentrate_pct")
            result["actual_forage_pct"] = _cf_actual.get("forage_pct")
            _cap_actual = calc_ca_p_ratio(_solved_ing, working_ingredients)
            result["actual_ca_p_ratio"] = _cap_actual
        else:
            result["actual_concentrate_pct"] = None
            result["actual_forage_pct"] = None
            result["actual_ca_p_ratio"] = None

        # 保存实际应用的营养上限约束（用于结果展示）
        result["_nutrient_max_applied"] = nutrient_max_reqs if nutrient_max_reqs else None

        st.session_state.result = result
        st.session_state.solved = True
        # 清除微调状态（新求解结果）
        for _k in ("_adj_initialized", "_adj_nut_cols",
                    "_adj_ing_df", "_adj_requirements", "_adj_original_pct",
                    "_adj_number_values"):
            st.session_state.pop(_k, None)

        if result["status"] == "optimal":
            status.update(label="✅ 求解成功！", state="complete")
        else:
            status.update(label=f"❌ 求解失败", state="error")


# ╔══════════════════════════════════════════════════════════════╗
# ║                   结果展示区                                   ║
# ╚══════════════════════════════════════════════════════════════╝

result = st.session_state.result

if st.session_state.solved and result:
    is_optimal = result["status"] == "optimal"

    # ── 状态标题 ────────────────────────────────────
    if is_optimal:
        st.success("### ✅ 求解成功")
    else:
        st.error(f"### ❌ 求解失败")
        st.error(f"**原因**: {result.get('message', '未知')}")

    # ── 配方概况（无论成功与否都展示）─────────────────
    st.markdown("---")
    st.markdown("### 📊 配方概况")

    num_ingredients = len(result.get("ingredients", {}))
    num_total_constraints = len(result.get("requirements", {}))
    num_passed = 0
    num_failed = 0
    if result.get("nutrients") and result.get("requirements"):
        for std_col, req_val in result.get("requirements", {}).items():
            if req_val is None:
                continue
            try:
                req_float = float(req_val)
            except (ValueError, TypeError):
                continue
            actual = result.get("nutrients", {}).get(std_col, 0.0)
            try:
                actual = float(actual)
            except (ValueError, TypeError):
                actual = 0.0
            # 用与对比表一致的精度判断（浮点数容差一致）
            if round(actual - req_float, 4) >= 0:
                num_passed += 1
            else:
                num_failed += 1

    card_cols = st.columns(5)
    with card_cols[0]:
        cost_val = result.get("cost", 0)
        if isinstance(cost_val, (int, float)) and cost_val > 0:
            st.metric("配方成本", f"¥{cost_val:.4f}元/kg")
        else:
            st.metric("配方成本", "N/A")
    with card_cols[1]:
        st.metric("原料种类", f"{num_ingredients} 种")
    with card_cols[2]:
        if num_total_constraints > 0:
            fail_text = f"{num_failed} 项未达标" if num_failed > 0 else "全部达标"
            st.metric("营养达标率", f"{num_passed}/{num_total_constraints}",
                      delta=fail_text if num_failed > 0 else None)
        else:
            st.metric("营养达标率", "N/A")
    with card_cols[3]:
        if result.get("standard_basis") == "风干饲粮基础":
            basis_label = f"DM(风干{result.get('air_dried_dm_pct', 88)}%)"
        else:
            basis_label = "DM基础"
        st.metric("计算基础", basis_label)
    with card_cols[4]:
        animal_labels = {"猪": "🐷 猪", "禽": "🐔 禽",
                         "牛": "🐄 牛", "羊": "🐑 羊", "其他": "🐱 其他"}
        st.metric("动物种类", animal_labels.get(result.get("animal_key", "禽"), "禽"))

    st.caption(f"饲养标准: **{result.get('standard_name', '未知')}**")

    # ── 精粗比 & 钙磷比（常驻参考值）───────────────────
    # 无论是否设置约束，始终展示实际配方的精粗比和钙磷比，供参考
    _actual_conc = result.get("actual_concentrate_pct")
    _actual_for = result.get("actual_forage_pct")
    _actual_cap = result.get("actual_ca_p_ratio")

    ref_cols = st.columns(2)
    with ref_cols[0]:
        if _actual_conc is not None or _actual_for is not None:
            conc_s = f"{_actual_conc:.1f}%" if _actual_conc is not None else "0%"
            for_s = f"{_actual_for:.1f}%" if _actual_for is not None else "0%"
            st.metric("🌿 精粗比(参考)", f"精{conc_s}/粗{for_s}")
        else:
            st.metric("🌿 精粗比(参考)", "—")
    with ref_cols[1]:
        if _actual_cap is not None:
            st.metric("🦴 钙磷比(参考)", f"{_actual_cap:.2f}:1")
        else:
            st.metric("🦴 钙磷比(参考)", "—")

    # ── 高级约束摘要 ──────────────────────────────
    _show_constraints = False
    _constraint_parts = []

    # 1. 原料用量限制（limit_pairs）
    _limit_pairs = st.session_state.get("limit_pairs", [])
    if _limit_pairs:
        _lp_lines = []
        for lp in _limit_pairs:
            _name = lp.get("name", "")
            _min_v = lp.get("min_val")
            _max_v = lp.get("max_val")
            if _min_v is not None and _max_v is not None:
                _lp_lines.append(f"{_name}: {_min_v}% ~ {_max_v}%")
            elif _min_v is not None:
                _lp_lines.append(f"{_name}: ≥ {_min_v}%")
            elif _max_v is not None:
                _lp_lines.append(f"{_name}: ≤ {_max_v}%")
        if _lp_lines:
            _constraint_parts.append(("📏 原料用量限制", _lp_lines))
            _show_constraints = True

    # 2. 营养上限约束
    _nutrient_max = result.get("_nutrient_max_applied") or st.session_state.get("nutrient_max_settings", {})
    if _nutrient_max:
        _nm_lines = [f"{k}: ≤ {v}" for k, v in _nutrient_max.items() if v is not None]
        if _nm_lines:
            _constraint_parts.append(("🛡 营养上限约束", _nm_lines))
            _show_constraints = True

    # 3. 精粗比约束
    _rc = result.get("ratio_constraints", {}) or {}
    _cf_set = any(_rc.get(k) is not None for k in ("concentrate_min", "concentrate_max", "forage_min", "forage_max"))
    if _cf_set:
        _cf_lines = []
        if _rc.get("concentrate_min") is not None or _rc.get("concentrate_max") is not None:
            _c_min = f"{_rc['concentrate_min']*100:.0f}%" if _rc.get("concentrate_min") is not None else "—"
            _c_max = f"{_rc['concentrate_max']*100:.0f}%" if _rc.get("concentrate_max") is not None else "—"
            _cf_lines.append(f"精料: {_c_min} ~ {_c_max}")
        if _rc.get("forage_min") is not None or _rc.get("forage_max") is not None:
            _f_min = f"{_rc['forage_min']*100:.0f}%" if _rc.get("forage_min") is not None else "—"
            _f_max = f"{_rc['forage_max']*100:.0f}%" if _rc.get("forage_max") is not None else "—"
            _cf_lines.append(f"粗料: {_f_min} ~ {_f_max}")
        if _cf_lines:
            _constraint_parts.append(("🌿 精粗比约束", _cf_lines))
            _show_constraints = True

    # 4. 钙磷比约束
    _cap_c = result.get("ca_p_constraints", {}) or {}
    if _cap_c.get("ca_p_min") is not None or _cap_c.get("ca_p_max") is not None:
        _cap_min_s = f"{_cap_c['ca_p_min']:.2f}" if _cap_c.get("ca_p_min") is not None else "—"
        _cap_max_s = f"{_cap_c['ca_p_max']:.2f}" if _cap_c.get("ca_p_max") is not None else "—"
        _constraint_parts.append(("🦴 钙磷比约束", [f"Ca/P: {_cap_min_s} : 1 ~ {_cap_max_s} : 1"]))
        _show_constraints = True

    if _show_constraints:
        st.markdown("---")
        st.markdown("### ⚙️ 本次求解的约束条件")
        for _ctitle, _clines in _constraint_parts:
            col_a, col_b = st.columns([1, 3])
            with col_a:
                st.markdown(f"**{_ctitle}**")
            with col_b:
                st.markdown(" · ".join(_clines))

    # ── 配方组成 ────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🧪 配方组成")

    formula_df = _build_formula_df(result, all_ingredients, result.get("auto_ingredients", []))
    
    
        

    if not formula_df.empty:
        def _highlight_total(row):
            if row.get("原料名称") == "【合计】":
                return ["background-color: #e8f5e9; font-weight: 700;"] * len(row)
            return [""] * len(row)

        try:
            def _safe_fmt(val, spec):
                """安全格式化：数值用 spec，字符串原样返回。"""
                try:
                    return format(float(val), spec)
                except (ValueError, TypeError):
                    return str(val)

            styled_formula = formula_df.style.apply(_highlight_total, axis=1).format({
                "配比(%)": lambda x: _safe_fmt(x, ".2f"),
                "单价(元/kg)": lambda x: _safe_fmt(x, ".2f"),
                "成本贡献(元)": lambda x: _safe_fmt(x, ".4f"),
            })
            st.dataframe(
                styled_formula,
                use_container_width=True,
                hide_index=True,
                height=min(35 * (len(formula_df) + 1), 600),
            )
        except Exception as fmt_err:
            st.warning(f"表格格式化出错，显示原始数据: {fmt_err}")
            st.dataframe(formula_df, use_container_width=True, hide_index=True)
    else:
        if is_optimal:
            st.info("配方组成为空（可能所有原料比例都低于阈值）")
        else:
            st.warning("⚠️ 未能生成配方组成。请检查：")
            st.markdown("""
            - 原料数据是否包含营养指标列？
            - 饲养标准的营养指标是否与原料数据列名匹配？
            - 是否选用了合适的原料？
            """)

    # 投料单选项（原样基础 as-fed）
    show_asfed = st.checkbox(
        "☐ 显示投料单（原样基础）",
        value=False,
        help="将DM基础配方转换回原样基础(as-fed)，用于实际投料。"
             "配方结果为干物质基础，投料单按各原料干物质含量换算回原样用量"
    )
    
    if show_asfed and result.get("dm_pct_map"):
        # 转换回as-fed基础
        asfed_ingredients = convert_result_to_asfed(
            result["ingredients"], result["dm_pct_map"]
        )
        # 构建as-fed基础的配方组成表格
        asfed_result = {"status": "optimal", "ingredients": asfed_ingredients, "cost": result.get("cost")}
        asfed_df = _build_formula_df(
            asfed_result, all_ingredients, result.get("auto_ingredients", [])
        )
        st.markdown("**📋 投料单（原样基础 as-fed）**")
        st.caption("配方结果为干物质基础，下表已按各原料干物质含量换算为原样基础的实际投料用量")
        if not asfed_df.empty:
            st.dataframe(asfed_df, use_container_width=True, hide_index=True)
        else:
            st.info("投料单为空")
    
    # ── 自动补充原料详情 ────────────────────────────
    auto_ing_list = result.get("auto_ingredients", [])
    expand_pool = st.session_state.get("auto_expand_pool")
    expand_nutrient_map = st.session_state.get("auto_expand_nutrient_map", {})
    if is_optimal and auto_ing_list and expand_pool is not None:
        st.markdown("---")
        st.markdown("### 🔍 自动补充原料营养详情")
        st.caption("以下原料由系统从总原料库中自动补充，建议将其录入自有原料库：")

        detail_df, cat_stats = _build_auto_ingredient_detail(
            result, expand_pool, expand_nutrient_map
        )

        if not detail_df.empty:
            try:
                fmt_spec = {}
                for col in detail_df.columns:
                    if col not in ("原料名称", "配比(%)", "中国分类"):
                        fmt_spec[col] = "{:.3f}"
                fmt_spec["配比(%)"] = "{:.2f}"

                detail_styled = detail_df.style.format(fmt_spec)
                st.dataframe(
                    detail_styled,
                    use_container_width=True,
                    hide_index=True,
                    height=min(38 * (len(detail_df) + 1), 500),
                )
            except Exception as fmt_err:
                st.dataframe(detail_df, use_container_width=True, hide_index=True)

            # ── 原料补充建议 ──────────────
            st.markdown("---")
            st.markdown("### 💡 原料补充建议")
            st.caption("基于自动补充原料的分类分析，建议您在自有原料库中补充以下类别原料：")

            recommendation = _build_feed_recommendation(
                auto_ing_list, cat_stats, expand_pool, expand_nutrient_map
            )
            if recommendation:
                st.markdown(recommendation)
        else:
            st.info("自动补充的原料配比极低（<0.01%），可忽略。")

    # ── 营养对比 ────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📈 营养指标对比")

    nut_df = _build_nutrient_df(result, all_ingredients)
    if not nut_df.empty:
        try:
            styled_nut = nut_df.style \
                .apply(_highlight_nutrition_row, axis=1) \
                .map(_color_diff, subset=["差值"]) \
                .format({
                    "需要量": "{:.4f}",
                    "实际含量": "{:.4f}",
                    "差值": "{:+.4f}",
                })
            st.dataframe(
                styled_nut,
                use_container_width=True,
                hide_index=True,
                height=min(35 * (len(nut_df) + 1), 700),
            )
        except Exception as fmt_err:
            st.warning(f"表格格式化出错，显示原始数据: {fmt_err}")
            st.dataframe(nut_df, use_container_width=True, hide_index=True)

        # 达标统计
        pass_count = (nut_df["达成"] == "✅").sum()
        fail_count = (nut_df["达成"] == "❌").sum()
        if fail_count == 0:
            st.success(f"✅ 全部 {pass_count} 项营养指标均已达标！")
        else:
            st.warning(f"✅ {pass_count} 项达标 | ❌ {fail_count} 项未达标")
            # 未达标项详细
            failed_items = nut_df[nut_df["达成"] == "❌"]
            if not failed_items.empty:
                with st.expander("🔍 查看未达标详情"):
                    for _, row in failed_items.iterrows():
                        indicator = row["营养指标"]
                        req = row["需要量"]
                        actual = row["实际含量"]
                        diff = row["差值"]
                        st.markdown(
                            f"- **{indicator}**: 需要 {req:.4f}, "
                            f"实际 {actual:.4f}, "
                            f"差 {abs(diff):.4f} (不足)"
                        )
    else:
        st.info("暂无营养对比数据。")
        # 诊断提示
        requirements = result.get("requirements", {})
        if requirements and not is_optimal:
            st.markdown("### 🔍 营养需求诊断")
            st.caption("以下营养指标被求解器要求满足，但无结果数据：")
            for key, val in list(requirements.items())[:10]:
                st.markdown(f"- **{key}**: 需要 {val}")

    # ── 配方微调 ────────────────────────────────────
    if is_optimal:
        st.markdown("---")
        # 检查是否需要自动重开dialog（添加原料/重置后触发）
        _should_reopen = st.session_state.pop("_should_reopen_adj_dialog", False)
        if _should_reopen:
            show_adjustment_dialog(result, all_ingredients)
        elif st.button("🔧 配方微调", key="open_adjustment_btn", use_container_width=True):
            show_adjustment_dialog(result, all_ingredients)

        # 如果有已保存的微调版本，显示"继续微调"按钮和下载选项
        _saved_pct = st.session_state.get("_saved_adjusted_pct")
        if _saved_pct is not None:
            _saved_cost = st.session_state.get("_saved_adjusted_cost", 0)
            adj_col1, adj_col2 = st.columns([1, 1])
            with adj_col1:
                if st.button("✏️ 继续微调", key="continue_adjustment_btn", use_container_width=True):
                    show_adjustment_dialog(result, all_ingredients)
            with adj_col2:
                _saved_ing_count = len(_saved_pct)
                st.caption(f"💾 已保存微调版（{_saved_ing_count}种原料，¥{_saved_cost:.4f}/kg）")

    # ── 排查建议（仅非optimal或有不达标时）──────────────
    if not is_optimal or (nut_df is not None and not nut_df.empty and (nut_df["达成"] == "❌").sum() > 0):
        st.markdown("---")
        with st.expander("🔧 排查建议", expanded=not is_optimal):
            st.markdown("""
            **常见原因和解决方法：**
            1. **营养约束过严** → 尝试放宽某些营养指标的最小要求
            2. **原料数据不完整** → 检查上传的原料文件是否包含所有必需的营养列
            3. **指标名不匹配** → 确保饲养标准中的营养指标名与原料列名一致
            4. **用量限制冲突** → 检查最小用量之和是否超过100%
            5. **能量单位不一致** → 检查是 Mcal/kg 还是 MJ/kg
            6. **缺少关键原料** → 某些微量元素需要特定载体（如石粉供钙）
            """)

    # ── 导出Excel ────────────────────────────────────
    st.markdown("---")
    st.markdown("### 💾 导出结果")

    # 投料单选项（as-fed基础）
    include_feeding_sheet = st.checkbox(
        "☐ 同时导出投料单工作表（原样基础）",
        value=False,
        help="在同一个Excel文件中增加一个「投料单」工作表，"
             "将DM基础配方按各原料干物质含量换算为原样基础的投料用量"
    )

    # 如果有已保存的微调版本，提供导出微调版选项
    _has_saved_adj = "_saved_adjusted_pct" in st.session_state
    if _has_saved_adj:
        export_version = st.radio(
            "导出版本",
            options=["原始求解结果", "微调后配方"],
            index=0,
            horizontal=True,
            help="选择要导出的配方版本",
        )
    else:
        export_version = "原始求解结果"

    export_col1, export_col2 = st.columns([1, 3])
    with export_col1:
        if st.button("📥 导出为 Excel 文件", type="secondary",
                     use_container_width=True):
            try:
                export_result = result.copy()

                # 如果选择导出微调后版本，替换原料配比和成本
                if export_version == "微调后配方":
                    _saved_adj_pct = st.session_state.get("_saved_adjusted_pct", {})
                    _saved_adj_cost = st.session_state.get("_saved_adjusted_cost", None)
                    if _saved_adj_pct:
                        export_result["ingredients"] = dict(_saved_adj_pct)
                        if _saved_adj_cost is not None:
                            export_result["cost"] = _saved_adj_cost

                # 获取求解用的原料数据和营养映射（修复营养贡献为0的问题）
                _export_ing_df = result.get("_working_ingredients_df", all_ingredients)
                _export_nutrient_map = result.get("nutrient_map", None)

                # 处理带标签的原料名称：导出时使用真实名称（去掉标签）
                if "ingredients" in export_result:
                    new_ingredients = {}
                    for name, pct in export_result["ingredients"].items():
                        real_name = name
                        if name.startswith("[上传] "):
                            real_name = name[5:]
                        elif name.startswith("[默认] "):
                            real_name = name[5:]
                        new_ingredients[real_name] = pct
                    export_result["ingredients"] = new_ingredients
                
                # 同时处理 auto_ingredients 列表
                auto_ingredients_clean = []
                for name in export_result.get("auto_ingredients", []):
                    real_name = name
                    if name.startswith("[上传] "):
                        real_name = name[5:]
                    elif name.startswith("[默认] "):
                        real_name = name[5:]
                    auto_ingredients_clean.append(real_name)
                export_result["auto_ingredients"] = auto_ingredients_clean

                import tempfile
                with tempfile.NamedTemporaryFile(
                    suffix=".xlsx", delete=False
                ) as tmp:
                    tmp_path = tmp.name
                save_result_to_excel(
                    export_result, tmp_path,
                    ingredients_df=_export_ing_df,
                    nutrient_map=_export_nutrient_map,
                )

                # 标注自动补充的原料（使用真实名称）
                if auto_ingredients_clean:
                    from openpyxl import load_workbook
                    wb = load_workbook(tmp_path)
                    if "配方结果" in wb.sheetnames:
                        ws = wb["配方结果"]
                        for row in ws.iter_rows(min_row=3):
                            cell = row[0]
                            if cell.value and str(cell.value).strip() in auto_ingredients_clean:
                                cell.value = f"{cell.value} *自动补充*"
                    wb.save(tmp_path)

                # ── 投料单工作表（原样基础 as-fed）──
                if include_feeding_sheet and result.get("dm_pct_map"):
                    from openpyxl import load_workbook
                    from feed_formulation import convert_result_to_asfed
                    wb = load_workbook(tmp_path)

                    # 转换为as-fed基础
                    asfed_ingredients = convert_result_to_asfed(
                        export_result["ingredients"], result["dm_pct_map"]
                    )
                    _write_feeding_sheet(
                        wb,
                        dm_ingredients=export_result["ingredients"],
                        asfed_ingredients=asfed_ingredients,
                        result=result,
                    )
                    wb.save(tmp_path)

                with open(tmp_path, "rb") as f:
                    excel_bytes = f.read()
                os.unlink(tmp_path)

                filename = f"配方结果_{selected_standard_name}.xlsx"
                st.download_button(
                    label="📥 点击下载",
                    data=excel_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.success("✅ Excel 文件已生成")
            except Exception as e:
                st.error(f"导出失败: {e}")

# ── 初始提示 ─────────────────────────────────────────────
if not st.session_state.solved:
    st.markdown("---")
    with st.expander("📋 数据概览", expanded=False):
        # 显示原料概览
        st.markdown("#### 🧂 可用原料（前20种）")
        preview_df = all_ingredients.head(20)
        key_cols = [c for c in ["价格", "粗蛋白%", "钙%", "总磷%",
                                 "猪消化能MC/Kg", "禽代谢能MC/Kg",
                                 "综合净能(牛)MC/Kg"] if c in preview_df.columns]
        st.dataframe(preview_df[key_cols], use_container_width=True)

        # 显示饲养标准概览
        st.markdown("#### 📋 可用饲养标准")
        for sheet_name, names in standards_data.items():
            if len(standards_data) > 1:
                st.markdown(f"**{sheet_name}** ({len(names)} 个标准)")
            for i, name in enumerate(names[:20]):
                st.caption(f"{i+1}. {name}")
            if len(names) > 20:
                st.caption(f"... 共 {len(names)} 个标准")
