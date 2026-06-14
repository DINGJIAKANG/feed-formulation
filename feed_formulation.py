"""
线性规划饲料配方程序
目标：在满足饲养标准约束的前提下，最小化配方总成本
"""
from __future__ import annotations

import sys
import os
import numpy as np
import pandas as pd
from scipy.optimize import linprog

# Windows GBK 终端 emoji 兼容性修复：将 stdout/stderr 强制设为 utf-8
# 避免 print() 含 emoji 时抛出 UnicodeEncodeError
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
if sys.stderr and hasattr(sys.stderr, "reconfigure"):
    try:
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# 引入现成模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from read_feed_ingredients import read_feed_ingredients
from read_feeding_standards import read_feeding_standards


# ─────────────────────────── 路径配置 ────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INGREDIENTS_FILE = os.path.join(BASE_DIR, "饲料常规营养含量.xls")
STANDARDS_FILE   = os.path.join(BASE_DIR, "饲养标准.xlsx")
FEED_TYPE_FILE   = os.path.join(BASE_DIR, "饲料类型分类.xlsx")  # 可选的饲料类型配置文件

# 有效磷生物利用率（按原料类型）
_PHOS_UTILIZATION = {
    "cereal": 0.30,    # 谷类（玉米、小麦、麦麸等）
    "oil_meal": 0.50,   # 粕类（豆粕、棉籽粕等）
    "animal": 0.70,      # 动物性原料（鱼粉、肉骨粉等）
    "other": 1.00,       # 其他（矿物质等，100%可利用）
}
# 原料分类关键词
_CEREAL_KW = ["玉米", "小麦", "次粉", "麦麸", "米糠", "大麦", "燕麦", "高粱"]
_ANIMAL_KW = ["鱼粉", "肉骨粉", "血粉", "羽毛粉", "蚕蛹"]


def _get_phosphorus_utilization(ing_name: str, custom_rates: dict | None = None,
                                 per_ingredient: dict | None = None) -> float:
    """根据原料名称返回有效磷利用率。
    优先级：逐个原料设置 > 类别自定义值 > 类别默认值

    custom_rates: 用户自定义的类别利用率，格式如 {"cereal": 0.35, "oil_meal": 0.55}
    per_ingredient: 逐个原料的利用率，格式如 {"玉米[1级]": 0.30, "鱼粉": 0.75}
                   若某原料在此字典中有值，直接返回该值（最高优先级）。
    """
    name = str(ing_name)

    # 最高优先级：逐个原料单独设置
    if per_ingredient and name in per_ingredient:
        return float(per_ingredient[name])

    # 合并默认值和类别自定义值
    rates = dict(_PHOS_UTILIZATION)
    if custom_rates:
        rates.update(custom_rates)

    if any(kw in name for kw in _ANIMAL_KW):
        return rates["animal"]
    elif any(kw in name for kw in _CEREAL_KW):
        return rates["cereal"]
    elif any(kw in name for kw in ["粕", "豆", "蛋白"]):
        return rates["oil_meal"]
    else:
        return rates["other"]


def _add_effective_phosphorus(df: pd.DataFrame,
                              enabled: bool = True,
                              custom_rates: dict | None = None,
                              per_ingredient: dict | None = None) -> pd.DataFrame:
    """
    为原料DataFrame添加/保留'有效磷%'列。

    逻辑：
    1. 原料已有「有效磷%」列且有数值 → 保留原值，不覆盖
    2. 原料无有效磷% 或 全为空值，但有总磷%：
       - enabled=True  → 有效磷 = 总磷 × 利用率
       - enabled=False → 有效磷 = 总磷（全部视为有效）
    3. 只有有效磷% 没有总磷% → 保留原值不变
    4. 两者都没有 → 不做任何处理
    """
    df = df.copy()

    # 检查已有数据情况
    _has_eff_p = "有效磷%" in df.columns
    _has_total_p = "总磷%" in df.columns

    if _has_eff_p:
        _ep_vals = pd.to_numeric(df["有效磷%"], errors="coerce")
        _eff_p_has_data = _ep_vals.notna().any()
        if _eff_p_has_data and not _has_total_p:
            # 场景3：只有有效磷没有总磷 → 直接使用
            print("  ℹ️ 使用原料自带的「有效磷%」列数据（无总磷%，无法换算）")
            return df
        elif _eff_p_has_data and _has_total_p:
            # 场景1：已有有效磷数据 → 保留用户数据
            print("  ℹ️ 原料已含「有效磷%」列且有数值，保留原值不覆盖")
            return df
        # 有列但全为空 → 继续走后面的逻辑填充

    if not _has_total_p:
        # 场景4：两者都没有
        return df

    # 需要从总磷%计算有效磷%
    if not enabled:
        df["有效磷%"] = pd.to_numeric(df["总磷%"], errors="coerce")
        print("  ℹ️ 有效磷换算未启用，有效磷% = 总磷%")
    else:
        _rates_str = ""
        if custom_rates:
            _rates_str = f"（用户自定义: {custom_rates}）"
        df["有效磷%"] = df.apply(
            lambda row: pd.to_numeric(row["总磷%"], errors="coerce") *
                        _get_phosphorus_utilization(row.name, custom_rates, per_ingredient),
            axis=1
        )
        print(f"  ℹ️ 有效磷已按生物利用率计算{_rates_str}")
    return df


def convert_to_dm_basis(df: pd.DataFrame) -> pd.DataFrame:
    """
    将原料营养值从原样基础(as-fed)转换为干物质基础(DM)。
    转换公式：营养值_DM = 营养值_as-fed / (DM% / 100)
    需要原料数据中包含'干物质%'列。
    """
    if "干物质%" not in df.columns:
        print("  ⚠️ 原料数据缺少'干物质%'列，无法转换为干物质基础")
        return df

    df_dm = df.copy()
    dm_raw = pd.to_numeric(df_dm["干物质%"], errors="coerce")

    # ── 第一步：异常值检查（只警告，不替换）─────────────────
    for idx_name in df_dm.index:
        v = dm_raw.get(idx_name, None)
        if pd.notna(v):
            vf = float(v)
            if vf < 5:
                print(f"  ⚠️ 警告: {idx_name} 的干物质%={vf}%，数值异常偏低(<5%)，请核实原始数据！")
            elif vf > 100:
                print(f"  ⚠️ 警告: {idx_name} 的干物质%={vf}%，数值异常偏高(>100%)，请核实原始数据！")

    # ── 第二步：缺失值用 100% 填充 ────────────────────────
    # DM% 缺失时默认 100%（认为营养数据已是 DM 基础，无需转换）
    _default_dm = pd.Series(100.0, index=df_dm.index)

    # 用原始值填充，缺失的才用默认值 100%（异常值保留原值，只警告）
    dm_pct = dm_raw / 100.0  # 直接使用，不替换异常值
    dm_pct = dm_pct.fillna(_default_dm / 100.0)  # 只填充缺失值
    dm_pct = dm_pct.replace([np.inf, -np.inf], np.nan).fillna(1.0)  # 兜底 100%
    
    # 需要转换的营养指标（百分比指标）
    nutrient_cols = [
        "粗蛋白%", "钙%", "总磷%", "有效磷%", "钠%", "氯%", "亚油酸%",
        "赖氨酸%", "蛋氨酸%", "胱氨酸%", "蛋胱氨酸%", "苏氨酸%", "色氨酸%",
        "异亮氨酸%", "亮氨酸%", "精氨酸%", "缬氨酸%", "组氨酸%", 
        "苯丙氨酸%", "酪氨酸%", "苯丙酪氨酸%", "脯氨酸%", "甘丝氨酸%",
        "粗脂肪%", "粗纤维%", "粗灰分%", "中性洗涤纤维%", "酸性洗涤纤维%",
    ]
    
    # 能量指标（Mcal/kg 或 MJ/kg）
    energy_cols = [col for col in df_dm.columns if "MC/Kg" in col or "MJ/Kg" in col]
    nutrient_cols.extend(energy_cols)
    
    # 执行转换
    for col in nutrient_cols:
        if col in df_dm.columns:
            vals = pd.to_numeric(df_dm[col], errors="coerce").fillna(0)
            df_dm[col] = vals / dm_pct
    
    # 添加标记列
    df_dm["_dm_basis"] = True
    if "_dm_basis" not in df.columns:
        df["_dm_basis"] = False
    
    print(f"  ✓ 已转换为干物质基础(DM)")
    return df_dm


def convert_requirements_to_dm(
    nutrient_requirements: dict,
    typical_dm_pct: float = 88.0,
) -> dict:
    """
    将饲养标准从风干饲粮基础转换为干物质基础(DM)。

    当饲养标准给出的是风干饲粮基础的营养浓度值时
    （如CP=16%表示每kg风干饲料含16%粗蛋白），
    需要转为DM基础时除以配合饲料的干物质百分比。

    参数:
        nutrient_requirements: {指标名: 浓度值}（风干饲粮基础）
        typical_dm_pct: 风干饲粮的干物质百分比（默认88%，即猪/禽配合饲料）
                        牛的配合饲料DM%通常约55-65%，羊约60-70%
                        如果饲养标准已是干物质基础(DM)，则无需调用此函数

    返回:
        {指标名: 浓度值}（DM基础）

    公式: 营养值_DM = 营养值_风干 / (DM% / 100)
    """
    dm_factor = typical_dm_pct / 100.0
    if dm_factor <= 0 or dm_factor > 1:
        return nutrient_requirements  # 无效值，不转换
    
    result = {}
    for key, val in nutrient_requirements.items():
        if val is not None and isinstance(val, (int, float)):
            result[key] = val / dm_factor
        else:
            result[key] = val
    
    return result


def convert_result_to_asfed(
    ingredients_dm: dict,
    dm_pct_map: dict,
) -> dict:
    """
    将干物质基础(DM)的配方结果转换回原样基础(as-fed)的投料单。
    
    参数:
        ingredients_dm: {原料名: DM基础下的配比(%)}  例如 玉米=73.08 表示73.08%
        dm_pct_map: {原料名: 干物质含量(%)}  （原样基础，如玉米86%）
    
    返回:
        {原料名: as-fed基础下的配比(%)}
        
    公式:
        DM配方中，配比之和=100%，每个原料占比为 DM配比_i/100
        DM基础下每种原料的干物质量 = (DM配比_i/100) * (DM%_i/100)
        as-fed配比_i = [(DM配比_i/100) * (DM%_i/100)] / Σ[(DM配比_j/100) * (DM%_j/100)] * 100
    """
    # 计算分母：Σ(DM配比_j/100 * DM%_j/100) = Σ(DM配比_j * DM%_j) / 10000
    denominator = 0.0
    for name, pct_dm in ingredients_dm.items():
        dm_pct = dm_pct_map.get(name, 88.0)  # 默认88%
        denominator += (pct_dm / 100.0) * (dm_pct / 100.0)
    
    if denominator <= 0:
        return ingredients_dm  # 无法转换，返回原结果
    
    # 转换每个原料
    result_asfed = {}
    for name, pct_dm in ingredients_dm.items():
        dm_pct = dm_pct_map.get(name, 88.0)
        # as-fed配比 = (DM配比/100 * DM%/100) / 分母 * 100
        pct_asfed = (pct_dm / 100.0) * (dm_pct / 100.0) / denominator * 100.0
        result_asfed[name] = round(pct_asfed, 4)
    
    return result_asfed


# ───────────────────────── 饲料类型分类 ────────────────────────
# 内置分类规则（当配置文件不存在时使用）
# 精料：能量饲料、蛋白质饲料、添加剂等
# 粗料：干草、秸秆、青贮、糟渣类等
_BUILTIN_FEED_TYPE = {
    "精料": [
        "玉米", "小麦", "次粉", "麦麸", "米糠", "大豆", "豆粕", "棉籽粕", "菜籽粕",
        "花生粕", "玉米蛋白粉", "玉米蛋白饲料", "鱼粉", "血粉", "羽毛粉", "肉骨粉",
        "淀粉", "糖", "磷酸", "碳酸钙", "食盐", "小苏打", "预混料", "氨基酸", "维生素",
    ],
    "粗料": [
        "秸秆", "稻草", "麦草", "干草", "青贮", "苜蓿", "草木樨", "槐树叶", "树叶",
        "酒糟", "醋糟", "酱糟", "糖渣", "豆腐渣", "啤酒糟", "甘薯藤", "花生藤",
    ],
}

def load_feed_type_classification() -> dict:
    """
    加载饲料类型分类（精料/粗料/其他）。
    优先从 FEED_TYPE_FILE 读取；若不存在则使用内置规则自动分类。
    返回 {原料名: '精料'/'粗料'/'其他'}
    """
    # 尝试从配置文件读取
    if os.path.exists(FEED_TYPE_FILE):
        try:
            df = pd.read_excel(FEED_TYPE_FILE)
            if "原料名称" in df.columns and "饲料类型" in df.columns:
                return dict(zip(df["原料名称"].astype(str).str.strip(),
                                df["饲料类型"].astype(str).str.strip()))
        except Exception as e:
            print(f"  ⚠️ 读取饲料类型配置文件失败: {e}")

    # 使用内置规则
    return {}


def classify_ingredient(name: str, classification: dict | None = None) -> str:
    """判断原料是精料/粗料/其他"""
    name = str(name).strip()
    # 优先使用配置文件
    if classification and name in classification:
        return classification[name]
    # 内置规则
    for kw in _BUILTIN_FEED_TYPE["粗料"]:
        if kw in name:
            return "粗料"
    for kw in _BUILTIN_FEED_TYPE["精料"]:
        if kw in name:
            return "精料"
    return "其他"


# ────────────────────────── 营养指标映射 ─────────────────────────
# 饲养标准列名  →  原料营养成分列名
# 注意：能量指标在运行时按动物种类自动切换（见 ENERGY_MAP）
NUTRIENT_MAP = {
    # 能量 —— 统一使用 Mcal/kg（MJ/kg 指标在 load_standards 中自动换算后也用此键）
    "Mcal/kg":                 "禽代谢能MC/Kg",
    "ME, Mcal/kg":             "禽代谢能MC/Kg",
    "ME, MJ/kg":               "禽代谢能MC/Kg",    # 会在load_standards中换算为Mcal/kg
    # 注意：ME, MJ/kg 在 load_standards 中换算为 ME, Mcal/kg 后再匹配
    # 蛋白质
    "粗蛋白质CP,%":             "粗蛋白%",
    "CP, %":                   "粗蛋白%",
    # 氨基酸（常用，列名与原料数据库保持一致）
    "赖氨酸%":                  "赖氨酸%",
    "蛋氨酸%":                  "蛋氨酸%",
    "蛋氨酸 %":                 "蛋氨酸%",
    "胱氨酸%":                  "胱氨酸%",
    "蛋胱氨酸%":                "蛋胱氨酸%",
    "苏氨酸%":                  "苏氨酸%",
    "色氨酸%":                  "色氨酸%",
    "异亮氨酸%":                "异亮氨酸%",
    "亮氨酸%":                  "亮氨酸%",
    "精氨酸%":                  "精氨酸%",
    "缬氨酸%":                  "缬氨酸%",
    "组氨酸%":                  "组氨酸%",
    "苯丙氨酸%":                "苯丙氨酸%",
    "酪氨酸%":                  "酪氨酸%",
    "苯丙酪氨酸%":              "苯丙酪氨酸%",
    # 矿物质
    "钙，%":                    "钙%",
    "总磷，%":                  "总磷%",
    "有效磷，%":                 "有效磷%",
    "有效磷":                   "有效磷%",
    "钠，%":                    "钠%",
    "氯，%":                    "氯%",
    "镁，%":                    None,    # 原料库中无镁%列
    "钾，%":                    "钾%",
    # 脂肪与纤维
    "粗脂肪%":                  "粗脂肪%",
    "粗纤维%":                  "粗纤维%",
    # 脂肪酸
    "亚油酸，%":                "亚油酸%",
    # 食盐
    "食盐%":                    "食盐%",
    "食盐，%":                  "食盐%",
    # 其他
    "胆碱g":                    None,    # 原料库中无胆碱%列
    # 反刍动物纤维指标（NDF/ADF）
    "NDF, %":                  "中性洗涤纤维%",
    "ADF, %":                  "酸性洗涤纤维%",
    "中性洗涤纤维NDF,%":        "中性洗涤纤维%",
    "酸性洗涤纤维ADF,%":        "酸性洗涤纤维%",
    # 反刍动物能量（肉牛）
    "肉牛维持净能, Mcal/kg":    "肉牛维持净能MC/Kg",
    "肉牛增重净能, Mcal/kg":    "肉牛增重净能MC/Kg",
    "综合净能, Mcal/kg":        "综合净能(牛)MC/Kg",
    "综合净能, MJ/kg":          "综合净能(牛)MC/Kg",
    # 反刍动物能量（奶牛）
    "产奶净能, Mcal/kg":        "产奶净能MC/Kg",
    "产奶净能, MJ/kg":          "产奶净能MC/Kg",
    # 反刍动物能量（羊）
    "羊消化能, Mcal/kg":        "羊消化能MC/Kg",
    "羊消化能, MJ/kg":          "羊消化能MC/Kg",
    "羊代谢能, Mcal/kg":        "羊代谢能MC/Kg",
    "羊代谢能, MJ/kg":          "羊代谢能MC/Kg",
    # 猪消化能
    "DE, Mcal/kg":             "猪消化能MC/Kg",
    "DE, MJ/kg":               "猪消化能MC/Kg",  # 会在load_standards中换算
    # ── 常见变体（兼容用户自定义文件的不同命名习惯）──
    "消化能 MC/Kg":            "猪消化能MC/Kg",
    "消化能Mcal/kg":           "猪消化能MC/Kg",
    "消化能, Mcal/kg":         "猪消化能MC/Kg",
    "DE,MC/kg":                "猪消化能MC/Kg",
    "禽代谢能 MC/Kg":          "禽代谢能MC/Kg",
    "代谢能 MC/Kg":            "禽代谢能MC/Kg",
    "ME,MC/kg":                "禽代谢能MC/Kg",
    # ── 其他动物（猫/狗/马/鱼等）──
    "脂肪%":                   "粗脂肪%",
    "灰分%":                   "粗灰分%",
    "Mg%":                     None,    # 原料库中无镁%列
    "K%":                      "钾%",
    "Na%":                     "钠%",
    "Cl%":                     "氯%",
    "S%":                      None,    # 原料库中无硫%列
    "总膳食纤维%":              "粗纤维%",   # 近似映射
    # 微量元素（mg/kg 单位）—— 原料库中暂无此数据，跳过
    "Cu，mg/kg":               None,
    "Fe，mg/kg":               None,
    "Mn，mg/kg":               None,
    "Zn，mg/kg":               None,
    "Se，mg/kg":               None,
    "I，mg/kg":                None,
    "维生素A，IU/kg":           None,
    "维生素D，IU/kg":           None,
    "维生素E，IU/kg":           None,
}

# 能量指标：按动物种类自动选择
ENERGY_MAP = {
    "猪": "猪消化能MC/Kg",
    "禽": "禽代谢能MC/Kg",
    "牛": "综合净能(牛)MC/Kg",
    "羊": "羊消化能MC/Kg",
    "其他": "消化能(其他)MC/Kg",   # 猫、狗、马等
}


# ── 动态营养列匹配 ──────────────────────────────────────────
def _normalize_col_name(name: str) -> str:
    """标准化列名用于模糊匹配：去空格、统一标点、小写"""
    return name.replace(" ", "").replace("_", "").replace("，", ",").replace("。", ".").lower()


def _find_best_column_matches(
    standard_keys: list[str],
    ing_columns: list[str],
    animal_key: str = "禽",
    nutrient_map_override: dict | None = None,
) -> tuple[dict[str, str], set[str]]:
    """
    为每个标准指标在原料库中查找最匹配的列名。

    匹配优先级：
      1. 标准键直接在原料列中存在
      2. NUTRIENT_MAP 硬编码映射（非 None）
      3. 标准化后精确匹配
      4. 核心关键词相互包含匹配
      5. 能量列按动物种类从 ENERGY_MAP 查找

    返回：
      - matched: {standard_key: ingredient_column}  只包含匹配成功的
      - unmatched: set[str]  未找到匹配的键
    """
    matched = {}
    unmatched = set()
    ing_set = set(ing_columns)
    ing_norm = {_normalize_col_name(c): c for c in ing_columns}

    for key in standard_keys:
        # 1. 直接匹配
        if key in ing_set:
            matched[key] = key
            continue

        # 2. NUTRIENT_MAP 硬编码映射（仅非 None）
        map_val = NUTRIENT_MAP.get(key) if nutrient_map_override is None else nutrient_map_override.get(key)
        if map_val is not None and map_val in ing_set:
            matched[key] = map_val
            continue

        # 3. 标准化后精确匹配
        key_norm = _normalize_col_name(key)
        if key_norm in ing_norm:
            matched[key] = ing_norm[key_norm]
            continue

        # 3b. 从 ing 列反查 key_norm
        # 检查任意 ing 列的标准化名是否等于 key_norm
        found = False
        for ic in ing_columns:
            if _normalize_col_name(ic) == key_norm:
                matched[key] = ic
                found = True
                break
        if found:
            continue

        # 4. 核心词匹配：提取 key 中的关键词，在 ing 列中搜索
        # 提取纯字母/中文部分作为关键词
        import re
        key_alpha = re.sub(r'[^a-zA-Z\u4e00-\u9fff]', '', key_norm)
        # 如果key里有能/蛋白/纤维等核心词，尝试在ing列中查找
        core_keywords = [
            # 从完整 key 中提取的字母+中文
            key_alpha,
        ]
        # 额外提取：常见核心词
        for kw in ["钙", "磷", "镁", "钾", "钠", "氯", "硫", "铁", "铜", "锌", "锰",
                    "硒", "碘", "钴", "能", "蛋白", "脂肪", "纤维", "灰分", "干物质",
                    "赖氨酸", "蛋氨酸", "苏氨酸", "色氨酸", "精氨酸", "亮氨酸", "缬氨酸",
                    "盐", "胆碱", "亚油酸",
                    "ca", "p", "mg", "k", "na", "cl", "s", "fe", "cu", "zn", "mn",
                    "se", "i", "co", "de", "me", "ne", "ge", "cp", "ndf", "adf"]:
            if kw in key_norm and kw not in core_keywords:
                core_keywords.append(kw)

        for ck in core_keywords:
            if len(ck) < 2:
                continue
            for ic in ing_columns:
                ic_norm = _normalize_col_name(ic)
                if ck in ic_norm:
                    matched[key] = ic
                    found = True
                    break
            if found:
                break
        if found:
            continue

        # 5. 能量列特殊处理：按动物种类从 ENERGY_MAP 查找
        if "能" in key or any(tag in key_norm for tag in ["de", "me", "ne", "ge", "mcal", "mj", "kj"]):
            energy_col = ENERGY_MAP.get(animal_key, "禽代谢能MC/Kg")
            if energy_col in ing_set:
                matched[key] = energy_col
                continue

        # 未匹配
        unmatched.add(key)

    return matched, unmatched


# ──────────────────────────── 数据加载 ───────────────────────────
def load_ingredients(file_path: str,
                     phos_enabled: bool = True,
                     phos_custom_rates: dict | None = None,
                     phos_per_ingredient: dict | None = None) -> pd.DataFrame:
    """将原料字典转为 DataFrame，行=原料，列=营养指标+价格+精粗类型
    phos_enabled: 是否启用有效磷利用率换算（默认True）
    phos_custom_rates: 用户自定义的类别有效磷利用率，如 {"cereal": 0.35, "oil_meal": 0.55}
    phos_per_ingredient: 逐个原料的有效磷利用率，如 {"玉米[1级]": 0.30, "鱼粉[进口]": 0.75}
    """
    raw = read_feed_ingredients(file_path)
    if not raw:
        raise FileNotFoundError(f"无法读取原料文件: {file_path}")
    records = []
    for name, data in raw.items():
        clean_name = str(name).strip()   # 去除前后空格
        row = {"原料名称": clean_name, "价格": data.get("价格", 0)}
        row.update(data.get("营养成分", {}))
        # 添加精粗类型（若有）
        ft = data.get("精粗类型")
        if ft is not None:
            ft_str = str(ft).strip()
            if ft_str and ft_str.lower() not in ("nan", "none", ""):
                row["精粗类型"] = ft_str
        records.append(row)
    df = pd.DataFrame(records).set_index("原料名称")
    # 价格为空或0的原料标记为不可用
    df["价格"] = pd.to_numeric(df["价格"], errors="coerce")
    # 价格为空/0/NaN 的原料设为 999（高价避免被LP主动选中，但保留在池中）
    _price_mask = (df["价格"].isna()) | (df["价格"] <= 0)
    if _price_mask.any():
        print(f"  ℹ️ {_price_mask.sum()} 种原料无价格，已设为 999 元/kg（可通过设置最小用量强制使用）")
    df["价格"] = df["价格"].fillna(999).where(df["价格"] > 0, 999)
    # 补充/修正常见原料的市场参考价（只补缺失，不覆盖用户值）
    df = _apply_price_overrides(df)
    # 计算有效磷（可配置是否启用及自定义利用率）
    df = _add_effective_phosphorus(df, enabled=phos_enabled, custom_rates=phos_custom_rates,
                                     per_ingredient=phos_per_ingredient)
    return df


def load_standards(file_path: str, standard_index: int = 0) -> tuple[dict, str]:
    """
    解析饲养标准Excel。

    支持两种格式：
    格式1（旧）：两列结构，A列=营养指标名，B列=需要量
    格式2（新）：多列结构，每行一个标准，每列一个营养指标

    参数:
        file_path: Excel文件路径
        standard_index: 选择第几个标准（0=第一个，默认）

    返回 (nutrient_requirements, standard_name)
    """
    try:
        # 尝试格式2（新格式）：读取"饲养标准"工作表
        df = pd.read_excel(file_path, sheet_name="饲养标准", header=0)
        if len(df) > 0 and "标准名称" in df.columns:
            # 新格式：每行一个标准
            if standard_index >= len(df):
                standard_index = 0  # 默认选第一个
            row = df.iloc[standard_index]
            standard_name = str(row.get("标准名称", "")).strip()
            # 调试：打印标准名称
            # print(f"DEBUG: standard_name={repr(standard_name)}")
            print(f"  已加载饲养标准：{standard_name}")

            # 跳过非营养指标列
            skip_cols = ["动物种类", "标准名称", "标准来源"]
            # MJ/kg → Mcal/kg 换算因子（仅用于 NUTRIENT_MAP 中已映射的指标）
            MJ_TO_MCAL = 1.0 / 4.184
            nutrients = {}
            for col in df.columns:
                if col in skip_cols:
                    continue
                val = row.get(col)
                if pd.isna(val):
                    continue
                try:
                    fval = float(val)
                except (ValueError, TypeError):
                    continue

                # 能量单位处理：
                # 仅当该指标在 NUTRIENT_MAP 中有映射时才换算（说明目标原料列使用 Mcal/kg）
                # 无映射的指标（如用户自定义的"其他消化能MJ/Kg"）保持原单位
                col_store = col
                if ("kJ/kg" in col or "kJ/Kg" in col) and (col in NUTRIENT_MAP):
                    fval = fval / 4184.0
                    col_store = col.replace("kJ/kg", "Mcal/kg").replace("kJ/Kg", "Mcal/kg")
                    if col_store in nutrients:
                        continue
                elif ("MJ/kg" in col or "MJ/Kg" in col) and (col in NUTRIENT_MAP):
                    fval = fval * MJ_TO_MCAL
                    col_store = col.replace("MJ/kg", "Mcal/kg").replace("MJ/Kg", "Mcal/kg")
                    if col_store in nutrients:
                        continue

                nutrients[col_store] = fval
            return nutrients, standard_name

        # 格式1（旧格式）：两列结构
        df = pd.read_excel(file_path, sheet_name=0, header=None)
        standard_name = str(df.iloc[0, 1]).strip()  # B1 为标准名称

        nutrients = {}
        for i in range(1, len(df)):
            key = str(df.iloc[i, 0]).strip()    # A列：指标名
            val = df.iloc[i, 1]                  # B列：需要量
            if not key or key.lower() in ("nan", "项目", ""):
                continue
            nutrients[key] = float(val) if not pd.isna(val) else None

        return nutrients, standard_name

    except Exception as e:
        print(f"读取饲养标准出错: {e}")
        return {}, ""


# ─────────────────────── 配方要求加载 ────────────────────────────
def _fuzzy_match_ingredient(short_name: str, ingredients_df: pd.DataFrame) -> str | None:
    """模糊匹配原料名称：精确→前缀→包含，支持简称匹配全名"""
    names = list(ingredients_df.index)

    if short_name in names:
        return short_name

    # 前缀匹配（如 "玉米" → "玉米[1级8.7%]"）
    for name in names:
        if name.startswith(short_name):
            return name

    # 包含匹配
    for name in names:
        if short_name in name or name in short_name:
            return name

    return None


def load_formula_requirements(
    file_path: str,
    ingredients_df: pd.DataFrame,
    feed_type_dict: dict | None = None,   # {原料名: '精料'/'粗料'/'其他'}
) -> tuple[dict, dict, dict, dict]:
    """
    读取饲养标准文件中的"配方要求"工作表，提取原料限量、精粗比约束、钙磷比约束。

    工作表格式（不存在该工作表则返回空）：
      第0行: 表头（原料, 限量, %）
      第1~n行: [原料名, 不等式(>= / <= / 空), 限量值(%)]
      特殊指令:
        - 精料比例: 精料占总日粮的最小/最大比例
        - 粗料比例: 粗料占总日粮的最小/最大比例
        - 钙磷比: 钙与总磷的比例范围（如 钙磷比 >= 1.2, 钙磷比 <= 2.0）

    返回 (min_bounds, max_bounds, ratio_constraints, ca_p_constraints)
      ratio_constraints = {
          'concentrate_min': float or None,
          'concentrate_max': float or None,
          'forage_min':      float or None,
          'forage_max':      float or None,
      }
      ca_p_constraints = {
          'ca_p_min': float or None,   # 钙磷比最小值（如 1.2 表示 1.2:1）
          'ca_p_max': float or None,   # 钙磷比最大值（如 2.0 表示 2.0:1）
      }
    """
    try:
        xl = pd.ExcelFile(file_path)
        if "配方要求" not in xl.sheet_names:
            return {}, {}, {}, {}

        df = pd.read_excel(file_path, sheet_name="配方要求", header=None)
        min_bounds = {}
        max_bounds = {}
        ratio_constraints: dict[str, float | None] = {
            "concentrate_min": None,
            "concentrate_max": None,
            "forage_min": None,
            "forage_max": None,
        }
        ca_p_constraints: dict[str, float | None] = {
            "ca_p_min": None,
            "ca_p_max": None,
        }

        for i in range(1, len(df)):
            ing_name_raw = str(df.iloc[i, 0]).strip()
            op_raw = (
                str(df.iloc[i, 1]).strip()
                if not pd.isna(df.iloc[i, 1])
                else ""
            )
            pct_raw = df.iloc[i, 2]

            if not ing_name_raw or ing_name_raw.lower() in ("nan", "", "原料"):
                continue
            if pd.isna(pct_raw):
                continue

            pct = float(pct_raw) / 100.0          # % → 比例

            # ── 处理精粗比指令 ──────────────────────────────────
            if ing_name_raw in ("精料比例", "精料", "concentrate"):
                key = "concentrate_min" if op_raw in (">=", ">", "") else "concentrate_max"
                if op_raw in (">=", ">"):
                    ratio_constraints["concentrate_min"] = pct
                else:
                    ratio_constraints["concentrate_max"] = pct
                print(f"  📋 配方要求: 精料比例 {op_raw or '≤'} {pct * 100:.0f}%")
                continue
            elif ing_name_raw in ("粗料比例", "粗料", "forage"):
                key = "forage_min" if op_raw in (">=", ">", "") else "forage_max"
                if op_raw in (">=", ">"):
                    ratio_constraints["forage_min"] = pct
                else:
                    ratio_constraints["forage_max"] = pct
                print(f"  📋 配方要求: 粗料比例 {op_raw or '≤'} {pct * 100:.0f}%")
                continue

            # ── 处理钙磷比指令 ────────────────────────────────
            elif ing_name_raw in ("钙磷比", "钙磷", "Ca:P", "钙/磷", "钙磷比例"):
                # 钙磷比的值不是百分比，不需要除以100
                ratio_val = float(pct_raw)   # 直接使用原始值，如 1.5 表示 1.5:1
                if op_raw in (">=", ">"):
                    ca_p_constraints["ca_p_min"] = ratio_val
                    print(f"  📋 配方要求: 钙磷比 ≥ {ratio_val:.2f}:1")
                elif op_raw in ("<=", "<"):
                    ca_p_constraints["ca_p_max"] = ratio_val
                    print(f"  📋 配方要求: 钙磷比 ≤ {ratio_val:.2f}:1")
                elif op_raw == "=":
                    ca_p_constraints["ca_p_min"] = ratio_val
                    ca_p_constraints["ca_p_max"] = ratio_val
                    print(f"  📋 配方要求: 钙磷比 = {ratio_val:.2f}:1")
                else:   # 空 → 上限
                    ca_p_constraints["ca_p_max"] = ratio_val
                    print(f"  📋 配方要求: 钙磷比 ≤ {ratio_val:.2f}:1")
                continue

            # ── 处理普通原料限量 ─────────────────────────────────
            matched = _fuzzy_match_ingredient(ing_name_raw, ingredients_df)
            if matched is None:
                print(f"  ⚠️ 配方要求中的原料未找到: {ing_name_raw}")
                continue

            # 处理等式约束和严格不等式
            if op_raw == "=":
                min_bounds[matched] = pct
                max_bounds[matched] = pct
                print(f"  📋 配方要求: {matched} = {pct * 100:.0f}%")
            elif op_raw in (">=",):
                min_bounds[matched] = pct
                print(f"  📋 配方要求: {matched} ≥ {pct * 100:.0f}%")
            elif op_raw in (">",):
                eps = max(pct * 0.01, 0.0005)
                min_bounds[matched] = pct + eps
                print(f"  📋 配方要求: {matched} > {pct * 100:.0f}% (实际≥{min_bounds[matched]*100:.2f}%)")
            elif op_raw in ("<=",):
                max_bounds[matched] = pct
                print(f"  📋 配方要求: {matched} ≤ {pct * 100:.0f}%")
            elif op_raw in ("<",):
                eps = max(pct * 0.01, 0.0005)
                max_bounds[matched] = pct - eps
                print(f"  📋 配方要求: {matched} < {pct * 100:.0f}% (实际≤{max_bounds[matched]*100:.2f}%)")
            else:   # 空 → 上限（默认）
                max_bounds[matched] = pct
                print(f"  📋 配方要求: {matched} ≤ {pct * 100:.0f}%")

        return min_bounds, max_bounds, ratio_constraints, ca_p_constraints

    except Exception as e:
        print(f"读取配方要求出错: {e}")
        return {}, {}, {}, {}


# ────────────────────────── 约束验证 ─────────────────────────────
def validate_constraints(
    ingredients_df: pd.DataFrame,
    nutrient_requirements: dict,
    nutrient_map: dict,
    min_bounds: dict | None = None,
    max_bounds: dict | None = None,
    concentrate_indices: list[int] | None = None,
    forage_indices: list[int] | None = None,
    concentrate_min: float | None = None,
    concentrate_max: float | None = None,
    forage_min: float | None = None,
    forage_max: float | None = None,
) -> list[str]:
    """
    验证约束是否合理，返回警告信息列表。
    检查项目：
      1. 最小用量之和 > 1.0
      2. 原料上下界矛盾（min > max）
      3. 精粗比约束与原料类型匹配
    """
    warnings = []
    names = list(ingredients_df.index)

    # 辅助函数：去掉原料名中的 [上传]/[默认] 标签前缀，用于标签无关匹配
    def _strip_tag(n):
        if n.startswith("[上传] ") or n.startswith("[默认] "):
            return n[5:]
        return n

    # 构建一个标签无关的名称集合（真实名 → 原始索引名映射）
    _real_to_index = {_strip_tag(n): n for n in names}

    # 0. 检查 min/max_bounds 中的原料名是否存在（标签无关匹配）
    if min_bounds:
        for name in min_bounds:
            _real_key = _strip_tag(name)
            if _real_key not in _real_to_index:
                warnings.append(f"⚠️ 最小用量设置中的原料 '{name}' 在当前原料列表中不存在")
    if max_bounds:
        for name in max_bounds:
            _real_key = _strip_tag(name)
            if _real_key not in _real_to_index:
                warnings.append(f"⚠️ 最大用量设置中的原料 '{name}' 在当前原料列表中不存在")

    # 1. 检查最小用量之和
    min_sum = 0.0
    if min_bounds:
        for name, val in min_bounds.items():
            _rk = _strip_tag(name)
            if _rk in _real_to_index:
                min_sum += val
    if min_sum > 1.0:
        warnings.append(f"⚠️ 原料最小用量之和 = {min_sum*100:.1f}% > 100%，问题必然不可行！")
    elif min_sum > 0.95:
        warnings.append(f"⚠️ 原料最小用量之和 = {min_sum*100:.1f}%，接近100%，可能无法满足营养约束")
    
    # 2. 检查原料上下界矛盾
    if min_bounds and max_bounds:
        for name in min_bounds:
            if name in max_bounds and min_bounds[name] > max_bounds[name]:
                warnings.append(f"⚠️ 原料 {name} 的最小用量({min_bounds[name]*100:.1f}%) > 最大用量({max_bounds[name]*100:.1f}%)")
    
    # 3. 检查精粗比约束
    if concentrate_min is not None and concentrate_max is not None and concentrate_min > concentrate_max:
        warnings.append(f"⚠️ 精料比例下限({concentrate_min*100:.1f}%) > 上限({concentrate_max*100:.1f}%)")
    if forage_min is not None and forage_max is not None and forage_min > forage_max:
        warnings.append(f"⚠️ 粗料比例下限({forage_min*100:.1f}%) > 上限({forage_max*100:.1f}%)")
    
    return warnings


# ────────────────────────── 原料筛选 ─────────────────────────────
# 非蛋白氮(NPN)原料 —— 仅反刍动物可用，单胃动物(猪/禽/其他)自动排除
_NPN_FEEDS = {'尿素', '双缩脲', '氯化铵', '硫酸铵', '液氨', '氨水'}

# 特殊原料默认用量上限（避免LP过度使用廉价的非常规原料）
# 键为原料名关键词(部分匹配)，值为最大比例(小数)
_DEFAULT_MAX_BOUNDS = {
    '血粉': 0.03,          # 血粉≤3%
    '羽毛粉': 0.04,        # 羽毛粉≤4%
    '肉骨粉': 0.08,        # 肉骨粉≤8%（猪/禽）
    '尿素': 0.02,          # 尿素≤2%（反刍动物）
}

# 原料名称常见别名（行业习惯 → 数据库名称关键词）
_INGREDIENT_ALIASES = {
    '豆粕': '大豆粕',
    '麸皮': '小麦麸',
    '棉粕': '棉籽粕',
    '菜粕': '菜籽粕',
    '豆饼': '大豆饼',
    '花生粕': '花生仁粕',
    '葵花粕': '向日葵仁粕',
    'DDGS': '玉米DDGS',
    '肉骨粉': '肉骨粉',
}

# 原料价格修正（补全常见大宗原料的参考市场价，元/kg）
_PRICE_OVERRIDES = {
    # 能量饲料
    '玉米[2级8%]': 1.55,
    '玉米[3级7.8%]': 1.50,
    '大麦[裸2级]': 1.35,        # 修复原数据1.0偏低
    '大麦[皮2级]': 1.25,
    '高粱[1级]': 1.30,
    '燕麦[全]': 1.60,
    '碎米[2级]': 1.70,
    '次粉[1级]': 1.45,
    '木薯干[泰国]': 1.10,
    # 蛋白质饲料
    '大豆粕[2级44.2%]': 3.50,
    '大豆粕[1级47.9%]': 3.70,
    '大豆粕[1级46.8%]': 3.60,
    '棉籽粕[2级40.2%]': 3.00,
    '菜籽粕[38.6%]': 2.80,
    '菜籽饼[35.7%]': 2.60,
    '花生仁粕[48.8%]': 3.60,
    '向日葵仁粕[36.5%]': 2.40,
    '玉米蛋白粉[63.5%]': 5.50,
    '玉米蛋白粉[44.3%]': 4.50,
    '玉米蛋白粉[51.3%]': 5.00,
    # 动物蛋白
    '鱼粉(CP60.2%)': 7.50,
    '鱼粉(CP53.5%)': 6.80,
    '肉骨粉[50%]': 4.50,
    # 麸皮糠类
    '小麦麸[1级11.9%]': 1.60,
    '小麦麸[1级15.7%]': 1.65,
    '小麦麸[2级]': 1.50,
    '米糠[2级13.0%]': 1.80,
    # 矿物质
    '石粉': 0.30,
    '磷酸氢钙': 3.50,
    '食盐': 0.80,
    '碳酸钙': 0.80,
    '贝壳粉': 0.50,
    # 氨基酸
    'L-赖氨酸盐酸盐': 8.00,
    'DL-蛋氨酸': 20.00,
    'L-苏氨酸': 12.00,
    'L-色氨酸': 60.00,
}

# 反刍动物默认精粗比建议（精料比例范围）
# 仅当代码主动传入这些参数时启用；设为None表示不启用默认约束
_DEFAULT_RUMINANT_RATIOS = {
    # animal_key: (concentrate_min, concentrate_max)
    '牛': (0.25, 0.75),   # 默认精料25%~75%
    '羊': (0.20, 0.70),   # 默认精料20%~70%
}


def _apply_price_overrides(df: pd.DataFrame) -> pd.DataFrame:
    """
    处理原料价格：
    1. 用户已设置的价格 → 保留（最高优先级）
    2. 价格缺失/为0 → 设为999（高价格，LP将避免使用）
    
    这样，未设置价格的原料一般不会被纳入计算（因为价格太高），
    但如果用户在用量限制中设置了下限（强制使用），仍可正常求解。
    """
    df = df.copy()
    set_high = []
    
    for name in df.index:
        price = df.loc[name, '价格']
        if pd.isna(price) or price == 0:
            df.loc[name, '价格'] = 999.0
            set_high.append(name)
    
    if set_high:
        print(f"  ⚠️ {len(set_high)} 种原料价格缺失，已设为999（高价格，将避免使用）")
        print(f"     如需使用这些原料，请在上传文件中填写价格，或在高级设置中设置用量下限强制使用。")
    
    return df


def _apply_special_max_bounds(
    names: list[str],
    max_bounds: dict | None,
    min_bounds: dict | None = None,
    animal_key: str = None,
) -> dict:
    """
    为特殊原料（血粉、羽毛粉、尿素等）自动添加默认用量上限。
    不覆盖用户已设置的 max_bounds 或 min_bounds。
    
    优先级：
    1. 用户设置了上限（max_bounds）→ 使用用户设置，不添加默认上限
    2. 用户设置了下限（min_bounds）→ 不添加默认上限（避免冲突）
    3. 用户未设置任何约束 → 添加默认上限
    """
    result = dict(max_bounds) if max_bounds else {}
    auto_added = []
    
    for name in names:
        # 检查用户是否已设置该原料的任何约束
        user_set_max = name in result  # 用户设置了上限
        user_set_min = min_bounds and name in min_bounds  # 用户设置了下限
        
        if user_set_max or user_set_min:
            # 用户已设置约束，不应用默认上限
            continue
        
        # 检查是否需要添加默认上限
        for kw, limit in _DEFAULT_MAX_BOUNDS.items():
            if kw in name:
                # 尿素仅对反刍动物限制，对单胃动物已通过 filter_ingredients 排除
                if kw == '尿素' and animal_key not in ('牛', '羊'):
                    continue
                result[name] = limit
                auto_added.append(f'{name}≤{limit*100:.0f}%')
                break
    
    if auto_added:
        print(f"  🔧 已为特殊原料自动添加用量上限: {', '.join(auto_added)}")
    return result


def _match_ingredient_alias(short_name: str, df: pd.DataFrame) -> str | None:
    """
    通过别名映射查找原料：先查别名表，再在数据库中模糊匹配。
    例如 "豆粕[2级43%]" → 查找含 "大豆粕" 关键词的原料名。
    """
    names = list(df.index)
    # 先精确匹配
    if short_name in names:
        return short_name
    # 别名替换查找
    for alias, keyword in _INGREDIENT_ALIASES.items():
        if alias in short_name and alias not in short_name.replace(alias, '', 1):
            # short_name starts with or contains alias as a distinct word
            # Try to find the keyword in the database
            candidates = [n for n in names if keyword in n]
            if candidates:
                # Try exact match first, then rank by similarity
                remaining = short_name.replace(alias, '').strip('[]（）()')
                for c in candidates:
                    c_no_bracket = c.split('[')[0].strip()
                    if remaining and remaining in c_no_bracket:
                        return c
                # If no match with remaining, return the first candidate with closest CP
                if remaining:
                    try:
                        target_cp = float(remaining.replace('%', ''))
                        best = None
                        best_diff = float('inf')
                        for c in candidates:
                            # Extract CP from name like "大豆粕[2级44.2%]"
                            import re
                            m = re.search(r'(\d+\.?\d*)%', c)
                            if m:
                                cp = float(m.group(1))
                                diff = abs(cp - target_cp)
                                if diff < best_diff:
                                    best_diff = diff
                                    best = c
                        if best and best_diff <= 5:  # CP差值≤5%才匹配
                            return best
                    except ValueError:
                        pass
                return candidates[0]  # fallback: return first match
    return None


def _get_feed_type_indices(
    ingredients_df: pd.DataFrame,
    feed_type_dict: dict | None = None,
) -> tuple[list[int], list[int]]:
    """
    根据饲料类型分类，返回精料和粗料的索引列表。
    优先级：
      1. 原料DataFrame中的"精粗类型"列（值="精料"或"粗料"）
      2. feed_type_dict（从饲料类型分类.xlsx读取）
      3. 内置关键词匹配（classify_ingredient）
    返回 (concentrate_indices, forage_indices)
    """
    concentrate_idx = []
    forage_idx = []
    
    # 检查是否有"精粗类型"列
    has_feed_type_col = "精粗类型" in ingredients_df.columns
    
    for i, name in enumerate(ingredients_df.index):
        # 优先使用"精粗类型"列
        if has_feed_type_col:
            ftype_val = str(ingredients_df.iloc[i]["精粗类型"]).strip()
            if ftype_val in ("精料", "粗料"):
                if ftype_val == "精料":
                    concentrate_idx.append(i)
                else:
                    forage_idx.append(i)
                continue  # 已分类，跳过后续逻辑
        
        # 回退：feed_type_dict 或 关键词匹配
        ftype = classify_ingredient(name, feed_type_dict)
        if ftype == "精料":
            concentrate_idx.append(i)
        elif ftype == "粗料":
            forage_idx.append(i)
    
    return concentrate_idx, forage_idx


def filter_ingredients(df: pd.DataFrame,
                       selected_names: list[str] | None = None,
                       energy_col: str = "禽代谢能MC/Kg",
                       animal_key: str | None = None,
                       exclude_npn: bool = True) -> pd.DataFrame:
    """
    按用户选定的原料名称过滤；
    单胃动物（猪/禽/其他）默认排除NPN原料（尿素等），可通过 exclude_npn=False 关闭；
    保留矿物质、氨基酸等特殊原料（允许能量为0）；
    返回干净的 DataFrame。
    """
    if selected_names:
        df = df.loc[df.index.intersection(selected_names)]

    # 单胃动物排除NPN原料（尿素、双缩脲等）——可通过 exclude_npn=False 关闭
    if exclude_npn and animal_key and animal_key not in ('牛', '羊'):
        npn_to_exclude = [n for n in _NPN_FEEDS if n in df.index]
        if npn_to_exclude:
            df = df.drop(npn_to_exclude)
            print(f"  🔧 已排除NPN原料（{animal_key}为单胃动物，不可利用非蛋白氮）: {npn_to_exclude}")

    return df.copy()


# ────────────────────────── 线性规划核心 ─────────────────────────
def build_and_solve(
    ingredients_df: pd.DataFrame,
    nutrient_requirements: dict,
    nutrient_map: dict,
    min_bounds: dict | None = None,
    max_bounds: dict | None = None,
    total_ratio: float = 1.0,
    concentrate_indices: list[int] | None = None,   # 精料索引列表
    forage_indices: list[int] | None = None,        # 粗料索引列表
    concentrate_min: float | None = None,           # 精料最小比例
    concentrate_max: float | None = None,           # 精料最大比例
    forage_min: float | None = None,                # 粗料最小比例
    forage_max: float | None = None,                # 粗料最大比例
    ca_p_min: float | None = None,                # 钙磷比最小值（如 1.2 表示 1.2:1）
    ca_p_max: float | None = None,                # 钙磷比最大值（如 2.0 表示 2.0:1）
    animal_key: str | None = None,               # 动物类型（用于默认max_bounds/精粗比）
    nutrient_max_requirements: dict | None = None,  # 营养指标上限（新增）
    auto_scale: bool = True,                      # 是否自动按 total_ratio 缩放营养需求
) -> dict:
    """
    线性规划求解最低成本配方。

    变量 x_i：第 i 种原料在配方中的质量比例（0~1）

    目标函数：min Σ c_i * x_i   （c_i = 价格）

    约束：
      1. Σ x_i = total_ratio           （总量约束，默认=1）
      2. 对每个营养指标 n：
           Σ a_ni * x_i >= 需要量下限   （营养下限，auto_scale=True 时自动×total_ratio）
      3. 对每个营养指标 n（可选）：
           Σ a_ni * x_i <= 需要量上限   （营养上限，auto_scale=True 时自动×total_ratio）
      4. lb_i <= x_i <= ub_i            （各原料用量上下限）
      5. 精粗比约束（可选）：
           Σ x_i (i∈精料) >= concentrate_min
           Σ x_i (i∈精料) <= concentrate_max
           Σ x_i (i∈粗料) >= forage_min
           Σ x_i (i∈粗料) <= forage_max
      6. 钙磷比约束（可选，线性近似）：
           Ca/P >= ca_p_min  →  Σ(钙%_i - ca_p_min*磷%_i) * x_i >= 0
           Ca/P <= ca_p_max  →  Σ(钙%_i - ca_p_max*磷%_i) * x_i <= 0

    返回：
      {'status': 'optimal'/'infeasible'/..., 'ingredients': {...}, 'cost': float, 'nutrients': {...}}
    """
    names = list(ingredients_df.index)
    _name_set = set(names)

    # 统一 min_bounds/max_bounds 的 key 与 names 一致（标签无关匹配）
    # 无论调用方传的是真实名还是带 [上传]/[默认] 标签的名，都映射到 names 中的标准名称
    _real_to_idx = {}
    for _idx in _name_set:
        if _idx.startswith("[上传] ") or _idx.startswith("[默认] "):
            _real_to_idx[_idx[5:]] = _idx
            _real_to_idx[_idx] = _idx
        else:
            _real_to_idx[_idx] = _idx

    if min_bounds:
        min_bounds = {_real_to_idx.get(k, k): v for k, v in min_bounds.items()}
    if max_bounds:
        max_bounds = {_real_to_idx.get(k, k): v for k, v in max_bounds.items()}

    n = len(names)
    if n == 0:
        return {"status": "no_ingredients", "message": "没有可用原料"}

    # ── 特殊原料默认用量上限（在total_ratio缩放之前应用）───
    _max_bounds = _apply_special_max_bounds(names, max_bounds, min_bounds, animal_key)

    # ── 自动缩放营养需求（total_ratio ≠ 1 时）─────────────────
    if auto_scale and total_ratio != 1.0:
        scaled_requirements = {k: v * total_ratio for k, v in nutrient_requirements.items()}
        # 同步缩放 min/max_bounds
        scaled_min = None
        scaled_max = None
        if min_bounds:
            scaled_min = {k: v * total_ratio for k, v in min_bounds.items()}
        if _max_bounds:
            scaled_max = {k: v * total_ratio for k, v in _max_bounds.items()}
        # 缩放营养上限
        scaled_max_reqs = None
        if nutrient_max_requirements:
            scaled_max_reqs = {k: v * total_ratio for k, v in nutrient_max_requirements.items()}
        if total_ratio > 1.0:
            print(f"  🔧 total_ratio={total_ratio}, 已自动缩放 {len(scaled_requirements)} 个营养需求")
        _nutrient_reqs = scaled_requirements
        _min_bounds = scaled_min if scaled_min else min_bounds
        _max_bounds = scaled_max if scaled_max else _max_bounds
        _nutrient_max_reqs = scaled_max_reqs if scaled_max_reqs else nutrient_max_requirements
    else:
        _nutrient_reqs = nutrient_requirements
        _min_bounds = min_bounds
        _nutrient_max_reqs = nutrient_max_requirements

    # ── 反刍动物默认精粗比约束 ───────────────────────────────
    # 仅当用户未设置精粗比约束时，为反刍动物启用默认值
    if animal_key in _DEFAULT_RUMINANT_RATIOS:
        defaults = _DEFAULT_RUMINANT_RATIOS[animal_key]
        if concentrate_min is None and concentrate_max is None and forage_min is None and forage_max is None:
            if concentrate_indices is None or forage_indices is None:
                concentrate_indices, forage_indices = _get_feed_type_indices(ingredients_df)
            if concentrate_indices and forage_indices:
                concentrate_min = defaults[0]
                concentrate_max = defaults[1]
                print(f"  🔧 已为{animal_key}启用默认精粗比: 精料{concentrate_min*100:.0f}%~{concentrate_max*100:.0f}%")

    # ── 约束验证 ─────────────────────────────────────────────
    warnings = validate_constraints(
        ingredients_df, _nutrient_reqs, nutrient_map,
        _min_bounds, _max_bounds,
        concentrate_indices, forage_indices,
        concentrate_min, concentrate_max, forage_min, forage_max,
    )
    for w in warnings:
        print(f"  {w}")

    # ── 目标函数（价格向量）───────────────────────────────
    prices = ingredients_df["价格"].values.astype(float)

    # ── 各原料上下界
    lb = np.zeros(n)
    ub = np.ones(n)
    if _min_bounds:
        for i, name in enumerate(names):
            if name in _min_bounds:
                lb[i] = _min_bounds[name]
    if _max_bounds:
        for i, name in enumerate(names):
            if name in _max_bounds:
                ub[i] = _max_bounds[name]

    bounds = list(zip(lb, ub))

    # ── 等式约束：总量 = total_ratio
    A_eq = np.ones((1, n))
    b_eq = np.array([total_ratio])

    # ── 不等式约束：营养指标
    A_ub_list = []
    b_ub_list = []

    for std_col, req_value in _nutrient_reqs.items():
        if req_value is None:
            continue
        req_value = float(req_value)

        # 找对应的原料列
        ing_col = nutrient_map.get(std_col)
        if ing_col is None or ing_col not in ingredients_df.columns:
            continue

        nutrient_vals = pd.to_numeric(ingredients_df[ing_col], errors="coerce").fillna(0).values

        # 下限约束：-Σa*x <= -req  （即 Σa*x >= req）
        A_ub_list.append(-nutrient_vals)
        b_ub_list.append(-req_value)

    if A_ub_list:
        A_ub = np.array(A_ub_list)
        b_ub = np.array(b_ub_list)
    else:
        A_ub = None
        b_ub = None

    # ── 营养上限约束 ─────────────────────────────────────
    if _nutrient_max_reqs:
        for std_col, max_val in _nutrient_max_reqs.items():
            if max_val is None:
                continue
            max_val = float(max_val)

            # 找对应的原料列
            ing_col = nutrient_map.get(std_col)
            if ing_col is None or ing_col not in ingredients_df.columns:
                continue

            nutrient_vals = pd.to_numeric(ingredients_df[ing_col], errors="coerce").fillna(0).values

            # 上限约束：Σa*x <= max_val
            if A_ub is not None:
                A_ub = np.vstack([A_ub, nutrient_vals])
                b_ub = np.concatenate([b_ub, [max_val]])
            else:
                A_ub = nutrient_vals.reshape(1, -1)
                b_ub = np.array([max_val])

            print(f"    约束: {std_col} ≤ {max_val:.4f}")

    # ── 精粗比约束 ───────────────────────────────────────────
    # 约束形式：
    #   精料最小比例: sum(x[concentrate_indices]) >= concentrate_min
    #   → -sum(x[concentrate_indices]) <= -concentrate_min
    #   精料最大比例: sum(x[concentrate_indices]) <= concentrate_max
    #   → sum(x[concentrate_indices]) <= concentrate_max
    ratio_A_ub = []
    ratio_b_ub = []

    if concentrate_min is not None and concentrate_indices is not None:
        row = np.zeros(n)
        row[concentrate_indices] = -1.0
        ratio_A_ub.append(row)
        ratio_b_ub.append(-concentrate_min)
        print(f"    约束: 精料比例 ≥ {concentrate_min*100:.1f}%")
    if concentrate_max is not None and concentrate_indices is not None:
        row = np.zeros(n)
        row[concentrate_indices] = 1.0
        ratio_A_ub.append(row)
        ratio_b_ub.append(concentrate_max)
        print(f"    约束: 精料比例 ≤ {concentrate_max*100:.1f}%")
    if forage_min is not None and forage_indices is not None:
        row = np.zeros(n)
        row[forage_indices] = -1.0
        ratio_A_ub.append(row)
        ratio_b_ub.append(-forage_min)
        print(f"    约束: 粗料比例 ≥ {forage_min*100:.1f}%")
    if forage_max is not None and forage_indices is not None:
        row = np.zeros(n)
        row[forage_indices] = 1.0
        ratio_A_ub.append(row)
        ratio_b_ub.append(forage_max)
        print(f"    约束: 粗料比例 ≤ {forage_max*100:.1f}%")

    if ratio_A_ub:
        ratio_A_ub = np.array(ratio_A_ub)
        ratio_b_ub = np.array(ratio_b_ub)
        if A_ub is not None:
            A_ub = np.vstack([A_ub, ratio_A_ub])
            b_ub = np.concatenate([b_ub, ratio_b_ub])
        else:
            A_ub = ratio_A_ub
            b_ub = ratio_b_ub

    # ── 钙磷比约束（线性近似）────────────────────────────────
    # 钙磷比 = Σ(钙%_i * x_i) / Σ(磷%_i * x_i)
    # 线性化：
    #   Ca/P >= ca_p_min  →  Σ(钙%_i - ca_p_min*磷%_i) * x_i >= 0
    #                       →  -Σ(钙%_i - ca_p_min*磷%_i) * x_i <= 0
    #   Ca/P <= ca_p_max  →  Σ(钙%_i - ca_p_max*磷%_i) * x_i <= 0
    if ca_p_min is not None or ca_p_max is not None:
        # 获取钙%和总磷%的原料数据
        ca_col = nutrient_map.get("钙，%", "钙%")
        p_col = nutrient_map.get("总磷，%", "总磷%")
        
        if ca_col in ingredients_df.columns and p_col in ingredients_df.columns:
            ca_vals = pd.to_numeric(ingredients_df[ca_col], errors="coerce").fillna(0).values
            p_vals = pd.to_numeric(ingredients_df[p_col], errors="coerce").fillna(0).values

            if ca_p_min is not None:
                # Ca/P >= ca_p_min  →  -Σ(ca - ca_p_min*p) * x <= 0
                row = -(ca_vals - ca_p_min * p_vals)
                if A_ub is not None:
                    A_ub = np.vstack([A_ub, row])
                    b_ub = np.concatenate([b_ub, [0.0]])
                else:
                    A_ub = row.reshape(1, -1)
                    b_ub = np.array([0.0])
                print(f"    约束: 钙磷比 ≥ {ca_p_min:.2f}:1")

            if ca_p_max is not None:
                # Ca/P <= ca_p_max  →  Σ(ca - ca_p_max*p) * x <= 0
                row = (ca_vals - ca_p_max * p_vals)
                if A_ub is not None:
                    A_ub = np.vstack([A_ub, row])
                    b_ub = np.concatenate([b_ub, [0.0]])
                else:
                    A_ub = row.reshape(1, -1)
                    b_ub = np.array([0.0])
                print(f"    约束: 钙磷比 ≤ {ca_p_max:.2f}:1")
        else:
            print(f"  ⚠️ 钙磷比约束需要'钙%'和'总磷%'列，但未找到")

    # ── 求解
    result = linprog(
        c=prices,
        A_ub=A_ub,
        b_ub=b_ub,
        A_eq=A_eq,
        b_eq=b_eq,
        bounds=bounds,
        method="highs",
    )

    if result.status != 0:
        status_msg = {
            1: "迭代次数超限",
            2: "问题不可行（约束冲突）",
            3: "问题无界",
            4: "数值问题",
        }.get(result.status, f"未知状态({result.status})")
        return {
            "status": "failed",
            "message": status_msg,
            "raw_status": result.status,
        }

    x = result.x
    # 过滤比例极小的原料（< 0.01%）
    threshold = 1e-4
    # total_ratio ≠ 1 时，x_i 需要换算回百分比
    scale = 100.0 / total_ratio if total_ratio != 0 else 100.0
    formula = {
        names[i]: round(x[i] * scale, 4)
        for i in range(n)
        if x[i] > threshold
    }

    # ── 计算实际精粗比和钙磷比 ────────────────────────────────
    actual_concentrate = None
    actual_forage = None
    actual_ca_p_ratio = None

    if concentrate_indices is not None and len(concentrate_indices) > 0:
        actual_concentrate = round(float(sum(x[i] for i in concentrate_indices)) / total_ratio * scale, 2) if total_ratio != 0 else 0.0
    if forage_indices is not None and len(forage_indices) > 0:
        actual_forage = round(float(sum(x[i] for i in forage_indices)) / total_ratio * scale, 2) if total_ratio != 0 else 0.0

    # 钙磷比 = Σ钙%*x / Σ磷%*x
    ca_col = None
    p_col = None
    for col in ingredients_df.columns:
        if col == "钙%":
            ca_col = col
        elif col == "总磷%":
            p_col = col
    if ca_col and p_col:
        ca_vals = pd.to_numeric(ingredients_df[ca_col], errors="coerce").fillna(0).values
        p_vals = pd.to_numeric(ingredients_df[p_col], errors="coerce").fillna(0).values
        total_ca = float(np.dot(ca_vals, x)) / total_ratio if total_ratio != 0 else 0.0
        total_p = float(np.dot(p_vals, x)) / total_ratio if total_ratio != 0 else 0.0
        if total_p > 1e-10:
            actual_ca_p_ratio = round(total_ca / total_p, 2)

    # ── 计算实际营养含量（仅计算饲养标准中要求的指标）──
    # 同样需要除以 total_ratio 还原到单位比例
    actual_nutrients = {}
    for std_col, req_val in _nutrient_reqs.items():
        if req_val is None:
            continue
        ing_col = nutrient_map.get(std_col, std_col)
        if ing_col in ingredients_df.columns:
            vals = pd.to_numeric(ingredients_df[ing_col], errors="coerce").fillna(0).values
            actual_nutrients[std_col] = round(float(np.dot(vals, x)) / total_ratio, 4) if total_ratio != 0 else 0.0
        else:
            actual_nutrients[std_col] = 0.0

    # requirements 也除以 total_ratio 还原到单位比例（用于显示对比）
    display_requirements = {}
    for k, v in _nutrient_reqs.items():
        display_requirements[k] = round(v / total_ratio, 4) if total_ratio != 0 and v is not None else v

    return {
        "status": "optimal",
        "ingredients": formula,          # {原料名: 百分比}
        "cost": round(float(result.fun) / total_ratio, 4) if total_ratio != 0 else round(float(result.fun), 4),  # 元/kg（归一化）
        "nutrients": actual_nutrients,   # {营养指标: 实际含量（单位比例）}
        "requirements": display_requirements,  # 需求值（单位比例，与 actual 同尺度）
        # 实际精粗比和钙磷比
        "actual_concentrate_pct": actual_concentrate,  # 精料实际百分比
        "actual_forage_pct": actual_forage,            # 粗料实际百分比
        "actual_ca_p_ratio": actual_ca_p_ratio,        # 实际钙磷比
        # 添加DM%映射，用于转换回as-fed基础的投料单
        "dm_pct_map": {
            names[i]: float(pd.to_numeric(ingredients_df["干物质%"], errors="coerce").fillna(88.0).iloc[i])
            for i in range(n)
            if i < len(ingredients_df)
        } if "干物质%" in ingredients_df.columns else {},
    }


# ─────────────────────────── 结果输出 ────────────────────────────
def print_result(result: dict):
    if result["status"] != "optimal":
        print(f"\n❌ 求解失败：{result.get('message', '未知原因')}")
        return

    # 计算基础标记（原料始终转为DM基础优化）
    dm_basis = result.get("dm_basis", False)
    basis_label = "干物质基础(DM)" if dm_basis else "原样基础(as-fed)"

    print("\n" + "=" * 60)
    print(f"✅  最优饲料配方（最低成本）[{basis_label}]")
    print("=" * 60)

    print(f"\n📦 原料组成（共 {len(result['ingredients'])} 种）：")
    print(f"  {'原料名称':<25} {'用量(%)':>10}")
    print(f"  {'-'*25} {'-'*10}")
    total_pct = 0
    for name, pct in sorted(result["ingredients"].items(), key=lambda x: -x[1]):
        print(f"  {name:<25} {pct:>10.2f}%")
        total_pct += pct
    print(f"  {'合计':<25} {total_pct:>10.2f}%")

    print(f"\n💰 配方成本：{result['cost']:.4f} 元/kg")

    print(f"\n🧪 营养含量对比（需要量 vs 实际）：")
    print(f"  {'指标':<15} {'需要量':>10} {'实际含量':>12} {'状态':>6}")
    print(f"  {'-'*15} {'-'*10} {'-'*12} {'-'*6}")
    reqs = result.get("requirements", {})
    for nutrient, actual in result["nutrients"].items():
        req = reqs.get(nutrient)
        if req is not None:
            req = float(req)
            flag = "✅" if actual >= req - 1e-6 else "⚠️"
            print(f"  {nutrient:<15} {req:>10.4f} {actual:>12.4f} {flag:>6}")
    print("=" * 60)


# ─────────────────────── 交互式主程序 ─────────────────────────────
def interactive_mode():
    print("\n" + "=" * 60)
    print("  🌾  线性规划饲料配方优化系统")
    print("=" * 60)

    # 1. 加载数据
    print("\n正在加载数据...")
    try:
        all_ingredients = load_ingredients(INGREDIENTS_FILE)
        nutrient_requirements, standard_name = load_standards(STANDARDS_FILE)
    except Exception as e:
        print(f"数据加载失败：{e}")
        sys.exit(1)

    if not nutrient_requirements:
        print("❌ 饲养标准为空，请检查文件")
        sys.exit(1)

    print(f"  已加载原料：{len(all_ingredients)} 种")
    print(f"  饲养标准：  {standard_name}")
    print(f"  营养约束指标数：{len(nutrient_requirements)}")

    # 1.5 加载饲料类型分类，计算精料/粗料索引
    feed_type_dict = load_feed_type_classification()
    # 为 all_ingredients 中的每个原料分类
    # 注意：这里只是预计算，实际索引在过滤后重新计算

    # 2. 自动推断能量指标
    animal_key = "禽"
    for key in ENERGY_MAP:
        if key in standard_name:
            animal_key = key
            break
    if any(kw in standard_name for kw in ["猪", "仔猪", "母猪"]):
        animal_key = "猪"
    elif any(kw in standard_name for kw in ["牛", "奶牛", "肉牛"]):
        animal_key = "牛"
    elif any(kw in standard_name for kw in ["羊"]):
        animal_key = "羊"
    elif any(kw in standard_name for kw in ["鸡", "鸭", "鹅", "禽", "肉鸡", "蛋鸡"]):
        animal_key = "禽"
    elif any(kw in standard_name for kw in ["猫", "狗", "马", "鱼", "其他"]):
        animal_key = "其他"
    energy_col = ENERGY_MAP.get(animal_key, "禽代谢能MC/Kg")
    nutrient_map = dict(NUTRIENT_MAP)
    nutrient_map["Mcal/kg"] = energy_col
    print(f"  能量指标：    {energy_col}（自动识别）")

    # 只保留 NUTRIENT_MAP 中有映射的约束，并对同一原料列去重（保留先出现的）
    seen_cols = set()
    deduped_requirements = {}
    for k, v in nutrient_requirements.items():
        if v is None or k not in NUTRIENT_MAP:
            continue
        target_col = NUTRIENT_MAP[k]
        # 检查是否同一物理量列的能量指标：用 nutrient_map（含动物特定映射）
        effective_col = nutrient_map.get(k, target_col)
        if effective_col in seen_cols:
            continue  # 跳过重复约束
        seen_cols.add(effective_col)
        deduped_requirements[k] = v
    nutrient_requirements = deduped_requirements

    # 3. 选择原料
    print(f"\n【原料选择】")
    print(f"  a) 使用全部原料（{len(all_ingredients)} 种，推荐）")
    print(f"  b) 手动选择原料")
    choice = input("  请选择 [a/b]：").strip().lower()

    selected_names = None
    min_bounds = {}
    max_bounds = {}

    if choice == "b":
        selected_names, min_bounds, max_bounds = _manual_ingredient_selection(all_ingredients)
    else:
        print("  将使用全部有效原料。")
        default_max = {
            "玉米[1级8.7%]": 0.65,
            "豆粕[1级44%]": 0.35,
            "麸皮": 0.15,
            "石粉": 0.08,
            "磷酸氢钙": 0.03,
            "食盐": 0.005,
        }
        for name, ub in default_max.items():
            if name in all_ingredients.index:
                max_bounds[name] = ub
        print(f"  已为常见原料设置默认上限")

    # 4. 过滤原料
    ing_df = filter_ingredients(all_ingredients, selected_names, energy_col, animal_key)
    print(f"\n  有效原料数：{len(ing_df)} 种")
    if len(ing_df) == 0:
        print("  ❌ 没有可用原料，请检查原料文件")
        return

    # 4.5 加载配方要求（自动应用原料限量+精粗比约束+钙磷比约束）
    print("\n【配方要求】")
    formula_min, formula_max, ratio_constraints, ca_p_constraints = load_formula_requirements(
        STANDARDS_FILE, ing_df, feed_type_dict,
    )
    # 合并：配方要求写入底层，用户手动设置的优先覆盖
    merged_min = dict(formula_min)
    merged_min.update(min_bounds)
    merged_max = dict(formula_max)
    merged_max.update(max_bounds)
    if not formula_min and not formula_max and not any(ratio_constraints.values()) and not any(ca_p_constraints.values()):
        print("  （无配方要求）")

    # 计算精料/粗料索引
    concentrate_idx, forage_idx = _get_feed_type_indices(ing_df, feed_type_dict)

    # 5. 求解（失败时自动扩充原料库）
    print("\n正在求解线性规划...")
    result = build_and_solve(
        ingredients_df=ing_df,
        nutrient_requirements=nutrient_requirements,
        nutrient_map=nutrient_map,
        min_bounds=merged_min if merged_min else None,
        max_bounds=merged_max if merged_max else None,
        concentrate_indices=concentrate_idx if concentrate_idx else None,
        forage_indices=forage_idx if forage_idx else None,
        concentrate_min=ratio_constraints.get("concentrate_min"),
        concentrate_max=ratio_constraints.get("concentrate_max"),
        forage_min=ratio_constraints.get("forage_min"),
        forage_max=ratio_constraints.get("forage_max"),
        ca_p_min=ca_p_constraints.get("ca_p_min"),
        ca_p_max=ca_p_constraints.get("ca_p_max"),
        animal_key=animal_key,
    )

    # 若手动选择原料导致不可行，自动扩充
    if (
        result["status"] != "optimal"
        and selected_names is not None
        and len(ing_df) < len(all_ingredients)
    ):
        print("\n⚠️  所选原料不足以满足全部营养约束，自动从原料库补充...")
        # 重新加载配方要求（针对全量原料）
        ing_df_full = filter_ingredients(all_ingredients, None, energy_col, animal_key)
        formula_min2, formula_max2, ratio_constraints2, ca_p_constraints2 = load_formula_requirements(
            STANDARDS_FILE, ing_df_full, feed_type_dict,
        )
        # 合并约束：配方要求 + 用户手动设置
        expanded_min = dict(formula_min2)
        expanded_min.update(min_bounds)
        expanded_max = dict(formula_max2)
        expanded_max.update(max_bounds)
        # 重新计算精料/粗料索引
        conc_idx_full, forage_idx_full = _get_feed_type_indices(ing_df_full, feed_type_dict)

        result = build_and_solve(
            ingredients_df=ing_df_full,
            nutrient_requirements=nutrient_requirements,
            nutrient_map=nutrient_map,
            min_bounds=expanded_min if expanded_min else None,
            max_bounds=expanded_max if expanded_max else None,
            concentrate_indices=conc_idx_full if conc_idx_full else None,
            forage_indices=forage_idx_full if forage_idx_full else None,
            concentrate_min=ratio_constraints2.get("concentrate_min"),
            concentrate_max=ratio_constraints2.get("concentrate_max"),
            forage_min=ratio_constraints2.get("forage_min"),
            forage_max=ratio_constraints2.get("forage_max"),
            ca_p_min=ca_p_constraints2.get("ca_p_min"),
            ca_p_max=ca_p_constraints2.get("ca_p_max"),
            animal_key=animal_key,
        )

        if result["status"] == "optimal":
            original_set = set(ing_df.index)
            new_ingredients = [
                n for n in result["ingredients"] if n not in original_set
            ]
            print(
                f"  ✅ 自动补充了 {len(new_ingredients)} 种原料"
            )
            ing_df = ing_df_full  # 后续 Excel 输出使用全量数据
        else:
            print(f"  ❌ 即便使用全部原料也无法求解: {result.get('message')}")

    # 6. 输出结果
    print_result(result)

    # 7. 保存结果
    save = input("\n是否保存结果到Excel? [y/n]：").strip().lower()
    if save == "y":
        short_name = standard_name[:15].replace("/", "_").replace("\\", "_")
        output_path = os.path.join(BASE_DIR, f"配方结果_{short_name}.xlsx")
        save_result_to_excel(
            result, output_path,
            standard_name=standard_name,
            ingredients_df=ing_df,
            nutrient_map=nutrient_map,
        )
        print(f"  结果已保存至：{output_path}")


def _get_int_input(prompt: str, lo: int, hi: int) -> int:
    while True:
        try:
            val = int(input(f"  {prompt} ({lo}-{hi})：").strip())
            if lo <= val <= hi:
                return val
        except ValueError:
            pass
        print(f"  请输入 {lo} 到 {hi} 之间的整数")


def _manual_ingredient_selection(all_df: pd.DataFrame):
    """交互式选择原料及其用量上下限"""
    names = list(all_df.index)
    print(f"\n  可用原料共 {len(names)} 种，请输入要使用的原料名称（每行一个，空行结束）：")
    print("  （或直接回车跳过，使用全部）")

    selected = []
    while True:
        line = input("  > ").strip()
        if not line:
            break
        # 模糊匹配
        matches = [n for n in names if line in n]
        if len(matches) == 1:
            selected.append(matches[0])
            print(f"    ✅ 已选：{matches[0]}")
        elif len(matches) > 1:
            print(f"    找到多个匹配：")
            for i, m in enumerate(matches[:10]):
                print(f"      {i+1}. {m}")
            idx = _get_int_input("请选择序号", 1, min(len(matches), 10))
            selected.append(matches[idx - 1])
        else:
            print(f"    ⚠️  未找到：{line}")

    if not selected:
        return None, {}, {}

    print(f"\n  已选原料（{len(selected)} 种）：{', '.join(selected)}")

    # 设置用量限制
    min_bounds = {}
    max_bounds = {}
    set_bounds = input("\n  是否为部分原料设置用量上下限? [y/n]：").strip().lower()
    if set_bounds == "y":
        for name in selected:
            print(f"  {name}：")
            lo_str = input(f"    最小比例(0-1，回车=0)：").strip()
            hi_str = input(f"    最大比例(0-1，回车=1)：").strip()
            lo = float(lo_str) if lo_str else 0.0
            hi = float(hi_str) if hi_str else 1.0
            if lo > 0:
                min_bounds[name] = lo
            if hi < 1:
                max_bounds[name] = hi

    return selected, min_bounds, max_bounds


# ──────────────────────────── Excel输出 ─────────────────────────
def save_result_to_excel(result: dict,
                         output_path: str,
                         standard_name: str = "",
                         ingredients_df: pd.DataFrame | None = None,
                         nutrient_map: dict | None = None):
    """
    按"结果格式参考.xlsx"格式输出配方结果。
    表格结构：
      第0行  标题（含配方成本）
      第1行  表头：原料 / 配比% / 各营养指标 / 每t成本
      第2~n行 各原料行（配比 + 营养贡献值）
      合计行   各列求和
      标准行   饲养标准需要量
      与标准差  合计 - 标准
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "配方结果"

    # ── 样式 ──────────────────────────────────────────────────
    title_font   = Font(name="微软雅黑", bold=True, size=11)
    header_font  = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2E75B6")
    bold_font    = Font(name="微软雅黑", bold=True, size=10)
    normal_font  = Font(name="微软雅黑", size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right        = Alignment(horizontal="right",  vertical="center")
    thin         = Side(style="thin", color="AAAAAA")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(ws, r, c, value, font=None, fill=None,
              alignment=None, border_=None):
        cell = ws.cell(row=r, column=c, value=value)
        cell.font     = font  or normal_font
        cell.alignment = alignment or center
        cell.border   = border_ or border
        if fill:
            cell.fill = fill
        return cell

    # ── 营养指标列顺序 ────────────────────────────────────────
    # 仅导出饲养标准中实际要求的营养指标（从 result["requirements"] 获取）
    requirements = result.get("requirements", {})
    if requirements:
        # 过滤掉值为 None 的项
        nutrient_cols = [k for k, v in requirements.items() if v is not None]
    elif nutrient_map:
        nutrient_cols = list(nutrient_map.keys())
    else:
        nutrient_cols = list(NUTRIENT_MAP.keys())

    # 把能量列提到第一位（在"配比%"之后显示）
    # 先检查 nutrient_cols 中是否已有能量指标
    energy_col_name = None
    for k, v in (nutrient_map or NUTRIENT_MAP).items():
        if v and "能" in str(v):
            energy_col_name = k
            break
    if energy_col_name and energy_col_name in nutrient_cols:
        # 已在列表中 -> 移到首位
        nutrient_cols.remove(energy_col_name)
        nutrient_cols = [energy_col_name] + nutrient_cols

    # ── 构建完整表头 ─────────────────────────────────────────
    headers = ["原料", "配比%"] + nutrient_cols + ["每t成本/元"]

    # ── 第0行：标题 ──────────────────────────────────────────
    cost_yuan_kg = result.get("cost", 0)
    title = f"全价饲料配方({cost_yuan_kg:.2f}元/kg）"
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c = _cell(ws, 1, 1, title, font=title_font, alignment=center)
    ws.row_dimensions[1].height = 25

    # ── 第1行：表头 ──────────────────────────────────────────
    ws.row_dimensions[2].height = 35
    for ci, hdr in enumerate(headers, 1):
        c = _cell(ws, 2, ci, hdr,
                   font=header_font, fill=header_fill, alignment=center)
        c.border = border

    # ── 辅助：查原料的营养值 ────────────────────────────────
    ing = ingredients_df if ingredients_df is not None else None

    def _nutrient_contribution(ing_name, nutr_key):
        """返回 配比% × 原料营养值（单利贡献）"""
        if ing is None or ing_name not in ing.index:
            return 0.0
        ing_col = (nutrient_map or {}).get(nutr_key)
        if not ing_col or ing_col not in ing.columns:
            # 尝试直接用 NUTRIENT_MAP
            ing_col = NUTRIENT_MAP.get(nutr_key)
        if not ing_col or ing_col not in ing.columns:
            return 0.0
        val = pd.to_numeric(ing.loc[ing_name, ing_col], errors="coerce")
        if pd.isna(val):
            return 0.0
        pct = result["ingredients"].get(ing_name, 0) / 100.0  # 转为比例
        return val * pct

    def _ingredient_cost(ing_name):
        """返回该原料在配方中每吨的贡献成本（元/t）"""
        if ing is None or ing_name not in ing.index:
            return 0.0
        price = pd.to_numeric(ing.loc[ing_name, "价格"], errors="coerce")
        if pd.isna(price):
            return 0.0
        pct = result["ingredients"].get(ing_name, 0) / 100.0
        return price * pct * 1000  # 元/t

    # ── 原料数据行 ────────────────────────────────────────────
    row_idx = 3
    for ing_name, pct in sorted(result["ingredients"].items(),
                                 key=lambda x: -x[1]):
        ws.row_dimensions[row_idx].height = 18
        # 原料名
        _cell(ws, row_idx, 1, ing_name, font=normal_font,
              alignment=Alignment(horizontal="left", vertical="center"))
        # 配比%
        _cell(ws, row_idx, 2, round(pct, 2),
              alignment=right)
        # 各营养贡献
        for ci, nk in enumerate(nutrient_cols, 3):
            contrib = _nutrient_contribution(ing_name, nk)
            _cell(ws, row_idx, ci, round(contrib, 4),
                  alignment=right)
        # 每吨成本贡献
        cost_contrib = _ingredient_cost(ing_name)
        _cell(ws, row_idx, len(headers), round(cost_contrib, 2),
              alignment=right)
        row_idx += 1

    # ── 合计行 ───────────────────────────────────────────────
    ws.row_dimensions[row_idx].height = 18
    _cell(ws, row_idx, 1, "合计", font=bold_font,
          alignment=center)
    # 配比% 合计
    total_pct = sum(result["ingredients"].values())
    _cell(ws, row_idx, 2, round(total_pct, 2),
          font=bold_font, alignment=right)
    # 营养合计（实际含量）
    for ci, nk in enumerate(nutrient_cols, 3):
        actual = result["nutrients"].get(nk, 0)
        _cell(ws, row_idx, ci, round(actual, 4),
              font=bold_font, alignment=right)
    # 总成本
    total_cost_per_t = cost_yuan_kg * 1000
    _cell(ws, row_idx, len(headers), round(total_cost_per_t, 2),
          font=bold_font, alignment=right)
    row_idx += 1

    # ── 标准行 ───────────────────────────────────────────────
    ws.row_dimensions[row_idx].height = 18
    _cell(ws, row_idx, 1, "标准", font=bold_font,
          alignment=center)
    _cell(ws, row_idx, 2, "", alignment=center)  # 配比% 无标准
    reqs = result.get("requirements", {})
    for ci, nk in enumerate(nutrient_cols, 3):
        req_val = reqs.get(nk)
        if req_val is not None:
            _cell(ws, row_idx, ci, round(float(req_val), 4),
                  font=bold_font, alignment=right)
        else:
            _cell(ws, row_idx, ci, "", alignment=center)
    _cell(ws, row_idx, len(headers), "", alignment=center)
    row_idx += 1

    # ── 与标准差行 ──────────────────────────────────────────
    ws.row_dimensions[row_idx].height = 18
    _cell(ws, row_idx, 1, "与标准的差", font=bold_font,
          alignment=center)
    _cell(ws, row_idx, 2, "", alignment=center)
    for ci, nk in enumerate(nutrient_cols, 3):
        actual = result["nutrients"].get(nk, 0)
        req_val = reqs.get(nk)
        if req_val is not None:
            diff = actual - float(req_val)
            _cell(ws, row_idx, ci, round(diff, 4),
                  font=bold_font,
                  alignment=right)
        else:
            _cell(ws, row_idx, ci, "", alignment=center)
    _cell(ws, row_idx, len(headers), "", alignment=center)

    # ── 列宽 ─────────────────────────────────────────────────
    col_widths = [22, 8] + [12] * len(nutrient_cols) + [14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)


# ─────────────────────── 非交互式API接口 ─────────────────────────
def formulate(
    selected_ingredients: list[str] | None = None,
    min_bounds: dict | None = None,
    max_bounds: dict | None = None,
    standard_index: int = 0,        # 选择第几个标准（0=第一个）
) -> dict:
    """
    程序化调用接口（供其他模块导入使用）。
    直接读取饲养标准文件，自动推断动物种类和能量指标。
    一律使用干物质基础(DM)计算。

    参数:
        selected_ingredients: 指定使用的原料名称列表（None=全部）
        min_bounds: {原料名: 最小比例}
        max_bounds: {原料名: 最大比例}
        standard_index: 选择第几个标准（0=第一个，1=第二个，...）

    返回:
        与 build_and_solve() 返回格式相同的字典
    """
    all_ingredients = load_ingredients(INGREDIENTS_FILE)
    nutrient_requirements, standard_name = load_standards(STANDARDS_FILE, standard_index=standard_index)

    if not nutrient_requirements:
        return {"status": "failed", "message": "饲养标准为空"}
    
    # 强制使用干物质基础(DM）计算
    print(f"  ℹ️ 按干物质基础(DM)计算配方（标准: {standard_name}）")
    all_ingredients = convert_to_dm_basis(all_ingredients)

    # 只保留有映射的营养指标（动态匹配：标准指标 → 原料库列名）
    animal_key = "禽"
    if any(kw in standard_name for kw in ["猪", "仔猪", "母猪"]):
        animal_key = "猪"
    elif any(kw in standard_name for kw in ["牛", "奶牛", "肉牛"]):
        animal_key = "牛"
    elif any(kw in standard_name for kw in ["羊"]):
        animal_key = "羊"
    elif any(kw in standard_name for kw in ["鸡", "鸭", "鹅", "禽", "肉鸡", "蛋鸡"]):
        animal_key = "禽"
    elif any(kw in standard_name for kw in ["猫", "狗", "马", "鱼", "其他"]):
        animal_key = "其他"
    energy_col = ENERGY_MAP.get(animal_key, "禽代谢能MC/Kg")

    # 先过滤原料（确定能量列后）
    ing_df_pre = filter_ingredients(all_ingredients, selected_ingredients, energy_col, animal_key)
    ing_cols = list(ing_df_pre.columns)

    # 动态匹配标准指标到原料列
    std_keys = list(nutrient_requirements.keys())
    matched_map, unmatched_keys = _find_best_column_matches(
        std_keys, ing_cols, animal_key=animal_key
    )

    # 构建 nutrient_map 和 deduped_mapped
    nutrient_map = matched_map.copy()
    deduped_mapped = {}
    seen_cols_f = set()
    for k, v in nutrient_requirements.items():
        if k in matched_map and v is not None:
            effective_col = matched_map[k]
            if effective_col in seen_cols_f:
                continue
            seen_cols_f.add(effective_col)
            deduped_mapped[k] = v

    if unmatched_keys:
        print(f"  ⚠️ {len(unmatched_keys)} 个指标未匹配到原料列，已跳过: {unmatched_keys}")
    nutrient_requirements = deduped_mapped
    print(f"  能量指标：    {energy_col}（自动识别 → {animal_key}）")
    print(f"  营养约束指标数（动态匹配后）: {len(nutrient_requirements)}")
    # 使用已过滤的原料df
    ing_df = ing_df_pre

    # 加载饲料类型分类
    feed_type_dict = load_feed_type_classification()
    conc_idx, forage_idx = _get_feed_type_indices(ing_df, feed_type_dict)

    # 自动加载配方要求（包括精粗比约束+钙磷比约束）
    formula_min, formula_max, ratio_constraints, ca_p_constraints = load_formula_requirements(
        STANDARDS_FILE, ing_df, feed_type_dict,
    )
    merged_min = dict(formula_min)
    if min_bounds:
        merged_min.update(min_bounds)
    merged_max = dict(formula_max)
    if max_bounds:
        merged_max.update(max_bounds)

    result = build_and_solve(
        ingredients_df=ing_df,
        nutrient_requirements=nutrient_requirements,
        nutrient_map=nutrient_map,
        min_bounds=merged_min if merged_min else None,
        max_bounds=merged_max if merged_max else None,
        concentrate_indices=conc_idx if conc_idx else None,
        forage_indices=forage_idx if forage_idx else None,
        concentrate_min=ratio_constraints.get("concentrate_min"),
        concentrate_max=ratio_constraints.get("concentrate_max"),
        forage_min=ratio_constraints.get("forage_min"),
        forage_max=ratio_constraints.get("forage_max"),
        ca_p_min=ca_p_constraints.get("ca_p_min"),
        ca_p_max=ca_p_constraints.get("ca_p_max"),
        animal_key=animal_key,
    )

    # 若手动指定原料但不可行，自动扩充到全量原料
    if (
        result["status"] != "optimal"
        and selected_ingredients is not None
        and len(ing_df) < len(all_ingredients)
    ):
        print("\n⚠️  所选原料不足以满足全部营养约束，自动从原料库补充...")
        ing_df_full = filter_ingredients(all_ingredients, None, energy_col, animal_key)
        expanded_min = dict(formula_min2)
        if min_bounds:
            expanded_min.update(min_bounds)
        expanded_max = dict(formula_max2)
        if max_bounds:
            expanded_max.update(max_bounds)
        # 重新计算索引
        conc_idx_full, forage_idx_full = _get_feed_type_indices(ing_df_full, feed_type_dict)

        result = build_and_solve(
            ingredients_df=ing_df_full,
            nutrient_requirements=nutrient_requirements,
            nutrient_map=nutrient_map,
            min_bounds=expanded_min if expanded_min else None,
            max_bounds=expanded_max if expanded_max else None,
            concentrate_indices=conc_idx_full if conc_idx_full else None,
            forage_indices=forage_idx_full if forage_idx_full else None,
            concentrate_min=ratio_constraints2.get("concentrate_min"),
            concentrate_max=ratio_constraints2.get("concentrate_max"),
            forage_min=ratio_constraints2.get("forage_min"),
            forage_max=ratio_constraints2.get("forage_max"),
            ca_p_min=ca_p_constraints2.get("ca_p_min"),
            ca_p_max=ca_p_constraints2.get("ca_p_max"),
            animal_key=animal_key,
        )
        if result["status"] == "optimal":
            original_set = set(ing_df.index)
            new_ingredients = [
                n for n in result["ingredients"] if n not in original_set
            ]
            print(f"  ✅ 自动补充了 {len(new_ingredients)} 种原料")

    # 标记是否按干物质基础计算（现在强制使用DM基础）
    result["dm_basis"] = True
    return result


# ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    interactive_mode()
